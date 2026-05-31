"""
╔══════════════════════════════════════════════════════════════════════╗
║         DASHBOARD DE ANÁLISIS FINANCIERO — v4.0                     ║
╚══════════════════════════════════════════════════════════════════════╝
Correr con:  python -m streamlit run dashboard_v2.py
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from scipy.optimize import minimize
import yfinance as yf
from datetime import datetime
import io
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(page_title="Dashboard Financiero", page_icon="📈",
                   layout="wide", initial_sidebar_state="collapsed")

# ─── ESTILOS ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:IBM Plex Sans,sans-serif;background:#080C14;color:#C8D4E3}
.stApp{background:#080C14}
.main-header{background:linear-gradient(135deg,#0A1628,#0D1F3C,#091520);border-bottom:1px solid #1A3A5C;
  padding:1.4rem 2rem 1.1rem;margin:-1rem -1rem 1.5rem;display:flex;align-items:center;justify-content:space-between}
.main-header h1{font-family:IBM Plex Mono,monospace;font-size:1.05rem;font-weight:600;color:#E8F4FD;
  letter-spacing:.15em;margin:0;text-transform:uppercase}
.main-header .subtitle{font-size:.7rem;color:#4A7FA5;letter-spacing:.08em;margin-top:.2rem}
.header-badge{background:#0E3460;border:1px solid #1A5C9A;border-radius:4px;padding:.3rem .8rem;
  font-family:IBM Plex Mono,monospace;font-size:.68rem;color:#4AB3F4;letter-spacing:.1em}
.stTabs [data-baseweb="tab-list"]{background:#0A1628;border-bottom:1px solid #1A3A5C;gap:0;padding:0 1rem}
.stTabs [data-baseweb="tab"]{font-family:IBM Plex Mono,monospace;font-size:.7rem;font-weight:500;
  letter-spacing:.1em;color:#4A7FA5;padding:.8rem 1.4rem;border-bottom:2px solid transparent;
  text-transform:uppercase;background:transparent!important}
.stTabs [aria-selected="true"]{color:#4AB3F4!important;border-bottom:2px solid #4AB3F4!important;background:transparent!important}
.stTabs [data-baseweb="tab-panel"]{padding:1.5rem .5rem;background:transparent}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:.75rem;margin-bottom:1.5rem}
.kpi-card{background:#0D1E35;border:1px solid #1A3A5C;border-radius:6px;padding:.9rem 1.1rem;position:relative;overflow:hidden}
.kpi-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,#1A5C9A,#4AB3F4)}
.kpi-label{font-family:IBM Plex Mono,monospace;font-size:.6rem;color:#4A7FA5;letter-spacing:.12em;text-transform:uppercase;margin-bottom:.35rem}
.kpi-value{font-family:IBM Plex Mono,monospace;font-size:1.3rem;font-weight:600;color:#E8F4FD;line-height:1}
.kpi-value.positive{color:#2ECC71}.kpi-value.negative{color:#E74C3C}
.kpi-delta{font-size:.62rem;margin-top:.25rem;color:#4A7FA5}
.section-header{font-family:IBM Plex Mono,monospace;font-size:.68rem;letter-spacing:.18em;
  text-transform:uppercase;color:#4AB3F4;border-left:3px solid #4AB3F4;padding-left:.8rem;margin:1.5rem 0 .8rem}
.styled-table{width:100%;border-collapse:collapse;font-family:IBM Plex Mono,monospace;
  font-size:.76rem;background:#0D1E35;border:1px solid #1A3A5C;border-radius:6px;overflow:hidden}
.styled-table th{background:#0A1628;color:#4AB3F4;font-size:.63rem;letter-spacing:.1em;
  text-transform:uppercase;padding:.65rem .9rem;border-bottom:1px solid #1A3A5C;text-align:right}
.styled-table th:first-child{text-align:left}
.styled-table td{padding:.55rem .9rem;border-bottom:1px solid #111E30;color:#C8D4E3;text-align:right}
.styled-table td:first-child{text-align:left;color:#7FA8C9}
.styled-table tr:last-child td{border-bottom:none}
.styled-table tr:hover td{background:#0F2540}
.td-positive{color:#2ECC71!important;font-weight:600}.td-negative{color:#E74C3C!important;font-weight:600}
.td-neutral{color:#F39C12!important;font-weight:600}.td-blue{color:#4AB3F4!important;font-weight:600}
.stNumberInput input,.stTextInput input{background:#0D1E35!important;border:1px solid #1A3A5C!important;
  color:#C8D4E3!important;border-radius:4px!important;font-family:IBM Plex Mono,monospace!important;font-size:.8rem!important}
.stNumberInput label,.stTextInput label,.stSelectbox label,.stMultiSelect label{
  font-family:IBM Plex Mono,monospace!important;font-size:.66rem!important;
  color:#4A7FA5!important;letter-spacing:.1em!important;text-transform:uppercase!important}
.stButton>button{background:#0E3460!important;border:1px solid #1A5C9A!important;color:#4AB3F4!important;
  font-family:IBM Plex Mono,monospace!important;font-size:.7rem!important;letter-spacing:.1em!important;
  text-transform:uppercase!important;border-radius:4px!important;padding:.5rem 1.2rem!important;transition:all .2s!important}
.stButton>button:hover{background:#1A5C9A!important;border-color:#4AB3F4!important;color:#E8F4FD!important}
.streamlit-expanderHeader{background:#0D1E35!important;border:1px solid #1A3A5C!important;border-radius:4px!important;
  font-family:IBM Plex Mono,monospace!important;font-size:.7rem!important;color:#4AB3F4!important;letter-spacing:.1em!important}
.model-badge{display:inline-block;background:#0E3460;border:1px solid #1A5C9A;border-radius:4px;
  padding:.2rem .7rem;font-family:IBM Plex Mono,monospace;font-size:.65rem;color:#4AB3F4;
  letter-spacing:.12em;text-transform:uppercase;margin-bottom:.5rem}
hr{border-color:#1A3A5C!important}
#MainMenu,footer,header{visibility:hidden}
.block-container{padding-top:.5rem!important}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
  <div><h1>📈 Dashboard de Análisis Financiero</h1>
    <div class="subtitle">Análisis Básico · Portafolios · Val. Estadística · Técnica · Fundamental</div></div>
  <div class="header-badge">v4.0 · 2026</div>
</div>""", unsafe_allow_html=True)

PLOT_BG="#080C14"; PAPER_BG="#0A1628"; GRID_COLOR="#1A3A5C"; FONT_COLOR="#C8D4E3"
PALETTE=["#4AB3F4","#2ECC71","#F39C12","#E74C3C","#9B59B6","#1ABC9C","#E67E22","#3498DB","#EC407A","#26C6DA"]

def base_layout(title="",height=420):
    return dict(
        title=dict(text=title,font=dict(family="IBM Plex Mono",size=11,color="#7FA8C9"),x=0.01),
        plot_bgcolor=PLOT_BG,paper_bgcolor=PAPER_BG,
        font=dict(family="IBM Plex Sans",size=11,color=FONT_COLOR),
        height=height,margin=dict(l=50,r=30,t=45,b=40),
        xaxis=dict(gridcolor=GRID_COLOR,showgrid=True,zeroline=False,tickfont=dict(size=10)),
        yaxis=dict(gridcolor=GRID_COLOR,showgrid=True,zeroline=False,tickfont=dict(size=10)),
        legend=dict(bgcolor="rgba(10,22,40,0.85)",bordercolor=GRID_COLOR,borderwidth=1,font=dict(size=11,color="#E8F4FD")),
        hoverlabel=dict(bgcolor="#0D1E35",bordercolor="#1A5C9A",font=dict(family="IBM Plex Mono",size=11)),
    )

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def fmt_fecha(ts):
    """Formatea cualquier fecha a dd/mm/aa."""
    try:
        return pd.Timestamp(ts).strftime("%d/%m/%y")
    except:
        return str(ts)

def fecha_anterior_cercana(idx, fecha):
    """
    Busca en idx la fecha más cercana <= fecha.
    Si ninguna cumple, retorna la primera disponible.
    """
    candidates = idx[idx <= fecha]
    if len(candidates) > 0:
        return candidates[-1]
    return idx[0]

def fmt_num(val, decimales=2, es_pct=False, es_usd=False):
    """Formatea número con coma decimal y punto miles (estilo colombiano/europeo)."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "N/A"
    if es_pct:
        s = f"{val*100:,.{decimales}f}"
    elif es_usd:
        s = f"{val:,.{decimales}f}"
    else:
        s = f"{val:,.{decimales}f}"
    # Convertir: punto miles → X, coma decimal → coma, X → punto
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    if es_pct:
        s += "%"
    elif es_usd:
        s = "$" + s
    return s

def periodo_legible(dias):
    if dias >= 365:
        a = dias/365
        return f"{int(round(a))} año{'s' if round(a)!=1 else ''}" if abs(a-round(a))<0.15 else f"{a:.1f} años"
    m = dias/30.44
    return f"{int(round(m))} mes{'es' if round(m)!=1 else ''}" if abs(m-round(m))<0.3 else f"{m:.1f} meses"

# ─── CARGA DE DATOS ───────────────────────────────────────────────────────────
@st.cache_data(ttl=3600)
def cargar_datos(tickers_str, benchmark, fecha_ini_str, fecha_fin_str):
    """
    fecha_ini_str / fecha_fin_str: 'YYYY-MM-DD' o '' para histórico completo.
    Lógica:
      - Descarga desde el inicio máximo posible
      - Recorta al inicio de la acción más joven
      - Aplica ffill para huecos intermedios
      - Respeta fechas inclusivas (busca anterior cercana si no existe)
    """
    tickers = [t.strip().upper() for t in tickers_str.split(",") if t.strip()]
    bench_t = benchmark.strip().upper()
    todos   = list(set(tickers + [bench_t]))

    # Siempre descargamos desde el máximo histórico disponible
    start_dl = "1990-01-01"
    end_dl   = datetime.today().strftime("%Y-%m-%d")
    raw = yf.download(todos, start=start_dl, end=end_dl, auto_adjust=True, progress=False)
    if raw.empty:
        return None, tickers

    precios = raw["Close"]
    if isinstance(precios, pd.Series):
        precios = precios.to_frame()
    precios.index = pd.to_datetime(precios.index)

    # ffill: rellena huecos intermedios con precio anterior
    precios = precios.ffill()

    # Inicio: la acción más joven define el inicio común
    tickers_validos = [t for t in tickers if t in precios.columns]
    if tickers_validos:
        primer_dato = max(precios[t].first_valid_index() for t in tickers_validos)
        precios = precios.loc[primer_dato:]

    # Aplicar rango solicitado por el usuario (inclusivo, anterior cercana)
    idx = precios.index
    if fecha_ini_str:
        fi_req = pd.Timestamp(fecha_ini_str)
        fi_use = fecha_anterior_cercana(idx, fi_req)
        precios = precios.loc[fi_use:]

    if fecha_fin_str:
        ff_req = pd.Timestamp(fecha_fin_str)
        ff_use = fecha_anterior_cercana(precios.index, ff_req)
        precios = precios.loc[:ff_use]

    precios = precios.dropna(how="all")
    return precios, tickers

@st.cache_data(ttl=3600)
def obtener_rf(ticker_rf="^TNX"):
    try:
        d = yf.download(ticker_rf, period="5d", auto_adjust=True, progress=False)
        if not d.empty:
            v = float(d["Close"].dropna().iloc[-1])
            return v/100 if v > 1 else v
    except:
        pass
    return 0.045

# ─── CÁLCULOS BÁSICOS ─────────────────────────────────────────────────────────
def calc_rendimientos(precios):
    return precios.pct_change().dropna()

def base_100(serie):
    return serie / serie.iloc[0] * 100

def rend_nominal(precios, fi, ff):
    """Rendimiento entre fi y ff usando fecha anterior cercana."""
    idx = precios.index
    fi_r = fecha_anterior_cercana(idx, fi)
    ff_r = fecha_anterior_cercana(idx, ff)
    return (precios.loc[ff_r] / precios.loc[fi_r]) - 1

def rend_efectivo(rn, dias):
    dias = max(dias, 1)
    return (1 + rn) ** (365/dias) - 1

def vol_nominal(rend, fi, ff):
    mask = (rend.index >= fi) & (rend.index <= ff)
    return rend.loc[mask].std()

def sharpe_ratio(re, ve, rf=0.0):
    """Para análisis básico: sin RF. Para portafolio se calcula directo."""
    return re / ve.replace(0, np.nan)

def pct_espectro(rend, niveles=[0,1,5,25,50,75,95,99,100]):
    return pd.DataFrame(
        {col: np.percentile(rend[col].dropna(), niveles) for col in rend.columns},
        index=[f"{p}%" for p in niveles]
    )

def pct_actual_serie(rend, ult_rend):
    return pd.Series({
        col: round((np.sum(rend[col].dropna().values <= ult_rend[col]) / len(rend[col].dropna())) * 100, 1)
        for col in rend.columns
    })

# ─── PORTAFOLIO ───────────────────────────────────────────────────────────────
def retorno_markowitz(rend_df): return rend_df.mean()*365
def retorno_capm(rend_df, rend_bench, rf):
    rm=rend_bench.mean()*365; betas={}
    for col in rend_df.columns:
        al=pd.concat([rend_df[col],rend_bench],axis=1).dropna()
        betas[col]=np.cov(al.iloc[:,0],al.iloc[:,1])[0,1]/np.var(al.iloc[:,1]) if len(al)>10 else 1.
    bs=pd.Series(betas); return rf+bs*(rm-rf), bs
def retorno_montecarlo(rend_df, n_sim=1000):
    return pd.Series({col: np.percentile(
        [np.mean(np.random.choice(rend_df[col].dropna().values,365,replace=True))*365 for _ in range(n_sim)],50)
        for col in rend_df.columns})
def vol_individual(rend_df): return rend_df.std()*np.sqrt(365)
def cov_anualizada(rend_df): return rend_df.cov()*365
def vol_portafolio(pesos,cov_anual):
    w=np.array(pesos); return np.sqrt(max(w@cov_anual.values@w,0))
def calcular_betas(rend_df,rend_bench):
    betas={}
    for col in rend_df.columns:
        al=pd.concat([rend_df[col],rend_bench],axis=1).dropna()
        betas[col]=round(np.cov(al.iloc[:,0],al.iloc[:,1])[0,1]/np.var(al.iloc[:,1]) if len(al)>10 else 1.,4)
    return pd.Series(betas)
def tracking_error(rp,rb): return (rp-rb).dropna().std()*np.sqrt(365)
def information_ratio(rp,rb,te): return (rp-rb)/te if te and te!=0 else np.nan
def optimizar_portafolio(retornos,cov_anual,rf,objetivo="sharpe"):
    n=len(retornos); w0=np.ones(n)/n; bounds=[(0.,1.)]*n
    cons=[{"type":"eq","fun":lambda w:np.sum(w)-1.}]
    fns={"sharpe":  lambda w:-(np.dot(w,retornos)-rf)/max(vol_portafolio(w,cov_anual),1e-9),
         "retorno": lambda w:-np.dot(w,retornos),
         "varianza":lambda w:vol_portafolio(w,cov_anual)**2,
         "ir":      lambda w:-(np.dot(w,retornos)-retornos.mean())/max(np.sqrt(np.sum((w-1/n)**2))*np.sqrt(365),1e-9)}
    res=minimize(fns.get(objetivo,fns["sharpe"]),w0,method="SLSQP",bounds=bounds,constraints=cons,
                 options={"maxiter":1000,"ftol":1e-9})
    w=np.clip(res.x if res.success else w0,0,1); return w/w.sum()

# ─── RENDER TABLAS HTML ───────────────────────────────────────────────────────
def fmt_val(val,fmt=".2%",umbral=0.):
    if isinstance(val,(int,float)) and not np.isnan(val):
        css="td-positive" if val>umbral else ("td-negative" if val<0 else "")
        return f"{val:{fmt}}",css
    return "N/A","td-neutral"

def tabla_html(df, fmt=".2%", umbral=0., titulo="", col_headers=None, row_headers=None):
    """
    df: activos en FILAS, rangos en COLUMNAS.
    col_headers: dict {col_name: (fecha_ini, fecha_fin)} para mostrar fechas bajo el header.
    row_headers: lista de nombres de filas (activos).
    """
    cols = list(df.columns)
    # Encabezados de columna con fechas opcionales
    hdr = ""
    for c in cols:
        fecha_str = ""
        if col_headers and c in col_headers:
            fi_s,ff_s = col_headers[c]
            fecha_str = f'<span style="font-size:.7rem;color:#7FA8C9;display:block;font-weight:400;margin-top:3px;">{fi_s} → {ff_s}</span>'
        hdr += f"<th style='text-align:right'>{c}{fecha_str}</th>"

    rows = ""
    for idx,row in df.iterrows():
        celdas = "".join(
            f'<td class="{fmt_val(v,fmt,umbral)[1]}">{fmt_val(v,fmt,umbral)[0]}</td>'
            for v in row)
        rows += f"<tr><td>{idx}</td>{celdas}</tr>"

    t = (f'<p style="font-family:\'IBM Plex Mono\',monospace;font-size:.63rem;'
         f'color:#4AB3F4;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.35rem;">'
         f'{titulo}</p>') if titulo else ""
    return (f"{t}<table class=\'styled-table\'>"
            f"<thead><tr><th>Activo</th>{hdr}</tr></thead>"
            f"<tbody>{rows}</tbody></table>")

# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────
def exportar_excel_bloque1(precios, rend, df_rn, df_re, df_vn, df_ve, df_sh,
                            pct_esp, pct_ult, ult_rend, tks_ok, bench, fechas_rangos,
                            port_data=None, ve_data=None, vt_data=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                      numbers as xl_numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        st.error("Instala openpyxl: pip install openpyxl")
        return None

    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja default

    # Colores
    AZUL_OSC  = "0D1E35"; AZUL_MED  = "1A3A5C"; AZUL_ACE  = "4AB3F4"
    VERDE     = "2ECC71"; ROJO      = "E74C3C"; GRIS_FILA = "F0F4F8"
    BLANCO    = "FFFFFF"; TEXTO_OSC = "0A1628"

    fill_header  = PatternFill("solid", fgColor=AZUL_OSC)
    fill_subhdr  = PatternFill("solid", fgColor=AZUL_MED)
    fill_title   = PatternFill("solid", fgColor=AZUL_ACE)
    fill_alt     = PatternFill("solid", fgColor=GRIS_FILA)
    fill_pos     = PatternFill("solid", fgColor="D5F5E3")
    fill_neg     = PatternFill("solid", fgColor="FADBD8")

    font_title   = Font(name="Calibri", bold=True, size=14, color=TEXTO_OSC)
    font_sub     = Font(name="Calibri", bold=True, size=9,  color=BLANCO)
    font_col     = Font(name="Calibri", bold=True, size=9,  color=AZUL_ACE)
    font_data    = Font(name="Calibri", size=9,    color=TEXTO_OSC)
    font_activo  = Font(name="Calibri", bold=True, size=9,  color=BLANCO)

    border_thin  = Border(
        left=Side(style="thin",color=AZUL_MED), right=Side(style="thin",color=AZUL_MED),
        top=Side(style="thin",color=AZUL_MED),  bottom=Side(style="thin",color=AZUL_MED))
    align_c      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_r      = Alignment(horizontal="right",  vertical="center")
    align_l      = Alignment(horizontal="left",   vertical="center")

    gen_date = datetime.today().strftime("%d/%m/%Y %H:%M")

    def write_title_block(ws, title, subtitle=""):
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 10
        c1 = ws.cell(1,1, f"ANÁLISIS BÁSICO — {title.upper()}")
        c1.font = font_title; c1.fill = fill_title; c1.alignment = align_l
        c2 = ws.cell(2,1, f"Dashboard Financiero  |  Generado: {gen_date}{('  |  '+subtitle) if subtitle else ''}")
        c2.font = Font(name="Calibri",size=8,color=BLANCO,italic=True)
        c2.fill = fill_header; c2.alignment = align_l
        # Fila 3 separadora
        ws.cell(3,1,"").fill = PatternFill("solid",fgColor=AZUL_ACE)

    def style_col_header(cell):
        cell.font = font_col; cell.fill = fill_subhdr
        cell.alignment = align_c; cell.border = border_thin

    def style_row_label(cell):
        cell.font = font_activo; cell.fill = fill_subhdr
        cell.alignment = align_l; cell.border = border_thin
        cell.font = Font(name="Calibri", bold=True, size=9, color=BLANCO)

    def style_data(cell, row_idx, val=None, fmt_type="pct"):
        cell.font = font_data; cell.border = border_thin
        cell.alignment = align_r
        cell.fill = fill_alt if row_idx % 2 == 0 else PatternFill("solid",fgColor=BLANCO)
        # Formato de número
        if fmt_type == "pct"  : cell.number_format = '0.00%'
        elif fmt_type == "usd": cell.number_format = '"$"#,##0.00'
        elif fmt_type == "dec": cell.number_format = '0.00'
        elif fmt_type == "dec3":cell.number_format = '0.000'
        elif fmt_type == "fecha": cell.number_format = 'DD/MM/YYYY'
        # Color condicional
        if val is not None and isinstance(val,(int,float)) and not np.isnan(val):
            if fmt_type in ("pct","dec","dec3") and val > 0: cell.fill = fill_pos
            elif fmt_type in ("pct","dec","dec3") and val < 0: cell.fill = fill_neg

    def autofit(ws, min_w=10, max_w=30):
        for col in ws.columns:
            length = max((len(str(c.value or "")) for c in col), default=min_w)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length+2, min_w), max_w)

    def merge_title(ws, ncols):
        for row in [1,2,3]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols,2))

    # ── 1. HOJA PRECIOS ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Precios")
    cols_p = tks_ok + ([bench] if bench in precios.columns else [])
    ncols = len(cols_p)+1
    merge_title(ws, ncols); write_title_block(ws,"Precios de Cierre Ajustados")
    r=4
    ws.cell(r,1,"Fecha").font=font_col; ws.cell(r,1).fill=fill_subhdr; ws.cell(r,1).alignment=align_c; ws.cell(r,1).border=border_thin
    for ci,t in enumerate(cols_p,2):
        style_col_header(ws.cell(r,ci,t))
    for ri,(idx,row) in enumerate(precios[cols_p].iterrows(),1):
        r+=1
        c=ws.cell(r,1,idx.to_pydatetime())
        style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
        for ci,t in enumerate(cols_p,2):
            v=row[t] if t in row else None
            if pd.notna(v): style_data(ws.cell(r,ci,float(v)),ri,float(v),"usd")
    autofit(ws)

    # ── 2. HOJA RENDIMIENTOS ───────────────────────────────────────────────────
    # Rendimientos incluye benchmark
    cols_rend_xl = tks_ok + ([bench] if bench in precios.columns else [])
    rend_full_xl = calc_rendimientos(precios[cols_rend_xl])
    ncols_rend = len(cols_rend_xl)+1
    ws2 = wb.create_sheet("Rendimientos")
    merge_title(ws2,ncols_rend); write_title_block(ws2,"Rendimientos Diarios")
    r=4
    ws2.cell(r,1,"Fecha").font=font_col; ws2.cell(r,1).fill=fill_subhdr; ws2.cell(r,1).alignment=align_c; ws2.cell(r,1).border=border_thin
    for ci,t in enumerate(cols_rend_xl,2): style_col_header(ws2.cell(r,ci,t))
    for ri,(idx,row) in enumerate(rend_full_xl.iterrows(),1):
        r+=1
        c=ws2.cell(r,1,idx.to_pydatetime()); style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
        for ci,t in enumerate(cols_rend_xl,2):
            v=row[t] if t in row else None
            if pd.notna(v): style_data(ws2.cell(r,ci,float(v)),ri,float(v),"pct")
    autofit(ws2)

    # ── 3. HOJA BASE 100 ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Base 100")
    merge_title(ws3,ncols); write_title_block(ws3,"Evolución Base 100")
    r=4
    ws3.cell(r,1,"Fecha").font=font_col; ws3.cell(r,1).fill=fill_subhdr; ws3.cell(r,1).alignment=align_c; ws3.cell(r,1).border=border_thin
    for ci,t in enumerate(tks_ok,2): style_col_header(ws3.cell(r,ci,t))
    b100_df = pd.DataFrame({t: base_100(precios[t].dropna()) for t in tks_ok})
    for ri,(idx,row) in enumerate(b100_df.iterrows(),1):
        r+=1
        c=ws3.cell(r,1,idx.to_pydatetime()); style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
        for ci,t in enumerate(tks_ok,2):
            v=row[t] if t in row.index else None
            if pd.notna(v): style_data(ws3.cell(r,ci,float(v)),ri,fmt_type="dec")
    autofit(ws3)

    # ── 3b. HOJA PRECIOS LN ──────────────────────────────────────────────────
    cols_ln = tks_ok + ([bench] if bench in precios.columns else [])
    ws3b = wb.create_sheet("Precios LN")
    merge_title(ws3b, len(cols_ln)+1); write_title_block(ws3b,"Precios Logarítmicos LN(Precio)")
    r=4
    ws3b.cell(r,1,"Fecha").font=font_col; ws3b.cell(r,1).fill=fill_subhdr
    ws3b.cell(r,1).alignment=align_c; ws3b.cell(r,1).border=border_thin
    for ci,t in enumerate(cols_ln,2): style_col_header(ws3b.cell(r,ci,t))
    ln_df = np.log(precios[cols_ln].replace(0,np.nan).dropna(how="all"))
    for ri,(idx,row) in enumerate(ln_df.iterrows(),1):
        r+=1
        c=ws3b.cell(r,1,idx.to_pydatetime()); style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
        for ci,t in enumerate(cols_ln,2):
            v=row[t] if t in row else None
            if pd.notna(v): style_data(ws3b.cell(r,ci,float(v)),ri,fmt_type="dec")
    autofit(ws3b)

    # ── HOJAS ANALÍTICAS (activos en filas, rangos en columnas) ───────────────
    def write_analitica(ws_a, df, titulo, fmt_type, subtitle=""):
        rangos = list(df.columns)
        ncols_a = len(rangos)+1
        merge_title(ws_a, ncols_a); write_title_block(ws_a, titulo, subtitle)
        r=4
        ws_a.cell(r,1,"Activo").font=font_col; ws_a.cell(r,1).fill=fill_subhdr
        ws_a.cell(r,1).alignment=align_c; ws_a.cell(r,1).border=border_thin
        for ci,rng in enumerate(rangos,2):
            fi_s,ff_s = fechas_rangos.get(rng,("",""))
            txt = f"{rng}\n{fi_s} → {ff_s}" if fi_s else rng
            c=ws_a.cell(r,ci,txt); style_col_header(c)
            ws_a.row_dimensions[r].height = 30
        for ri,(activo,row) in enumerate(df.iterrows(),1):
            r+=1
            c=ws_a.cell(r,1,activo); style_row_label(c)
            for ci,rng in enumerate(rangos,2):
                v=row[rng] if rng in row else np.nan
                if pd.notna(v): style_data(ws_a.cell(r,ci,float(v)),ri,float(v),fmt_type)
        autofit(ws_a)

    write_analitica(wb.create_sheet("Rend. Nominal"),   df_rn, "Rendimiento Nominal",             "pct")
    write_analitica(wb.create_sheet("Rend. Efectivo"),  df_re, "Rendimiento Efectivo Anualizado",  "pct")
    write_analitica(wb.create_sheet("Vol. Nominal"),    df_vn, "Volatilidad Nominal",              "pct")
    write_analitica(wb.create_sheet("Vol. Efectiva"),   df_ve, "Volatilidad Efectiva Anualizada",  "pct")
    write_analitica(wb.create_sheet("Sharpe"),          df_sh, "Sharpe Ratio",                    "dec3")

    # ── HOJA PERCENTILES ESPECTRO ──────────────────────────────────────────────
    ws_pe = wb.create_sheet("Percentiles Espectro")
    ncols_pe = len(tks_ok)+1
    merge_title(ws_pe,ncols_pe); write_title_block(ws_pe,"Espectro de Percentiles","Rendimientos diarios — Histórico completo")
    r=4
    ws_pe.cell(r,1,"Percentil").font=font_col; ws_pe.cell(r,1).fill=fill_subhdr
    ws_pe.cell(r,1).alignment=align_c; ws_pe.cell(r,1).border=border_thin
    for ci,t in enumerate(tks_ok,2): style_col_header(ws_pe.cell(r,ci,t))
    for ri,(idx,row) in enumerate(pct_esp.iterrows(),1):
        r+=1
        c=ws_pe.cell(r,1,idx); style_row_label(c)
        for ci,t in enumerate(tks_ok,2):
            v=row[t] if t in row else np.nan
            if pd.notna(v): style_data(ws_pe.cell(r,ci,float(v)),ri,float(v),"pct")
    autofit(ws_pe)

    # ── HOJA PERCENTIL ÚLTIMO RENDIMIENTO ─────────────────────────────────────
    ws_pu = wb.create_sheet("Percentil Últ. Rend.")
    merge_title(ws_pu,len(tks_ok)+1)
    write_title_block(ws_pu,"Percentil del Último Rendimiento",
                      f"Fecha: {fmt_fecha(rend.index[-1])}")
    r=4
    ws_pu.cell(r,1,"Métrica").font=font_col; ws_pu.cell(r,1).fill=fill_subhdr
    ws_pu.cell(r,1).alignment=align_c; ws_pu.cell(r,1).border=border_thin
    for ci,t in enumerate(tks_ok,2): style_col_header(ws_pu.cell(r,ci,t))
    # Fila 1: Último rendimiento
    r+=1; ws_pu.cell(r,1,"Últ. Rendimiento"); style_row_label(ws_pu.cell(r,1))
    for ci,t in enumerate(tks_ok,2):
        v=float(ult_rend[t])
        style_data(ws_pu.cell(r,ci,v),1,v,"pct")
    # Fila 2: Percentil
    r+=1; ws_pu.cell(r,1,"Percentil"); style_row_label(ws_pu.cell(r,1))
    for ci,t in enumerate(tks_ok,2):
        v=float(pct_ult[t])/100  # guardar como decimal para formato %
        style_data(ws_pu.cell(r,ci,v),2,v,"pct")
    autofit(ws_pu)

    # ── HOJA PORTAFOLIO (opcional) ───────────────────────────────────────────────
    if port_data:
        ws_p = wb.create_sheet("Portafolio")
        tks_p   = port_data["tks_ok"]
        modelo  = port_data["modelo"]
        rets_e  = port_data["rets_e"]
        vol_ind = port_data["vol_ind"]
        sh_ind  = port_data["sh_ind"]
        betas_s = port_data["betas_s"]
        ret_p   = port_data["ret_p"]
        vol_p   = port_data["vol_p"]
        sh_p    = port_data["sh_p"]
        ir      = port_data["ir"]
        te      = port_data["te"]
        ret_bi  = port_data["ret_bi"]
        pesos_n = port_data["pesos_norm"]
        bench_n = port_data.get("bench","Benchmark")
        escen_data = port_data.get("escenarios", {})  # {modelo: {escenario: {ret,vol,sh,ir,pesos}}}

        cols_p_t = tks_p + ["Portafolio"]
        ncols_pt = len(cols_p_t)+1
        merge_title(ws_p, ncols_pt)
        write_title_block(ws_p, f"Análisis de Portafolio — Modelo: {modelo}")

        # ── Sección 1: KPIs del portafolio ────────────────────────────────
        r = 4
        kpis = [
            ("Modelo",          modelo,   None),
            ("Retorno Portafolio", f"{ret_p:.2%}", ret_p),
            ("Volatilidad Portafolio", f"{vol_p:.2%}", None),
            (f"Sharpe Portafolio", f"{sh_p:.3f}", sh_p),
            ("IR",              f"{ir:.3f}" if not np.isnan(ir) else "N/A", ir if not np.isnan(ir) else None),
            ("Tracking Error",  f"{te:.2%}" if not np.isnan(te) else "N/A", None),
            (f"Retorno {bench_n}", f"{ret_bi:.2%}", ret_bi),
        ]
        # Título sección
        c_sec = ws_p.cell(r,1,"MÉTRICAS DEL PORTAFOLIO")
        c_sec.font = Font(name="Calibri",bold=True,size=10,color=BLANCO)
        c_sec.fill = fill_header; c_sec.alignment = align_l; c_sec.border = border_thin
        ws_p.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
        r+=1
        for ki,(label,val,num) in enumerate(kpis):
            cl=ws_p.cell(r,1,label); cl.font=font_activo; cl.fill=fill_subhdr
            cl.alignment=align_l; cl.border=border_thin
            cv=ws_p.cell(r,2,val); cv.font=font_data; cv.border=border_thin
            cv.alignment=align_r
            cv.fill = fill_alt if ki%2==0 else PatternFill("solid",fgColor=BLANCO)
            if num is not None:
                cv.fill = fill_pos if num>0 else fill_neg
            r+=1

        r+=1  # separador

        # ── Sección 2: Tabla principal W, Retorno, Vol, Sharpe, Beta, IR ──
        c_sec2=ws_p.cell(r,1,"TABLA PRINCIPAL DEL PORTAFOLIO")
        c_sec2.font=Font(name="Calibri",bold=True,size=10,color=BLANCO)
        c_sec2.fill=fill_header; c_sec2.alignment=align_l; c_sec2.border=border_thin
        ws_p.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols_pt)
        r+=1
        # Header fila
        ws_p.cell(r,1,"Métrica").font=font_col; ws_p.cell(r,1).fill=fill_subhdr
        ws_p.cell(r,1).alignment=align_c; ws_p.cell(r,1).border=border_thin
        for ci,t in enumerate(cols_p_t,2): style_col_header(ws_p.cell(r,ci,t))
        r+=1

        metricas_port = [
            ("W (%)",        [f"{pesos_n[i]*100:.1f}%" for i in range(len(tks_p))]+["100.0%"]),
            ("Retorno Esp.", [f"{rets_e.get(t,0):.2%}" for t in tks_p]+[f"{ret_p:.2%}"]),
            ("Volatilidad",  [f"{vol_ind.get(t,0):.2%}" for t in tks_p]+[f"{vol_p:.2%}"]),
            ("Sharpe",       [f"{sh_ind.get(t,0):.3f}" for t in tks_p]+[f"{sh_p:.3f}"]),
            ("Beta",         [f"{betas_s.get(t,1.):.4f}" for t in tks_p]+["—"]),
            ("IR",           ["—"]*len(tks_p)+[f"{ir:.3f}" if not np.isnan(ir) else "N/A"]),
        ]
        for mi,(met,vals) in enumerate(metricas_port):
            c_m=ws_p.cell(r,1,met); style_row_label(c_m)
            for ci,v in enumerate(vals,2):
                cel=ws_p.cell(r,ci,v); cel.font=font_data; cel.border=border_thin
                cel.alignment=align_r
                cel.fill=fill_alt if mi%2==0 else PatternFill("solid",fgColor=BLANCO)
                # color condicional donde aplica
                try:
                    num_v=float(v.replace("%","").replace("—",""))
                    if met in ("Retorno Esp.","W (%)") and num_v>0: cel.fill=fill_pos
                    elif met=="Retorno Esp." and num_v<0: cel.fill=fill_neg
                    elif met=="Sharpe" and num_v>1: cel.fill=fill_pos
                    elif met=="Sharpe" and num_v<0: cel.fill=fill_neg
                except: pass
            r+=1

        r+=1  # separador

        # ── Sección 3: Tablas de escenarios ───────────────────────────────
        if escen_data:
            c_sec3=ws_p.cell(r,1,"ESCENARIOS ÓPTIMOS POR MODELO")
            c_sec3.font=Font(name="Calibri",bold=True,size=10,color=BLANCO)
            c_sec3.fill=fill_header; c_sec3.alignment=align_l; c_sec3.border=border_thin
            ws_p.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6+len(tks_p))
            r+=1

            cols_esc=["Escenario","Retorno","Riesgo","Sharpe","IR"]+[f"W {t}" for t in tks_p]
            for mod_name, escens in escen_data.items():
                # Sub-título modelo
                cs=ws_p.cell(r,1,f"► {mod_name}")
                cs.font=Font(name="Calibri",bold=True,size=9,color=AZUL_ACE)
                cs.fill=fill_subhdr; cs.alignment=align_l; cs.border=border_thin
                ws_p.merge_cells(start_row=r,start_column=1,end_row=r,end_column=len(cols_esc))
                r+=1
                # Header escenarios
                for ci,ch in enumerate(cols_esc,1):
                    style_col_header(ws_p.cell(r,ci,ch))
                r+=1
                for ei,(esc_name,esc_vals) in enumerate(escens.items()):
                    style_row_label(ws_p.cell(r,1,esc_name))
                    ret_v=esc_vals.get("ret",0); vol_v=esc_vals.get("vol",0)
                    sh_v=esc_vals.get("sh",0);   ir_v=esc_vals.get("ir",np.nan)
                    pesos_v=esc_vals.get("pesos",[])
                    vals_esc=[f"{ret_v:.2%}",f"{vol_v:.2%}",f"{sh_v:.3f}",
                              f"{ir_v:.3f}" if not np.isnan(ir_v) else "N/A"]+[f"{w*100:.1f}%" for w in pesos_v]
                    for ci,v in enumerate(vals_esc,2):
                        cel=ws_p.cell(r,ci,v); cel.font=font_data; cel.border=border_thin; cel.alignment=align_r
                        cel.fill=fill_alt if ei%2==0 else PatternFill("solid",fgColor=BLANCO)
                        try:
                            nv=float(v.replace("%",""))
                            if ci==2 and nv>0: cel.fill=fill_pos
                            elif ci==2 and nv<0: cel.fill=fill_neg
                        except: pass
                    r+=1
                r+=1  # separador entre modelos

        autofit(ws_p)

    # ── HOJA(S) VALORACIÓN ESTADÍSTICA ───────────────────────────────────────
    if ve_data:
        AZUL_VE    = "1A3A5C"
        fill_ve_hdr  = PatternFill("solid", fgColor="0B3D2E")
        fill_ve_sub  = PatternFill("solid", fgColor=AZUL_VE)
        fill_ve_acc  = PatternFill("solid", fgColor="0D2B1A")

        # ve_data es una lista de dicts, uno por ticker
        ve_list = ve_data if isinstance(ve_data, list) else [ve_data]
        for ve_item in ve_list:
            ws_ve      = wb.create_sheet(f"VE {ve_item['ticker']}"[:31])
            ticker_ve  = ve_item["ticker"]
            bench_ve_n = ve_item["bench"]
            rango_ve_n = ve_item["rango"]
            ncols_ve   = 3

            # Título verde diferenciado
            ws_ve.merge_cells("A1:D1"); ws_ve.merge_cells("A2:D2"); ws_ve.merge_cells("A3:D3")
            c_t = ws_ve.cell(1,1,f"VALORACIÓN ESTADÍSTICA — {ticker_ve.upper()}")
            c_t.font = Font(name="Calibri",bold=True,size=14,color="FFFFFF")
            c_t.fill = fill_ve_hdr; c_t.alignment = align_l
            c_s = ws_ve.cell(2,1,f"Dashboard Financiero  |  {ticker_ve}  |  Benchmark: {bench_ve_n}  |  Rango: {rango_ve_n}  |  {gen_date}")
            c_s.font = Font(name="Calibri",size=8,color="A8D5B5",italic=True)
            c_s.fill = fill_ve_acc; c_s.alignment = align_l
            ws_ve.cell(3,1,"").fill = PatternFill("solid",fgColor="2ECC71")
            ws_ve.row_dimensions[1].height=28; ws_ve.row_dimensions[2].height=16; ws_ve.row_dimensions[3].height=4

            r = 4
            # Estadísticos regresión
            sec=ws_ve.cell(r,1,"ESTADÍSTICOS DE REGRESIÓN LINEAL")
            sec.font=Font(name="Calibri",bold=True,size=10,color="FFFFFF")
            sec.fill=fill_ve_hdr; sec.alignment=align_l; sec.border=border_thin
            ws_ve.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
            r+=1
            for ci,v in enumerate(["Estadístico","Valor"],1):
                c=ws_ve.cell(r,ci,v)
                c.font=Font(name="Calibri",bold=True,size=9,color="A8D5B5")
                c.fill=fill_ve_sub; c.alignment=align_c; c.border=border_thin
            r+=1
            for ei,(label,val,num,fmt_t) in enumerate(ve_item["stats"]):
                cl=ws_ve.cell(r,1,label)
                cl.font=Font(name="Calibri",bold=True,size=9,color="FFFFFF")
                cl.fill=fill_ve_sub; cl.alignment=align_l; cl.border=border_thin
                cv=ws_ve.cell(r,2,val)
                cv.font=font_data; cv.border=border_thin; cv.alignment=align_r
                cv.fill=fill_alt if ei%2==0 else PatternFill("solid",fgColor=BLANCO)
                if num is not None and isinstance(num,(int,float)) and not np.isnan(num):
                    if fmt_t=="pct":  cv.number_format="0.00%";    cv.value=num
                    elif fmt_t=="dec2": cv.number_format="0.00";   cv.value=num
                    elif fmt_t=="usd":  cv.number_format='"$"#,##0.00'; cv.value=num
                    if num>0 and fmt_t in ("pct","dec2"): cv.fill=fill_pos
                    elif num<0 and fmt_t in ("pct","dec2"): cv.fill=fill_neg
                r+=1
            r+=1

            # Percentiles
            sec2=ws_ve.cell(r,1,"PERCENTILES DE PRECIOS")
            sec2.font=Font(name="Calibri",bold=True,size=10,color="FFFFFF")
            sec2.fill=fill_ve_hdr; sec2.alignment=align_l; sec2.border=border_thin
            ws_ve.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
            r+=1
            for ci,v in enumerate(["Percentil","Precio"],1):
                c=ws_ve.cell(r,ci,v)
                c.font=Font(name="Calibri",bold=True,size=9,color="A8D5B5")
                c.fill=fill_ve_sub; c.alignment=align_c; c.border=border_thin
            r+=1
            for pi,(pk,pv) in enumerate(ve_item["percentiles"].items()):
                cl=ws_ve.cell(r,1,pk)
                cl.font=Font(name="Calibri",bold=True,size=9,color="FFFFFF")
                cl.fill=fill_ve_sub; cl.alignment=align_l; cl.border=border_thin
                cv=ws_ve.cell(r,2,pv); cv.font=font_data
                cv.number_format='"$"#,##0.00'; cv.border=border_thin; cv.alignment=align_r
                cv.fill=fill_alt if pi%2==0 else PatternFill("solid",fgColor=BLANCO)
                r+=1
            r+=1

            # Recomendación
            sec3=ws_ve.cell(r,1,"RECOMENDACIÓN ESTADÍSTICA PONDERADA")
            sec3.font=Font(name="Calibri",bold=True,size=10,color="FFFFFF")
            sec3.fill=fill_ve_hdr; sec3.alignment=align_l; sec3.border=border_thin
            ws_ve.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
            r+=1
            for ci,v in enumerate(["Método","Precio Objetivo","Potencial","Peso %"],1):
                c=ws_ve.cell(r,ci,v)
                c.font=Font(name="Calibri",bold=True,size=9,color="A8D5B5")
                c.fill=fill_ve_sub; c.alignment=align_c; c.border=border_thin
            r+=1
            for ri2,(metodo,p_obj,potencial,peso) in enumerate(ve_item["recomendacion"]):
                is_rec = metodo=="⭐ Recomendación"
                fr=PatternFill("solid",fgColor="0B3D2E") if is_rec else (fill_alt if ri2%2==0 else PatternFill("solid",fgColor=BLANCO))
                for ci,v in enumerate([metodo,p_obj,potencial,peso],1):
                    c=ws_ve.cell(r,ci,v)
                    c.font=Font(name="Calibri",bold=True if is_rec else False,size=9,
                                color="A8D5B5" if is_rec else TEXTO_OSC)
                    c.fill=fr; c.border=border_thin
                    c.alignment=align_r if ci>1 else align_l
                    if ci==2 and isinstance(v,(int,float)): c.number_format='"$"#,##0.00'
                    if ci==3 and isinstance(v,(int,float)):
                        c.number_format="0.00%"
                        c.fill=fill_pos if v>0 else fill_neg
                    if ci==4 and isinstance(v,(int,float)): c.number_format="0.00%"
                r+=1
            autofit(ws_ve)
    # ── HOJAS ANÁLISIS TÉCNICO ───────────────────────────────────────────────
    if vt_data:
        AZUL_VT   = "0D2B4A"
        fill_vt_h = PatternFill("solid", fgColor="0A1F35")
        fill_vt_s = PatternFill("solid", fgColor=AZUL_VT)

        def vt_col_header(cell, val):
            cell.value = val
            cell.font  = Font(name="Calibri",bold=True,size=9,color="26C6DA")
            cell.fill  = fill_vt_s; cell.alignment=align_c; cell.border=border_thin

        def vt_title(ws_t, titulo):
            ws_t.merge_cells("A1:Z1"); ws_t.merge_cells("A2:Z2"); ws_t.merge_cells("A3:Z3")
            c=ws_t.cell(1,1,f"ANÁLISIS TÉCNICO — {titulo.upper()}")
            c.font=Font(name="Calibri",bold=True,size=13,color="FFFFFF")
            c.fill=fill_vt_h; c.alignment=align_l
            c2=ws_t.cell(2,1,f"Dashboard Financiero  |  {gen_date}")
            c2.font=Font(name="Calibri",size=8,color="7FA8C9",italic=True)
            c2.fill=PatternFill("solid",fgColor="06111E"); c2.alignment=align_l
            ws_t.cell(3,1,"").fill=PatternFill("solid",fgColor="26C6DA")
            ws_t.row_dimensions[1].height=26; ws_t.row_dimensions[2].height=14; ws_t.row_dimensions[3].height=4

        tks_vt  = [k for k in vt_data.keys() if k != "__params__"]
        params  = vt_data.get("__params__",{})
        mm_ps   = params.get("mms",[5,20,100,200])
        mf,ms   = params.get("macd_fast",12), params.get("macd_slow",26)
        rsi_p   = params.get("rsi_per",14)
        rsi_sb  = params.get("rsi_sob",80)
        rsi_sv  = params.get("rsi_sov",20)

        # ── Hoja MM ───────────────────────────────────────────────────────
        ws_mm = wb.create_sheet("MM")
        vt_title(ws_mm, "Medias Móviles")
        r=4
        ws_mm.cell(r,1,"Fecha"); ws_mm.cell(r,1).font=Font(name="Calibri",bold=True,size=9,color="26C6DA")
        ws_mm.cell(r,1).fill=fill_vt_s; ws_mm.cell(r,1).alignment=align_c; ws_mm.cell(r,1).border=border_thin
        ci=2
        for tkr in tks_vt:
            if tkr=="__params__": continue
            for col_lbl in [f"{tkr}_Precio"]+[f"{tkr}_MM{p}" for p in mm_ps]:
                vt_col_header(ws_mm.cell(r,ci), col_lbl); ci+=1
        for ri,idx in enumerate(vt_data[tks_vt[0]]["serie"].index,1):
            r+=1
            c=ws_mm.cell(r,1,idx.to_pydatetime()); style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
            ci=2
            for tkr in tks_vt:
                if tkr=="__params__": continue
                td=vt_data[tkr]
                s=td["serie"]; mms_d=td["mms"]
                vals=[float(s.iloc[ri-1]) if ri-1<len(s) else np.nan]
                for p in mm_ps:
                    mm_s=mms_d.get(p)
                    vals.append(float(mm_s.iloc[ri-1]) if mm_s is not None and ri-1<len(mm_s) else np.nan)
                for v in vals:
                    if not np.isnan(v): style_data(ws_mm.cell(r,ci,v),ri,fmt_type="usd")
                    ci+=1
        autofit(ws_mm)

        # ── Hoja MACD ─────────────────────────────────────────────────────
        ws_macd = wb.create_sheet("MACD")
        vt_title(ws_macd, f"MACD ({mf},{ms}) Signal({params.get('macd_sig',9)})")
        r=4
        ws_macd.cell(r,1,"Fecha"); ws_macd.cell(r,1).font=Font(name="Calibri",bold=True,size=9,color="26C6DA")
        ws_macd.cell(r,1).fill=fill_vt_s; ws_macd.cell(r,1).alignment=align_c; ws_macd.cell(r,1).border=border_thin
        ci=2
        for tkr in tks_vt:
            for lbl in [f"{tkr}_EMA{mf}",f"{tkr}_EMA{ms}",f"{tkr}_MACD",f"{tkr}_Signal"]:
                vt_col_header(ws_macd.cell(r,ci), lbl); ci+=1
        sig_p = params.get("macd_sig",9)
        for ri,idx in enumerate(vt_data[tks_vt[0]]["serie"].index,1):
            r+=1
            c=ws_macd.cell(r,1,idx.to_pydatetime()); style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
            ci=2
            for tkr in tks_vt:
                td=vt_data[tkr]; serie_t=td["serie"]
                # Rolling simple means
                ema_f = serie_t.rolling(window=mf).mean()
                ema_s = serie_t.rolling(window=ms).mean()
                macd  = ema_f - ema_s
                sig   = macd.rolling(window=sig_p).mean()
                vals=[ema_f.iloc[ri-1], ema_s.iloc[ri-1],
                      macd.iloc[ri-1],  sig.iloc[ri-1]]
                for v in vals:
                    if pd.notna(v):
                        cel=ws_macd.cell(r,ci,float(v)); cel.number_format="0.0000"
                        style_data(cel,ri,fmt_type="dec")
                    ci+=1
        autofit(ws_macd)

        # ── Hoja RSI ──────────────────────────────────────────────────────
        ws_rsi = wb.create_sheet("RSI")
        vt_title(ws_rsi, f"RSI ({rsi_p})")
        r=4
        ws_rsi.cell(r,1,"Fecha"); ws_rsi.cell(r,1).font=Font(name="Calibri",bold=True,size=9,color="26C6DA")
        ws_rsi.cell(r,1).fill=fill_vt_s; ws_rsi.cell(r,1).alignment=align_c; ws_rsi.cell(r,1).border=border_thin
        ci=2
        rsi_cols=["Rendimiento","Alzas","Bajas",
                  f"Suma{rsi_p}Alzas",f"Suma{rsi_p}Bajas",
                  "R/S","RSI","Sobrecompra","Sobreventa"]
        for tkr in tks_vt:
            for lbl in [f"{tkr}_{c}" for c in rsi_cols]:
                vt_col_header(ws_rsi.cell(r,ci), lbl); ci+=1
        # Pre-calcular todas las series RSI por ticker
        rsi_series = {}
        for tkr in tks_vt:
            serie_t = vt_data[tkr]["serie"]
            rend_s  = serie_t.pct_change()
            # Alzas: valor si positivo, 0 si negativo (para sumar correctamente)
            alzas_val = rend_s.apply(lambda x: x if (pd.notna(x) and x > 0) else np.nan)
            bajas_val = rend_s.apply(lambda x: x if (pd.notna(x) and x < 0) else np.nan)
            # Para la suma: reemplazar NaN con 0 dentro de la ventana
            alzas_fill = rend_s.apply(lambda x: x if (pd.notna(x) and x > 0) else 0.0)
            bajas_fill = rend_s.apply(lambda x: x if (pd.notna(x) and x < 0) else 0.0)
            sum_a_s = alzas_fill.rolling(window=rsi_p, min_periods=rsi_p).sum()
            sum_b_s = bajas_fill.rolling(window=rsi_p, min_periods=rsi_p).sum()
            rs_s    = sum_a_s / sum_b_s.abs().replace(0, np.nan)
            rsi_s   = 100 - (100 / (1 + rs_s))
            rsi_series[tkr] = {
                "rend": rend_s, "alzas": alzas_val, "bajas": bajas_val,
                "sum_a": sum_a_s, "sum_b": sum_b_s,
                "rs": rs_s, "rsi": rsi_s
            }

        def _gv(s, i):
            val = s.iloc[i] if i < len(s) else np.nan
            return float(val) if pd.notna(val) else np.nan

        for ri,idx in enumerate(vt_data[tks_vt[0]]["serie"].index,1):
            r+=1
            c=ws_rsi.cell(r,1,idx.to_pydatetime()); style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
            ci=2
            for tkr in tks_vt:
                sr = rsi_series[tkr]
                vals = [
                    (_gv(sr["rend"],  ri-1), "pct"),
                    (_gv(sr["alzas"], ri-1), "pct"),
                    (_gv(sr["bajas"], ri-1), "pct"),
                    (_gv(sr["sum_a"], ri-1), "pct"),
                    (_gv(sr["sum_b"], ri-1), "pct"),
                    (_gv(sr["rs"],    ri-1), "dec2"),
                    (_gv(sr["rsi"],   ri-1), "dec2"),
                    (float(rsi_sb),          "dec2"),
                    (float(rsi_sv),          "dec2"),
                ]
                for v, fmt_t in vals:
                    if not np.isnan(v):
                        cel = ws_rsi.cell(r, ci, v)
                        style_data(cel, ri, v, fmt_t)
                        if fmt_t == "pct":  cel.number_format = "0.00%"
                        elif fmt_t == "dec2": cel.number_format = "0.00"
                    ci += 1
        autofit(ws_rsi)

        # ── Hoja Fibo Alza ────────────────────────────────────────────────
        ws_fa = wb.create_sheet("Fibo Alza")
        vt_title(ws_fa, "Fibonacci — Niveles de Alza")
        r=4
        fibo_hdrs=["Activo","Tendencia","Pivote Máx","Fecha Máx","Pivote Mín","Fecha Mín",
                   "Fibo1 (23,6%)","Fibo2 (38,2%)","Fibo3 (50,0%)","Fibo4 (61,8%)","Fibo5 (100%)"]
        for ci,h in enumerate(fibo_hdrs,1): vt_col_header(ws_fa.cell(r,ci),h)
        for ri,tkr in enumerate([t for t in tks_vt if t!="__params__"],1):
            r+=1; td=vt_data[tkr]
            vals=[tkr, td.get("tendencia_fibo","N/A"),
                  td.get("maximo",np.nan), fmt_fecha(td.get("fecha_max_fibo","")),
                  td.get("minimo",np.nan), fmt_fecha(td.get("fecha_min_fibo",""))]
            vals += [td["alza"].get(k,np.nan) for k in td["alza"]]
            for ci,v in enumerate(vals,1):
                cel=ws_fa.cell(r,ci,v)
                cel.font=font_data; cel.border=border_thin; cel.alignment=align_r if ci>2 else align_l
                cel.fill=fill_alt if ri%2==0 else PatternFill("solid",fgColor=BLANCO)
                if isinstance(v,float) and not np.isnan(v) and ci>4:
                    cel.number_format='"$"#,##0.00'
        autofit(ws_fa)

        # ── Hoja Fibo Baja ────────────────────────────────────────────────
        ws_fb = wb.create_sheet("Fibo Baja")
        vt_title(ws_fb, "Fibonacci — Niveles de Baja")
        r=4
        for ci,h in enumerate(fibo_hdrs,1): vt_col_header(ws_fb.cell(r,ci),h)
        for ri,tkr in enumerate([t for t in tks_vt if t!="__params__"],1):
            r+=1; td=vt_data[tkr]
            vals=[tkr, td.get("tendencia_fibo","N/A"),
                  td.get("maximo",np.nan), fmt_fecha(td.get("fecha_max_fibo","")),
                  td.get("minimo",np.nan), fmt_fecha(td.get("fecha_min_fibo",""))]
            vals += [td["baja"].get(k,np.nan) for k in td["baja"]]
            for ci,v in enumerate(vals,1):
                cel=ws_fb.cell(r,ci,v)
                cel.font=font_data; cel.border=border_thin; cel.alignment=align_r if ci>2 else align_l
                cel.fill=fill_alt if ri%2==0 else PatternFill("solid",fgColor=BLANCO)
                if isinstance(v,float) and not np.isnan(v) and ci>4:
                    cel.number_format='"$"#,##0.00'
        autofit(ws_fb)

        # ── Hoja Fibo Alza Diario ─────────────────────────────────────────
        ws_fad = wb.create_sheet("Fibo Alza Diario")
        vt_title(ws_fad, "Fibonacci Alza — Niveles Diarios")
        r=4
        ws_fad.cell(r,1,"Fecha"); ws_fad.cell(r,1).font=Font(name="Calibri",bold=True,size=9,color="26C6DA")
        ws_fad.cell(r,1).fill=fill_vt_s; ws_fad.cell(r,1).alignment=align_c; ws_fad.cell(r,1).border=border_thin
        ci=2
        fibo_lbl=["Fibo1(23,6%)","Fibo2(38,2%)","Fibo3(50,0%)","Fibo4(61,8%)","Fibo5(100%)"]
        for tkr in tks_vt:
            vt_col_header(ws_fad.cell(r,ci), f"{tkr}_Precio"); ci+=1
            for lbl in [f"{tkr}_{l}" for l in fibo_lbl]:
                vt_col_header(ws_fad.cell(r,ci), lbl); ci+=1
        for ri,idx in enumerate(vt_data[tks_vt[0]]["serie"].index,1):
            r+=1
            c=ws_fad.cell(r,1,idx.to_pydatetime()); style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
            ci=2
            for tkr in tks_vt:
                td=vt_data[tkr]
                precio_v = float(td["serie"].iloc[ri-1]) if ri-1<len(td["serie"]) else np.nan
                cel=ws_fad.cell(r,ci,precio_v if not np.isnan(precio_v) else None)
                if not np.isnan(precio_v): style_data(cel,ri,fmt_type="usd"); cel.number_format='"$"#,##0.00'
                ci+=1
                for nivel in td["alza"].values():
                    cel=ws_fad.cell(r,ci,float(nivel))
                    style_data(cel,ri,fmt_type="usd"); cel.number_format='"$"#,##0.00'; ci+=1
        autofit(ws_fad)

        # ── Hoja Fibo Baja Diario ──────────────────────────────────────────
        ws_fbd = wb.create_sheet("Fibo Baja Diario")
        vt_title(ws_fbd, "Fibonacci Baja — Niveles Diarios")
        r=4
        ws_fbd.cell(r,1,"Fecha"); ws_fbd.cell(r,1).font=Font(name="Calibri",bold=True,size=9,color="26C6DA")
        ws_fbd.cell(r,1).fill=fill_vt_s; ws_fbd.cell(r,1).alignment=align_c; ws_fbd.cell(r,1).border=border_thin
        ci=2
        for tkr in tks_vt:
            vt_col_header(ws_fbd.cell(r,ci), f"{tkr}_Precio"); ci+=1
            for lbl in [f"{tkr}_{l}" for l in fibo_lbl]:
                vt_col_header(ws_fbd.cell(r,ci), lbl); ci+=1
        for ri,idx in enumerate(vt_data[tks_vt[0]]["serie"].index,1):
            r+=1
            c=ws_fbd.cell(r,1,idx.to_pydatetime()); style_data(c,ri,fmt_type="fecha"); c.alignment=align_l
            ci=2
            for tkr in tks_vt:
                td=vt_data[tkr]
                precio_v = float(td["serie"].iloc[ri-1]) if ri-1<len(td["serie"]) else np.nan
                cel=ws_fbd.cell(r,ci,precio_v if not np.isnan(precio_v) else None)
                if not np.isnan(precio_v): style_data(cel,ri,fmt_type="usd"); cel.number_format='"$"#,##0.00'
                ci+=1
                for nivel in td["baja"].values():
                    cel=ws_fbd.cell(r,ci,float(nivel))
                    style_data(cel,ri,fmt_type="usd"); cel.number_format='"$"#,##0.00'; ci+=1
        autofit(ws_fbd)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ─── TABS ─────────────────────────────────────────────────────────────────────
# ── EXCEL HELPER FUNCTIONS (accesibles desde cualquier pestaña) ──────────────
def _construir_port_data_xl():
    if "port_calc" not in st.session_state or "precios" not in st.session_state:
        return None
    pc    = st.session_state["port_calc"]
    p     = st.session_state["precios"]
    bench = st.session_state.get("bench","")
    tks   = [t for t in st.session_state.get("tickers_list",[]) if t in p.columns]
    rend  = calc_rendimientos(p[tks])
    bench_ok = bench in p.columns
    rend_b   = calc_rendimientos(p[[bench]]).iloc[:,0] if bench_ok else None
    cov_xl   = cov_anualizada(rend)
    rf       = st.session_state.get("rf",0.045)
    escen_xl = {}
    modelos  = [("Markowitz", retorno_markowitz(rend))]
    if bench_ok:
        modelos.append(("CAPM", retorno_capm(rend,rend_b,rf)[0]))
    rm_mc = retorno_montecarlo(rend, 300)
    modelos.append(("Montecarlo", rm_mc))
    for mk_name,rm_xl in modelos:
        escen_xl[mk_name] = {}
        for ne,oe in [("Mín. Varianza","varianza"),("Máx. IR","ir"),
                      ("Máx. Sharpe","sharpe"),("Máx. Retorno","retorno")]:
            wo  = optimizar_portafolio(rm_xl,cov_xl,rf,oe)
            re_ = float(np.dot(wo,rm_xl)); ve_ = vol_portafolio(wo,cov_xl)
            se_ = (re_-rf)/ve_ if ve_>0 else np.nan
            rpd_= (rend*wo).sum(axis=1)
            te_ = tracking_error(rpd_,rend_b) if bench_ok else np.nan
            ie_ = information_ratio(re_,pc["ret_bi"],te_)
            escen_xl[mk_name][ne]={"ret":re_,"vol":ve_,"sh":se_,"ir":ie_,"pesos":list(wo)}
    return {
        "tks_ok":     tks,
        "modelo":     pc["modelo"],
        "rets_e":     {t: float(pc["rets_e"][t]) for t in tks},
        "vol_ind":    {t: float(pc["vol_ind"][t]) for t in tks},
        "sh_ind":     {t: float(pc["sh_ind"][t]) for t in tks},
        "betas_s":    {t: float(pc["betas_s"].get(t,1.)) for t in tks},
        "ret_p":      pc["ret_p"],"vol_p":pc["vol_p"],"sh_p":pc["sh_p"],
        "ir":         pc["ir"],"te":pc["te"],"ret_bi":pc["ret_bi"],
        "pesos_norm": pc["pesos_norm"],"bench":bench,"escenarios":escen_xl,
    }

def _construir_ve_data_xl():
    ve_keys = [k for k in st.session_state if k.startswith("ve_calc_")]
    if not ve_keys or "precios" not in st.session_state: return None
    p     = st.session_state["precios"]
    bench = st.session_state.get("bench","")
    tks   = [t for t in st.session_state.get("tickers_list",[]) if t in p.columns]
    bench_ok = bench in p.columns
    ve_n_xl  = st.session_state.get("ve_n", 1)
    ve_u_xl  = st.session_state.get("ve_u", "Años")
    if ve_u_xl == "Todo histórico":
        dias     = 99999
        rango_ve = "Todo el histórico"
    elif ve_u_xl == "Años":
        dias     = int(ve_n_xl * 365)
        rango_ve = f"{ve_n_xl} {ve_u_xl}"
    elif ve_u_xl == "Meses":
        dias     = int(ve_n_xl * 30.44)
        rango_ve = f"{ve_n_xl} {ve_u_xl}"
    else:
        dias     = int(ve_n_xl)
        rango_ve = f"{ve_n_xl} {ve_u_xl}"
    ff   = pd.Timestamp(p.index[-1])
    fi   = p.index[0] if dias >= 99999 else fecha_anterior_cercana(p.index, ff - pd.Timedelta(days=dias))
    mask = (p.index>=fi)&(p.index<=ff)
    from scipy import stats as sp_s
    all_ve = []
    for ticker_ve in tks:
        p_r = p.loc[mask,ticker_ve].dropna()
        b_r = p.loc[mask,bench].dropna() if bench_ok else None
        if len(p_r)<10: continue
        ln_a = np.log(p_r); ult_p = float(p_r.iloc[-1])
        if b_r is not None and len(b_r)>10:
            ln_b  = np.log(b_r); ult_b = float(b_r.iloc[-1])
            df_ln = pd.concat([ln_a,ln_b],axis=1).dropna()
            sl,ic,rv,_,_ = sp_s.linregress(df_ln.iloc[:,1].values, df_ln.iloc[:,0].values)
            r2   = rv**2; cor = float(np.corrcoef(df_ln.iloc[:,0].values,df_ln.iloc[:,1].values)[0,1])
            val_e= float(np.exp(ic+sl*np.log(ult_b))); pot_reg=(val_e/ult_p)-1
            ln20 = ln_a.iloc[-20:]; dv20=float(ln20.std()); pm20=float(ln20.mean())
            cv   = dv20/pm20 if pm20!=0 else np.nan
            conf = "Confiable" if r2>0.75 else "No confiable"
            est  = "Sobrevalorada" if pot_reg<0 else "Subvalorada"
        else:
            sl=ic=cor=r2=dv20=pm20=cv=0.; val_e=ult_p; pot_reg=0.; conf=est="N/A"
        p_arr = p_r.values; p50=float(np.percentile(p_arr,50)); pot_pct=(p50/ult_p)-1
        pct_a = (np.sum(p_arr<=ult_p)/len(p_arr))*100
        w_reg = st.session_state.get("peso_reg",60)/100; w_pct=1-w_reg
        pot_p = pot_reg*w_reg+pot_pct*w_pct; p_obj=ult_p*(1+pot_p)
        all_ve.append({
            "ticker":ticker_ve,"bench":bench,"rango":rango_ve,
            "stats":[
                ("Factor Correlación",f"{cor:.4f}",cor,"dec2"),
                ("Alfa",f"{ic:.4f}",ic,"dec2"),("Beta",f"{sl:.4f}",sl,"dec2"),
                ("R²",f"{r2:.4f}",r2,"dec2"),("Confiabilidad",conf,None,""),
                ("Valoración Estad.",f"${val_e:.2f}",val_e,"usd"),
                ("Potencial Reg.",f"{pot_reg:.2%}",pot_reg,"pct"),
                ("Estado",est,None,""),("Desvest 20d",f"{dv20:.2f}",dv20,"dec2"),
                ("Promedio 20d",f"{pm20:.2f}",pm20,"dec2"),
                ("Coef. Variación",f"{cv:.2%}" if not np.isnan(cv) else "N/A",
                 cv if not np.isnan(cv) else None,"pct"),
                ("Percentil Actual",f"{pct_a:.1f}%",pct_a/100,"pct"),
            ],
            "percentiles":{f"{n}%":float(np.percentile(p_arr,n)) for n in [0,1,5,25,50,75,95,99,100]},
            "recomendacion":[
                ("Regresión Lineal",val_e,pot_reg,w_reg),
                ("Percentiles P50",p50,pot_pct,w_pct),
                ("⭐ Recomendación",p_obj,pot_p,1.0),
            ],
        })
    return all_ve if all_ve else None

def _construir_vt_data_xl():
    """Recoge datos técnicos del session_state para el Excel."""
    if not any(k.startswith("vt_calc_") for k in st.session_state): return None
    if "precios" not in st.session_state: return None
    precios_x = st.session_state["precios"]
    tks_x     = [t for t in st.session_state.get("tickers_list",[]) if t in precios_x.columns]
    mm1_x = st.session_state.get("mm1",5); mm2_x = st.session_state.get("mm2",20)
    mm3_x = st.session_state.get("mm3",100); mm4_x = st.session_state.get("mm4",200)
    mf_x  = st.session_state.get("macd_fast",12); ms_x = st.session_state.get("macd_slow",26)
    ms2_x = st.session_state.get("macd_sig",9)
    rp_x  = st.session_state.get("rsi_per",14)
    rs_x  = st.session_state.get("rsi_sob",80); rv_x = st.session_state.get("rsi_sov",20)
    fs_x  = st.session_state.get("fibo_sens",10)

    rango_vt_x = st.session_state.get("rango_vt","1 Año")
    dvm_x = {"1 Mes":30,"3 Meses":90,"6 Meses":180,"1 Año":365,"2 Años":730,
              "3 Años":1095,"5 Años":1825,"Todo":99999}
    dias_x = dvm_x.get(rango_vt_x, 365)
    ff_x   = pd.Timestamp(precios_x.index[-1])
    fi_x   = precios_x.index[0] if dias_x>=99999 else fecha_anterior_cercana(
                 precios_x.index, ff_x - pd.Timedelta(days=dias_x))
    mask_x = (precios_x.index >= fi_x) & (precios_x.index <= ff_x)

    result = {"__params__": {"mms":[mm1_x,mm2_x,mm3_x,mm4_x],
                              "macd_fast":mf_x,"macd_slow":ms_x,
                              "rsi_per":rp_x,"rsi_sob":rs_x,"rsi_sov":rv_x}}
    for tkr in tks_x:
        serie = precios_x.loc[mask_x, tkr].dropna()
        if len(serie) < 30: continue
        mms_d  = {p: serie.rolling(p).mean() for p in [mm1_x,mm2_x,mm3_x,mm4_x]}
        ema_f  = serie.rolling(window=mf_x).mean()
        ema_s  = serie.rolling(window=ms_x).mean()
        macd_s = ema_f - ema_s
        sig_s  = macd_s.rolling(window=ms2_x).mean()
        hist_s = macd_s - sig_s
        rend_s  = serie.pct_change()
        alzas_s = rend_s.apply(lambda x: x if (pd.notna(x) and x>0) else np.nan)
        bajas_s = rend_s.apply(lambda x: x if (pd.notna(x) and x<0) else np.nan)
        alzas_f = rend_s.apply(lambda x: x if (pd.notna(x) and x>0) else 0.0)
        bajas_f = rend_s.apply(lambda x: x if (pd.notna(x) and x<0) else 0.0)
        sum_a_s = alzas_f.rolling(window=rp_x, min_periods=rp_x).sum()
        sum_b_s = bajas_f.rolling(window=rp_x, min_periods=rp_x).sum()
        rs_s    = sum_a_s / sum_b_s.abs().replace(0,np.nan)
        rsi_s   = 100-(100/(1+rs_s))
        # Fibonacci
        ratios = [0.236,0.382,0.500,0.618,1.000]
        labels = ["Fibo 1 (23,6%)","Fibo 2 (38,2%)","Fibo 3 (50,0%)","Fibo 4 (61,8%)","Fibo 5 (100%)"]
        n = max(int(fs_x),2); vals_s = serie.values; idx_a = serie.index
        pmx=None; pmn=None
        for i in range(n,len(vals_s)-n):
            v=vals_s[i-n:i+n+1]
            if vals_s[i]==v.max(): pmx=i
            if vals_s[i]==v.min(): pmn=i
        if pmx is None: pmx=int(np.argmax(vals_s))
        if pmn is None: pmn=int(np.argmin(vals_s))
        maximo=float(vals_s[pmx]); minimo=float(vals_s[pmn])
        tendencia="BAJISTA" if idx_a[pmx]>idx_a[pmn] else "ALCISTA"
        rng=maximo-minimo if maximo!=minimo else maximo*0.01
        alza_d={lb:maximo-r*rng for lb,r in zip(labels,ratios)}
        baja_d={lb:minimo+r*rng for lb,r in zip(labels,ratios)}
        result[tkr] = dict(serie=serie,mms=mms_d,macd_s=macd_s,sig_s=sig_s,hist_s=hist_s,
                           rsi_serie=rsi_s,alza=alza_d,baja=baja_d,
                           maximo=maximo,minimo=minimo,tendencia_fibo=tendencia,
                           fecha_max_fibo=idx_a[pmx],fecha_min_fibo=idx_a[pmn])
    return result if len(result)>1 else None

def _exportar_excel(inc_basico=True, port_data=None, ve_data=None, vt_data=None):
    if "precios" not in st.session_state: return None
    p     = st.session_state["precios"]
    bench = st.session_state.get("bench","")
    tks   = [t for t in st.session_state.get("tickers_list",[]) if t in p.columns]
    rend  = calc_rendimientos(p[tks])
    ff_dt = pd.Timestamp(p.index[-1]); fi_dt = pd.Timestamp(p.index[0])
    dias_total=(ff_dt-fi_dt).days
    # Rebuild analysis tables
    n1=st.session_state.get("n1",1); u1=st.session_state.get("u1","Meses")
    n2=st.session_state.get("n2",3); u2=st.session_state.get("u2","Meses")
    n3=st.session_state.get("n3",1); u3=st.session_state.get("u3","Años")
    def a_dias(n,u): return int(n*365) if u=="Años" else (int(n*30.44) if u=="Meses" else int(n))
    def lbl(n,u): return f"{n} {u}"
    rangos=[(lbl(n1,u1),a_dias(n1,u1)),(lbl(n2,u2),a_dias(n2,u2)),
            (lbl(n3,u3),a_dias(n3,u3)),("Total histórico",dias_total)]
    rn_r={}; re_r={}; vn_r={}; ve_r={}; sh_r={}; fechas_r={}
    rf=st.session_state.get("rf",0.045)
    for nombre,dias in rangos:
        fi_r=fi_dt if dias>=dias_total else ff_dt-pd.Timedelta(days=dias)
        fi_real=fecha_anterior_cercana(p.index,fi_r)
        ff_real=fecha_anterior_cercana(p.index,ff_dt)
        fechas_r[nombre]=(fmt_fecha(fi_real),fmt_fecha(ff_real))
        rn=rend_nominal(p[tks],fi_r,ff_dt); dias_r=max((ff_dt-fi_r).days,1)
        re=rend_efectivo(rn,dias_r); vn=vol_nominal(rend,fi_r,ff_dt)
        ve_=vn*np.sqrt(365); sh=re/ve_.replace(0,np.nan)
        rn_r[nombre]=rn; re_r[nombre]=re; vn_r[nombre]=vn; ve_r[nombre]=ve_; sh_r[nombre]=sh
    df_rn=pd.DataFrame(rn_r); df_re=pd.DataFrame(re_r)
    df_vn=pd.DataFrame(vn_r); df_ve=pd.DataFrame(ve_r); df_sh=pd.DataFrame(sh_r)
    pct_esp=pct_espectro(rend); ult_rend=rend.iloc[-1]
    pct_ult=pct_actual_serie(rend,ult_rend)
    return exportar_excel_bloque1(p,rend,df_rn,df_re,df_vn,df_ve,df_sh,
                                   pct_esp,pct_ult,ult_rend,tks,bench,fechas_r,
                                   port_data=port_data if inc_basico else None,
                                   ve_data=ve_data,
                                   vt_data=vt_data)

tab1,tab2,tab3,tab4,tab5,tab6 = st.tabs([
    "01 · Análisis Básico","02 · Portafolios",
    "03 · Val. Estadística","04 · Val. Técnica",
    "05 · Val. Fundamental","06 · Exportar PDF",
])

# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 1 — ANÁLISIS BÁSICO
# ══════════════════════════════════════════════════════════════════════════════
with tab1:

    with st.expander("⚙️  CONFIGURACIÓN GENERAL", expanded=True):
        c1,c2 = st.columns([3,1])
        with c1: tickers_input = st.text_input("Tickers (separados por coma)", value="AAPL, MSFT, GOOGL, AMZN")
        with c2: bench_input   = st.text_input("Benchmark", value="^GSPC")

        # RF
        crf1,crf2,crf3 = st.columns([1,1,2])
        with crf1: rf_ticker = st.text_input("Ticker RF", value="^TNX", help="^TNX=Bono USA 10Y")
        with crf2: rf_manual = st.number_input("RF manual (%)", 0.,30.,0.,0.1,format="%.4f",
                                                help="0 = usar el ticker")
        with crf3:
            st.markdown("<br>",unsafe_allow_html=True)
            if st.button("📡 Obtener RF del mercado",key="btn_rf"):
                rf_live = obtener_rf(rf_ticker)
                st.session_state["rf_live"] = rf_live
                st.success(f"RF obtenida: {rf_live*100:.4f}% ({rf_ticker})")
        rf = rf_manual/100 if rf_manual>0 else st.session_state.get("rf_live", obtener_rf("^TNX"))
        st.caption(f"RF activa: **{rf*100:.4f}%**")

        # Fechas opcionales
        st.markdown("---")
        st.markdown('<div class="section-header">Rango de la Base de Datos (opcional)</div>', unsafe_allow_html=True)
        st.caption("Dejar en blanco para obtener el máximo histórico disponible.")
        cf1,cf2 = st.columns(2)
        with cf1:
            usar_fi = st.checkbox("Definir fecha inicio", value=False)
            fi_global = st.date_input("Fecha inicio (inclusiva)", value=datetime(2020,1,1),
                                       min_value=datetime(1900,1,1),
                                       max_value=datetime.today(),
                                       disabled=not usar_fi) if usar_fi else None
        with cf2:
            usar_ff = st.checkbox("Definir fecha fin", value=False)
            ff_global = st.date_input("Fecha fin (inclusiva)", value=datetime.today(),
                                       min_value=datetime(1900,1,1),
                                       max_value=datetime.today(),
                                       disabled=not usar_ff) if usar_ff else None

        fi_str = str(fi_global) if usar_fi and fi_global else ""
        ff_str = str(ff_global) if usar_ff and ff_global else ""

        # Rangos
        st.markdown("---")
        st.markdown('<div class="section-header">Rangos de Tiempo</div>', unsafe_allow_html=True)
        unidades=["Días","Meses","Años"]
        def a_dias(n,u): return int(n*365) if u=="Años" else (int(n*30.44) if u=="Meses" else int(n))
        def label_rango(n,u): return f"{n} {u}"
        cr1a,cr1b,cr2a,cr2b,cr3a,cr3b = st.columns(6)
        with cr1a: n1=st.number_input("Rango 1",1,999,1,key="n1")
        with cr1b: u1=st.selectbox("u1",unidades,index=2,key="u1",label_visibility="hidden")
        with cr2a: n2=st.number_input("Rango 2",1,999,3,key="n2")
        with cr2b: u2=st.selectbox("u2",unidades,index=2,key="u2",label_visibility="hidden")
        with cr3a: n3=st.number_input("Rango 3",1,999,1,key="n3")
        with cr3b: u3=st.selectbox("u3",unidades,index=2,key="u3",label_visibility="hidden")
        d1=a_dias(n1,u1); d2=a_dias(n2,u2); d3=a_dias(n3,u3)

        btn_cargar = st.button("🔄  CARGAR Y CALCULAR",key="btn1")

    # ── CARGA ─────────────────────────────────────────────────────────────────
    if btn_cargar or "precios" in st.session_state:
        if btn_cargar:
            with st.spinner("Descargando datos de Yahoo Finance..."):
                precios,tickers_list = cargar_datos(tickers_input,bench_input,fi_str,ff_str)
                if precios is None:
                    st.error("Error al descargar. Verifica los tickers."); st.stop()
                st.session_state.update({"precios":precios,"tickers_list":tickers_list,
                    "bench":bench_input.strip().upper(),"rf":rf})

        precios      = st.session_state["precios"]
        tickers_list = st.session_state["tickers_list"]
        bench        = st.session_state["bench"]
        rf           = st.session_state.get("rf",0.045)
        tks_ok       = [t for t in tickers_list if t in precios.columns]
        bench_ok     = bench in precios.columns
        if not tks_ok: st.error("Ningún ticker válido."); st.stop()

        rend       = calc_rendimientos(precios[tks_ok])
        ff_dt      = pd.Timestamp(precios.index[-1])
        fi_dt      = pd.Timestamp(precios.index[0])
        dias_total = (ff_dt-fi_dt).days

        rangos_def = [(label_rango(n1,u1),d1),(label_rango(n2,u2),d2),
                      (label_rango(n3,u3),d3),("Total histórico",dias_total)]

        # ── KPI ───────────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">Resumen</div>', unsafe_allow_html=True)
        periodo_txt = periodo_legible(dias_total)
        fi_fmt=fmt_fecha(fi_dt); ff_fmt=fmt_fecha(ff_dt)
        st.markdown(f"""
        <div class="kpi-grid">
          <div class="kpi-card"><div class="kpi-label">Activos</div>
            <div class="kpi-value">{len(tks_ok)}</div>
            <div class="kpi-delta">{', '.join(tks_ok[:5])}{'...' if len(tks_ok)>5 else ''}</div></div>
          <div class="kpi-card"><div class="kpi-label">Observaciones</div>
            <div class="kpi-value">{len(precios):,}</div>
            <div class="kpi-delta">{fi_fmt} → {ff_fmt}</div></div>
          <div class="kpi-card"><div class="kpi-label">Período</div>
            <div class="kpi-value" style="font-size:1rem">{periodo_txt}</div>
            <div class="kpi-delta">{fi_fmt} → {ff_fmt}</div></div>
          <div class="kpi-card"><div class="kpi-label">RF</div>
            <div class="kpi-value">{rf*100:.4f}%</div><div class="kpi-delta">anual</div></div>
          <div class="kpi-card"><div class="kpi-label">Benchmark</div>
            <div class="kpi-value" style="font-size:.9rem">{bench}</div>
            <div class="kpi-delta">{'✓ OK' if bench_ok else '✗ No encontrado'}</div></div>
        </div>""", unsafe_allow_html=True)

        # ── CONSTRUIR TABLAS ───────────────────────────────────────────────────
        rn_rows={}; re_rows={}; vn_rows={}; ve_rows={}; sh_rows={}; fechas_rangos={}
        for nombre,dias in rangos_def:
            fi_r = fi_dt if dias>=dias_total else ff_dt-pd.Timedelta(days=dias)
            ff_r = ff_dt; dias_real=max((ff_r-fi_r).days,1)
            idx_p = precios.index
            fi_real = fecha_anterior_cercana(idx_p,fi_r)
            ff_real = fecha_anterior_cercana(idx_p,ff_r)
            fechas_rangos[nombre] = (fmt_fecha(fi_real),fmt_fecha(ff_real))
            rn=rend_nominal(precios[tks_ok],fi_r,ff_r)
            re=rend_efectivo(rn,dias_real)
            vn=vol_nominal(rend,fi_r,ff_r)
            ve=vn*np.sqrt(365)
            sh=sharpe_ratio(re,ve,rf)
            rn_rows[nombre]=rn; re_rows[nombre]=re
            vn_rows[nombre]=vn; ve_rows[nombre]=ve; sh_rows[nombre]=sh

        # DataFrames: activos en filas, rangos en columnas
        df_rn=pd.DataFrame(rn_rows); df_re=pd.DataFrame(re_rows)
        df_vn=pd.DataFrame(vn_rows); df_ve=pd.DataFrame(ve_rows); df_sh=pd.DataFrame(sh_rows)

        # ── TABLAS ────────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">Rendimientos</div>', unsafe_allow_html=True)
        ct1,ct2=st.columns(2)
        with ct1: st.markdown(tabla_html(df_rn,".2%",col_headers=fechas_rangos,titulo="Rendimiento Nominal"),unsafe_allow_html=True)
        with ct2: st.markdown(tabla_html(df_re,".2%",col_headers=fechas_rangos,titulo="Rendimiento Efectivo Anualizado"),unsafe_allow_html=True)

        st.markdown('<div class="section-header">Volatilidad</div>', unsafe_allow_html=True)
        ct3,ct4=st.columns(2)
        with ct3: st.markdown(tabla_html(df_vn,".2%",col_headers=fechas_rangos,titulo="Volatilidad Nominal"),unsafe_allow_html=True)
        with ct4: st.markdown(tabla_html(df_ve,".2%",col_headers=fechas_rangos,titulo="Volatilidad Efectiva Anualizada"),unsafe_allow_html=True)

        st.markdown('<div class="section-header">Sharpe Ratio</div>', unsafe_allow_html=True)
        st.markdown(tabla_html(df_sh,".3f",0.5,col_headers=fechas_rangos,titulo="Sharpe (Ret. efectivo / Vol. efectiva)"),unsafe_allow_html=True)

        # ── PERCENTILES ───────────────────────────────────────────────────────
        st.markdown('<div class="section-header">Análisis de Percentiles</div>', unsafe_allow_html=True)
        pct_esp  = pct_espectro(rend)
        ult_rend = rend.iloc[-1]
        pct_ult  = pct_actual_serie(rend,ult_rend)
        ult_fecha_fmt = fmt_fecha(rend.index[-1])

        # Espectro: percentiles en filas, activos en columnas
        st.markdown(tabla_html(pct_esp.T,".2%",titulo="Espectro de rendimientos diarios (histórico completo)"),unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)
        # Último rendimiento: activos en filas, métricas en columnas
        df_ult = pd.DataFrame({
            "Últ. Rendimiento": ult_rend,
            "Percentil":        pct_ult/100,
        })
        hdr_ult = "<th>Últ. Rendimiento</th><th>Percentil</th>"
        rows_ult = ""
        for t in tks_ok:
            rv=ult_rend[t]; pv=pct_ult[t]/100
            cr="td-positive" if rv>0 else "td-negative"
            cp="td-positive" if pv<0.3 else ("td-negative" if pv>0.7 else "td-neutral")
            rows_ult += f'<tr><td>{t}</td><td class="{cr}">{rv:.2%}</td><td class="{cp}">{pv:.1%}</td></tr>'
        st.markdown(
            f'<p style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#4AB3F4;'
            f'letter-spacing:.1em;text-transform:uppercase;margin-bottom:.35rem;">'
            f'Percentil del último rendimiento ({ult_fecha_fmt})</p>'
            f'<table class="styled-table"><thead><tr><th>Activo</th>{hdr_ult}</tr></thead>'
            f'<tbody>{rows_ult}</tbody></table>',unsafe_allow_html=True)

        # ── GRÁFICAS ──────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">Gráficas Interactivas</div>', unsafe_allow_html=True)
        rango_visual=st.selectbox("Rango visual",["1M","3M","6M","1A","3A","5A","Todo"],index=3)
        dvm={"1M":30,"3M":90,"6M":180,"1A":365,"3A":1095,"5A":1825,"Todo":99999}
        fi_vis=ff_dt-pd.Timedelta(days=dvm[rango_visual]); mv=precios.index>=fi_vis
        gt1,gt2,gt3=st.tabs(["Precios","Rendimientos Diarios","Base 100"])
        with gt1:
            fig=go.Figure()
            for i,t in enumerate(tks_ok):
                s=precios.loc[mv,t].dropna()
                fig.add_trace(go.Scatter(x=s.index,y=s.values,name=t,
                    line=dict(color=PALETTE[i%len(PALETTE)],width=1.6),
                    hovertemplate=f"<b>{t}</b><br>%{{x|%d/%m/%y}}<br>${{y:.2f}}<extra></extra>"))
            lo=base_layout("Precios de Cierre Ajustados",450)
            lo["xaxis"]["rangeslider"]=dict(visible=True,bgcolor="#0A1628",bordercolor="#1A3A5C")
            fig.update_layout(**lo); st.plotly_chart(fig,use_container_width=True)
        with gt2:
            fig2=go.Figure()
            for i,t in enumerate(tks_ok):
                s=rend.loc[rend.index>=fi_vis,t].dropna()
                fig2.add_trace(go.Bar(x=s.index,y=s.values*100,name=t,
                    marker_color=PALETTE[i%len(PALETTE)],opacity=0.7,
                    hovertemplate=f"<b>{t}</b><br>%{{x|%d/%m/%y}}<br>%{{y:.2f}}%<extra></extra>"))
            lo2=base_layout("Rendimientos Diarios (%)",450); lo2["barmode"]="overlay"
            fig2.update_layout(**lo2); st.plotly_chart(fig2,use_container_width=True)
        with gt3:
            fig3=go.Figure()
            for i,t in enumerate(tks_ok+([bench] if bench_ok else [])):
                if t in precios.columns:
                    s=precios.loc[mv,t].dropna(); b=base_100(s)
                    fig3.add_trace(go.Scatter(x=b.index,y=b.values,name=t,
                        line=dict(color=PALETTE[i%len(PALETTE)],
                                  width=2. if t==bench else 1.6,
                                  dash="dot" if t==bench else "solid"),
                        hovertemplate=f"<b>{t}</b><br>%{{x|%d/%m/%y}}<br>%{{y:.1f}}<extra></extra>"))
            fig3.add_hline(y=100,line_dash="dash",line_color="#4A7FA5",opacity=0.4)
            fig3.update_layout(**base_layout("Evolución Base 100",450))
            st.plotly_chart(fig3,use_container_width=True)

        # ── DECISIÓN ──────────────────────────────────────────────────────────
        st.markdown('<div class="section-header">Tabla de Decisión</div>', unsafe_allow_html=True)
        sh_r3=df_sh.iloc[:,2]  # columna Rango 3
        dec='<table class="styled-table"><thead><tr><th>Activo</th><th>Sharpe (R3)</th><th>Percentil Actual</th><th>Señal</th></tr></thead><tbody>'
        for t in tks_ok:
            sv=sh_r3.get(t,np.nan); pv=pct_ult.get(t,50.)
            señal=("🟢 ATRACTIVO" if (not np.isnan(sv) and sv>1 and pv<40) else
                   "🟡 MODERADO"  if (not np.isnan(sv) and sv>0.5 and pv<60) else
                   "🔴 NEGATIVO"  if (np.isnan(sv) or sv<0) else "⚪ NEUTRAL")
            sc="td-positive" if (not np.isnan(sv) and sv>0) else "td-negative"
            pc="td-positive" if pv<30 else ("td-negative" if pv>70 else "td-neutral")
            sv_str=f"{sv:.3f}" if not np.isnan(sv) else "N/A"
            dec+=f'<tr><td class="td-label">{t}</td><td class="{sc}">{sv_str}</td><td class="{pc}">{pv:.1f}%</td><td>{señal}</td></tr>'
        dec+="</tbody></table>"
        st.markdown(dec,unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 2 — PORTAFOLIOS
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    if "precios" not in st.session_state:
        st.info("⬅️  Primero carga los datos en **01 · Análisis Básico**")
    else:
        precios=st.session_state["precios"]; tickers_list=st.session_state["tickers_list"]
        bench=st.session_state["bench"]; rf=st.session_state.get("rf",0.045)
        tks_ok=[t for t in tickers_list if t in precios.columns]
        bench_ok=bench in precios.columns
        rend=calc_rendimientos(precios[tks_ok])
        rend_bench=calc_rendimientos(precios[[bench]]).iloc[:,0] if bench_ok else None

        with st.expander("⚙️  CONFIGURACIÓN DEL PORTAFOLIO",expanded=True):
            cp1,cp2,cp3=st.columns([2,1,1])
            with cp1: modelo=st.selectbox("Modelo",["Markowitz","CAPM","Montecarlo"],key="modelo_port")
            with cp2:
                st.markdown(f'<div style="font-family:IBM Plex Mono,monospace;font-size:.63rem;'
                            f'color:#4A7FA5;letter-spacing:.1em;text-transform:uppercase;'
                            f'margin-bottom:.35rem;">RF (del Bloque 1)</div>'
                            f'<div style="font-family:IBM Plex Mono,monospace;font-size:1.1rem;'
                            f'font-weight:600;color:#4AB3F4;">{rf*100:.4f}%</div>',
                            unsafe_allow_html=True)
                rf_port = rf
            with cp3: bench_label=st.text_input("Nombre índice",value=bench,key="bench_label")
            st.markdown("---")
            st.markdown('<div class="section-header">Pesos W (deben sumar 100%)</div>',unsafe_allow_html=True)
            cols_w=st.columns(len(tks_ok)); pesos_vals=[]; dw=round(100/len(tks_ok),1)
            for col,t in zip(cols_w,tks_ok):
                with col: pesos_vals.append(st.number_input(f"{t}",0.,100.,dw,1.,format="%.1f",key=f"w_{t}"))
            total_w=sum(pesos_vals)
            if abs(total_w-100)>0.01: st.warning(f"⚠️  Los pesos suman {total_w:.1f}%. Se normalizarán a 100%.")
            pesos_norm=[p/total_w for p in pesos_vals]
            btn_port=st.button("📐  CALCULAR PORTAFOLIO",key="btn_port")

        if btn_port or "port_calc" in st.session_state:
            if btn_port:
                with st.spinner(f"Calculando ({modelo})..."):
                    if modelo=="Markowitz":
                        rets_e=retorno_markowitz(rend)
                        betas_s=calcular_betas(rend,rend_bench) if bench_ok else pd.Series({t:1. for t in tks_ok})
                    elif modelo=="CAPM":
                        rets_e,betas_s=retorno_capm(rend,rend_bench,rf_port) if bench_ok else (retorno_markowitz(rend),pd.Series({t:1. for t in tks_ok}))
                    else:
                        rets_e=retorno_montecarlo(rend,1000)
                        betas_s=calcular_betas(rend,rend_bench) if bench_ok else pd.Series({t:1. for t in tks_ok})
                    cov_an=cov_anualizada(rend); vol_ind=vol_individual(rend)
                    ret_p=float(np.dot(pesos_norm,rets_e)); vol_p=vol_portafolio(pesos_norm,cov_an)
                    sh_p=(ret_p-rf_port)/vol_p if vol_p>0 else np.nan
                    ret_bi=float(rend_bench.mean()*365) if bench_ok else 0.
                    rpd=(rend*pesos_norm).sum(axis=1)
                    te=tracking_error(rpd,rend_bench) if bench_ok else np.nan
                    ir=information_ratio(ret_p,ret_bi,te)
                    sh_ind=(rets_e-rf_port)/vol_ind
                    st.session_state["port_calc"]=dict(modelo=modelo,rets_e=rets_e,betas_s=betas_s,
                        cov_an=cov_an,vol_ind=vol_ind,sh_ind=sh_ind,ret_p=ret_p,vol_p=vol_p,
                        sh_p=sh_p,ret_bi=ret_bi,te=te,ir=ir,pesos_norm=pesos_norm,rf_port=rf_port)

            pc=st.session_state["port_calc"]
            rets_e=pc["rets_e"]; betas_s=pc["betas_s"]; cov_an=pc["cov_an"]
            vol_ind=pc["vol_ind"]; sh_ind=pc["sh_ind"]; ret_p=pc["ret_p"]
            vol_p=pc["vol_p"]; sh_p=pc["sh_p"]; ret_bi=pc["ret_bi"]
            te=pc["te"]; ir=pc["ir"]; pesos_norm=pc["pesos_norm"]; rf_port=pc["rf_port"]
            # Asegurar que pesos_norm tiene el mismo largo que tks_ok
            if len(pesos_norm) != len(tks_ok):
                if len(pesos_norm) > len(tks_ok):
                    pesos_norm = pesos_norm[:len(tks_ok)]
                else:
                    pesos_norm = list(pesos_norm) + [0.0]*(len(tks_ok)-len(pesos_norm))
                total_fix = sum(pesos_norm)
                if total_fix > 0:
                    pesos_norm = [p/total_fix for p in pesos_norm]

            st.markdown('<div class="section-header">Métricas del Portafolio</div>',unsafe_allow_html=True)
            st.markdown(f"""<div class="kpi-grid">
              <div class="kpi-card"><div class="kpi-label">Modelo</div><div class="kpi-value" style="font-size:.9rem">{pc['modelo']}</div></div>
              <div class="kpi-card"><div class="kpi-label">Retorno Esperado</div>
                <div class="kpi-value {'positive' if ret_p>0 else 'negative'}">{ret_p:.2%}</div><div class="kpi-delta">anualizado</div></div>
              <div class="kpi-card"><div class="kpi-label">Volatilidad</div><div class="kpi-value">{vol_p:.2%}</div><div class="kpi-delta">anualizada</div></div>
              <div class="kpi-card"><div class="kpi-label">Sharpe</div>
                <div class="kpi-value {'positive' if sh_p>1 else ('negative' if sh_p<0 else '')}">{sh_p:.3f}</div></div>
              <div class="kpi-card"><div class="kpi-label">IR</div>
                <div class="kpi-value {'positive' if not np.isnan(ir) and ir>0 else 'negative'}">{f"{ir:.3f}" if not np.isnan(ir) else "N/A"}</div></div>
              <div class="kpi-card"><div class="kpi-label">Tracking Error</div>
                <div class="kpi-value">{f"{te:.2%}" if not np.isnan(te) else "N/A"}</div></div>
              <div class="kpi-card"><div class="kpi-label">Ret. {bench_label}</div>
                <div class="kpi-value {'positive' if ret_bi>0 else 'negative'}">{ret_bi:.2%}</div><div class="kpi-delta">promedio×365</div></div>
            </div>""",unsafe_allow_html=True)

            st.markdown('<div class="section-header">Tabla Principal</div>',unsafe_allow_html=True)
            cols_t=tks_ok+["Portafolio"]
            rows_t={"W (%)":         {t:(f"{pesos_norm[i]*100:.1f}%" if i<len(pesos_norm) else "N/A") for i,t in enumerate(tks_ok)}|{"Portafolio":"100.0%"},
                    "Retorno Esp.":  {t:f"{rets_e[t]:.2%}" for t in tks_ok}|{"Portafolio":f"{ret_p:.2%}"},
                    "Volatilidad":   {t:f"{vol_ind[t]:.2%}" for t in tks_ok}|{"Portafolio":f"{vol_p:.2%}"},
                    "Sharpe":        {t:f"{sh_ind[t]:.3f}" for t in tks_ok}|{"Portafolio":f"{sh_p:.3f}"},
                    "Beta":          {t:f"{betas_s.get(t,1.):.4f}" for t in tks_ok}|{"Portafolio":"—"},
                    "IR":            {t:"—" for t in tks_ok}|{"Portafolio":f"{ir:.3f}" if not np.isnan(ir) else "N/A"}}
            def css_v(v,m):
                if "%" in str(v):
                    n=float(v.replace("%",""))/100
                    return "td-positive" if (m=="Retorno Esp." and n>0) else ("td-negative" if (m=="Retorno Esp." and n<0) else ("td-blue" if m=="W (%)" else ""))
                try:
                    n=float(v)
                    if m=="Sharpe": return "td-positive" if n>1 else ("td-negative" if n<0 else "td-neutral")
                    if m=="IR":     return "td-positive" if n>0 else "td-negative"
                except: pass
                return ""
            hdr_t="".join(f"<th>{c}</th>" for c in cols_t)
            filas="".join(f"<tr><td class='td-label'>{m}</td>"+"".join(
                f'<td class="{css_v(rows_t[m].get(c,"—"),m)}">{rows_t[m].get(c,"—")}</td>'
                for c in cols_t)+"</tr>" for m in rows_t)
            st.markdown(f'<table class="styled-table"><thead><tr><th>Métrica</th>{hdr_t}</tr></thead><tbody>{filas}</tbody></table>',unsafe_allow_html=True)

            st.markdown('<div class="section-header">Matriz de Correlación</div>',unsafe_allow_html=True)
            corr=rend.corr()
            fc=go.Figure(go.Heatmap(z=corr.values,x=corr.columns.tolist(),y=corr.index.tolist(),
                colorscale=[[0,"#E74C3C"],[.5,"#0D1E35"],[1,"#2ECC71"]],zmid=0,zmin=-1,zmax=1,
                text=np.round(corr.values,2),texttemplate="%{text}",textfont=dict(size=11,family="IBM Plex Mono")))
            lo_corr=base_layout("Correlación de Rendimientos Diarios",350)
            lo_corr["yaxis"]["autorange"]="reversed"
            fc.update_layout(**lo_corr)
            st.plotly_chart(fc,use_container_width=True)

            st.markdown('<div class="section-header">Escenarios Óptimos — Los 3 Modelos</div>',unsafe_allow_html=True)
            st.caption("Optimización: Σ W = 100%, W ≥ 0  ·  Los 3 modelos siempre visibles")
            escenarios=[("Mín. Varianza","varianza"),("Máx. IR","ir"),("Máx. Sharpe","sharpe"),("Máx. Retorno","retorno")]

            def render_escenarios(mod_n, rm_calc):
                rows_e=[]; pw={}
                for ne,oe in escenarios:
                    wo=optimizar_portafolio(rm_calc,cov_an,rf_port,oe)
                    re_=float(np.dot(wo,rm_calc)); ve_=vol_portafolio(wo,cov_an)
                    se_=(re_-rf_port)/ve_ if ve_>0 else np.nan
                    rpd_=(rend*wo).sum(axis=1)
                    te_=tracking_error(rpd_,rend_bench) if bench_ok else np.nan
                    ie_=information_ratio(re_,ret_bi,te_)
                    rows_e.append({"Escenario":ne,"Retorno":f"{re_:.2%}","Riesgo":f"{ve_:.2%}",
                                   "Sharpe":f"{se_:.3f}","IR":f"{ie_:.3f}" if not np.isnan(ie_) else "N/A"})
                    pw[ne]=wo
                eh='<table class="styled-table"><thead><tr><th>Escenario</th><th>Retorno</th><th>Riesgo</th><th>Sharpe</th><th>IR</th>'
                eh+="".join(f"<th>W {t}</th>" for t in tks_ok)+"</tr></thead><tbody>"
                for row in rows_e:
                    esc=row["Escenario"]; w=pw[esc]
                    rn_=float(row["Retorno"].replace("%",""))/100; sn_=float(row["Sharpe"])
                    cr="td-positive" if rn_>0 else "td-negative"
                    cs="td-positive" if sn_>1 else ("td-negative" if sn_<0 else "td-neutral")
                    eh+=f'<tr><td class="td-label">{esc}</td><td class="{cr}">{row["Retorno"]}</td><td>{row["Riesgo"]}</td><td class="{cs}">{row["Sharpe"]}</td><td>{row["IR"]}</td>'
                    eh+="".join(f'<td class="td-blue">{wi*100:.1f}%</td>' for wi in w)+"</tr>"
                eh+="</tbody></table>"
                st.markdown(eh,unsafe_allow_html=True)
                fw=go.Figure()
                for i,(en,we) in enumerate(pw.items()):
                    fw.add_trace(go.Bar(name=en,x=tks_ok,y=[wi*100 for wi in we],
                        marker_color=PALETTE[i%len(PALETTE)],
                        hovertemplate=f"<b>{en}</b><br>%{{x}}: %{{y:.1f}}%<extra></extra>"))
                lo_w=base_layout(f"Distribución de Pesos — {mod_n}",320)
                lo_w["barmode"]="group"; lo_w["yaxis"]["title"]="Peso (%)"
                fw.update_layout(**lo_w); st.plotly_chart(fw,use_container_width=True)
                return pw

            tabs_m=st.tabs(["📐 Markowitz","📐 CAPM","📐 Montecarlo"])
            with tabs_m[0]:
                rm_mk=retorno_markowitz(rend)
                pw_mk=render_escenarios("Markowitz",rm_mk)
            with tabs_m[1]:
                rm_cp,_=retorno_capm(rend,rend_bench,rf_port) if bench_ok else (retorno_markowitz(rend),None)
                pw_cp=render_escenarios("CAPM",rm_cp)
            with tabs_m[2]:
                with st.spinner("Ejecutando Montecarlo (500 simulaciones)..."):
                    rm_mc=retorno_montecarlo(rend,500)
                pw_mc=render_escenarios("Montecarlo",rm_mc)
            st.markdown('<div class="section-header">Frontera Eficiente (Markowitz)</div>',unsafe_allow_html=True)
            with st.spinner("Calculando frontera..."):
                rmk=retorno_markowitz(rend); rmin=float(rmk.min()); rmax=float(rmk.max())
                frets=[]; fvols=[]
                for tgt in np.linspace(rmin,rmax,60):
                    cons_=[{"type":"eq","fun":lambda w:np.sum(w)-1.},
                           {"type":"eq","fun":lambda w,t=tgt:float(np.dot(w,rmk))-t}]
                    res_=minimize(lambda w:vol_portafolio(w,cov_an),np.ones(len(tks_ok))/len(tks_ok),
                                  method="SLSQP",bounds=[(0,1)]*len(tks_ok),constraints=cons_,options={"maxiter":500})
                    if res_.success: frets.append(tgt); fvols.append(vol_portafolio(res_.x,cov_an))
            ffe=go.Figure()
            ffe.add_trace(go.Scatter(x=[v*100 for v in fvols],y=[r*100 for r in frets],mode="lines",
                name="Frontera",line=dict(color="#4AB3F4",width=2.5),
                hovertemplate="Vol: %{x:.2f}%<br>Ret: %{y:.2f}%<extra></extra>"))
            for i,t in enumerate(tks_ok):
                ffe.add_trace(go.Scatter(x=[float(vol_ind[t])*100],y=[float(rmk[t])*100],
                    mode="markers+text",name=t,text=[t],textposition="top center",
                    textfont=dict(size=10,family="IBM Plex Mono"),
                    marker=dict(size=10,color=PALETTE[i%len(PALETTE)],line=dict(width=1,color="#0A1628"))))
            ffe.add_trace(go.Scatter(x=[vol_p*100],y=[ret_p*100],mode="markers",name="Portafolio actual",
                marker=dict(size=14,color="#F39C12",symbol="star",line=dict(width=1,color="#0A1628")),
                hovertemplate="Portafolio actual<br>Vol: %{x:.2f}%<br>Ret: %{y:.2f}%<extra></extra>"))
            lo_fe=base_layout("Frontera Eficiente de Markowitz",450)
            lo_fe["xaxis"]["title"]="Volatilidad (%)"; lo_fe["yaxis"]["title"]="Retorno Esperado (%)"
            ffe.update_layout(**lo_fe); st.plotly_chart(ffe,use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# BLOQUES 3-6
# ══════════════════════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 3 — VALORACIÓN ESTADÍSTICA
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    if "precios" not in st.session_state:
        st.info("⬅️  Primero carga los datos en **01 · Análisis Básico**")
    else:
        precios_ve  = st.session_state["precios"]
        tks_ve      = st.session_state["tickers_list"]
        bench_ve    = st.session_state["bench"]
        tks_ok_ve   = [t for t in tks_ve if t in precios_ve.columns]
        bench_ok_ve = bench_ve in precios_ve.columns

        with st.expander("⚙️  CONFIGURACIÓN VALORACIÓN ESTADÍSTICA", expanded=True):
            cve1, cve2 = st.columns([2,2])
            with cve1:
                ve_c1, ve_c2 = st.columns([1,1])
                with ve_c1:
                    ve_n = st.number_input("Cantidad", min_value=1, max_value=999, value=1, step=1, key="ve_n")
                with ve_c2:
                    ve_u = st.selectbox("Unidad", ["Días","Meses","Años","Todo histórico"],
                                        index=2, key="ve_u", label_visibility="visible")
                def ve_a_dias(n,u):
                    if u=="Todo histórico": return 99999
                    return int(n*365) if u=="Años" else (int(n*30.44) if u=="Meses" else int(n))
                def ve_label(n,u): return "Todo el histórico" if u=="Todo histórico" else f"{n} {u}"
                rango_ve = ve_label(ve_n, ve_u)
            cve3, cve4 = st.columns(2)
            with cve3:
                st.markdown('<span style="font-family:IBM Plex Mono,monospace;font-size:.66rem;color:#E8F4FD;letter-spacing:.1em;text-transform:uppercase;">Peso Regresión Lineal (%)</span>', unsafe_allow_html=True)
                peso_reg = st.slider("", 0, 100, 60, 5, key="peso_reg", label_visibility="collapsed")
            with cve4:
                peso_pct = 100 - peso_reg
                st.markdown(f'<div style="padding:.6rem 0"><div class="kpi-label">Peso Percentiles</div><div style="font-family:IBM Plex Mono,monospace;font-size:1.3rem;font-weight:600;color:#4AB3F4;">{peso_pct}%</div></div>', unsafe_allow_html=True)

        btn_ve = st.button("📐  CALCULAR TODOS LOS ACTIVOS", key="btn_ve")

        dias_ve = ve_a_dias(ve_n, ve_u)
        ff_ve   = pd.Timestamp(precios_ve.index[-1])
        fi_ve   = precios_ve.index[0] if dias_ve>=99999 else fecha_anterior_cercana(
                      precios_ve.index, ff_ve - pd.Timedelta(days=dias_ve))

        if btn_ve or any(k.startswith("ve_calc_") for k in st.session_state):
            if btn_ve:
                for t in tks_ok_ve:
                    st.session_state[f"ve_calc_{t}"] = True

            from scipy import stats as sp_stats
            resultados_ve = {}
            for ticker_sel in tks_ok_ve:
                mask_ve = (precios_ve.index>=fi_ve)&(precios_ve.index<=ff_ve)
                p_rango = precios_ve.loc[mask_ve, ticker_sel].dropna()
                b_rango = precios_ve.loc[mask_ve, bench_ve].dropna() if bench_ok_ve else None
                if len(p_rango)<10: continue
                ln_accion = np.log(p_rango)
                ult_precio    = float(p_rango.iloc[-1])
                ult_precio_ln = float(ln_accion.iloc[-1])
                ult_bench=ult_bench_ln=None
                slope=intercept=correl=r2=dv20=pm20=cv=0.
                val_est=ult_precio; pot_reg=0.; estado=conf="N/A"; val_serie=None
                x_arr=y_arr=None
                if b_rango is not None and len(b_rango)>10:
                    ln_bench = np.log(b_rango)
                    ult_bench=float(b_rango.iloc[-1]); ult_bench_ln=float(ln_bench.iloc[-1])
                    df_ln = pd.concat([ln_accion,ln_bench],axis=1).dropna()
                    y_arr=df_ln.iloc[:,0].values; x_arr=df_ln.iloc[:,1].values
                    slope,intercept,r_val,_,_ = sp_stats.linregress(x_arr,y_arr)
                    r2=r_val**2; correl=float(np.corrcoef(y_arr,x_arr)[0,1])
                    val_est=float(np.exp(intercept+slope*ult_bench_ln))
                    pot_reg=(val_est/ult_precio)-1
                    estado="🔴 Sobrevalorada" if pot_reg<0 else "🟢 Subvalorada"
                    conf="✅ Confiable" if r2>0.75 else "⚠️ No confiable"
                    ln20=ln_accion.iloc[-20:]; dv20=float(ln20.std()); pm20=float(ln20.mean())
                    cv=dv20/pm20 if pm20!=0 else np.nan
                    val_serie=np.exp(intercept+slope*np.log(b_rango))
                p_arr=p_rango.values; p50=float(np.percentile(p_arr,50))
                p_pcts={f"{n}%":np.percentile(p_arr,n) for n in [0,1,5,25,50,75,95,99,100]}
                pot_pct=(p50/ult_precio)-1
                pct_act=(np.sum(p_arr<=ult_precio)/len(p_arr))*100
                w_reg=peso_reg/100; w_pct=peso_pct/100
                pot_pond=pot_reg*w_reg+pot_pct*w_pct; p_obj=ult_precio*(1+pot_pond)
                señal=("🟢 COMPRAR" if pot_pond>0.10 else "🟡 POSIBLE COMPRA" if pot_pond>0.03
                       else "🔴 VENDER" if pot_pond<-0.10 else "⚠️ POSIBLE VENTA" if pot_pond<-0.03
                       else "⚪ MANTENER")
                resultados_ve[ticker_sel]=dict(
                    p_rango=p_rango,b_rango=b_rango,ln_accion=ln_accion,
                    ult_precio=ult_precio,ult_precio_ln=ult_precio_ln,
                    ult_bench=ult_bench,ult_bench_ln=ult_bench_ln,
                    slope=slope,intercept=intercept,correl=correl,r2=r2,
                    val_est=val_est,pot_reg=pot_reg,estado=estado,conf=conf,
                    dv20=dv20,pm20=pm20,cv=cv,val_serie=val_serie,
                    p_pcts=p_pcts,p50=p50,pot_pct=pot_pct,pct_act=pct_act,
                    pot_pond=pot_pond,p_obj=p_obj,señal=señal,
                    x_arr=x_arr,y_arr=y_arr)

            if not resultados_ve:
                st.warning("No hay suficientes datos."); st.stop()

            fi_fmt_ve=fmt_fecha(fi_ve); ff_fmt_ve=fmt_fecha(ff_ve)

            # Resumen comparativo
            st.markdown('<div class="section-header">Resumen Comparativo</div>', unsafe_allow_html=True)
            sh="<th>Activo</th><th>Últ. Precio</th><th>Val. Estadística</th><th>Pot. Reg.</th><th>Estado</th><th>P50</th><th>Pot. P50</th><th>Pot. Ponderado</th><th>P. Objetivo</th><th>Señal</th>"
            sr=""
            for t,r in resultados_ve.items():
                cp="td-positive" if r["pot_pond"]>0 else "td-negative"
                cr="td-positive" if r["pot_reg"]>0 else "td-negative"
                cpc="td-positive" if r["pot_pct"]>0 else "td-negative"
                sr+=(f'<tr><td class="td-blue">{t}</td>'
                     f'<td>{fmt_num(r["ult_precio"],2,es_usd=True)}</td>'
                     f'<td>{fmt_num(r["val_est"],2,es_usd=True)}</td>'
                     f'<td class="{cr}">{fmt_num(r["pot_reg"],2,es_pct=True)}</td>'
                     f'<td>{r["estado"]}</td>'
                     f'<td>{fmt_num(r["p50"],2,es_usd=True)}</td>'
                     f'<td class="{cpc}">{fmt_num(r["pot_pct"],2,es_pct=True)}</td>'
                     f'<td class="{cp}">{fmt_num(r["pot_pond"],2,es_pct=True)}</td>'
                     f'<td>{fmt_num(r["p_obj"],2,es_usd=True)}</td>'
                     f'<td>{r["señal"]}</td></tr>')
            st.markdown(f'<table class="styled-table"><thead><tr>{sh}</tr></thead><tbody>{sr}</tbody></table>', unsafe_allow_html=True)

            # Pestañas por activo
            st.markdown('<div class="section-header">Detalle por Activo</div>', unsafe_allow_html=True)
            tabs_t=st.tabs([f"📊 {t}" for t in resultados_ve])
            for tab_tk,(tkr,res) in zip(tabs_t,resultados_ve.items()):
                with tab_tk:
                    up=res["ult_precio"]; ve_=res["val_est"]; pr=res["pot_reg"]
                    pp=res["pot_pct"]; pond=res["pot_pond"]; po=res["p_obj"]
                    p50=res["p50"]; pa=res["pct_act"]; r2=res["r2"]
                    conf=res["conf"]; est=res["estado"]
                    # KPI
                    st.markdown(f"""<div class="kpi-grid">
                      <div class="kpi-card"><div class="kpi-label">Activo</div>
                        <div class="kpi-value" style="font-size:1rem">{tkr}</div></div>
                      <div class="kpi-card"><div class="kpi-label">Último Precio</div>
                        <div class="kpi-value">{fmt_num(up,2,es_usd=True)}</div>
                        <div class="kpi-delta">{ff_fmt_ve}</div></div>
                      <div class="kpi-card"><div class="kpi-label">Precio Objetivo</div>
                        <div class="kpi-value {'positive' if po>up else 'negative'}">{fmt_num(po,2,es_usd=True)}</div></div>
                      <div class="kpi-card"><div class="kpi-label">Potencial Ponderado</div>
                        <div class="kpi-value {'positive' if pond>0 else 'negative'}">{fmt_num(pond,2,es_pct=True)}</div></div>
                      <div class="kpi-card"><div class="kpi-label">Señal</div>
                        <div class="kpi-value" style="font-size:.8rem">{res["señal"]}</div></div>
                      <div class="kpi-card"><div class="kpi-label">Confiabilidad</div>
                        <div class="kpi-value" style="font-size:.8rem">{conf}</div>
                        <div class="kpi-delta">R²={fmt_num(r2,4)}</div></div>
                    </div>""", unsafe_allow_html=True)

                    col_l,col_r=st.columns([1,2])
                    with col_l:
                        bname=bench_ve if bench_ok_ve else "N/A"
                        ub=fmt_num(res["ult_bench"],2,es_usd=True) if res["ult_bench"] else "N/A"
                        ubl=fmt_num(res["ult_bench_ln"],4) if res["ult_bench_ln"] else "N/A"
                        st.markdown(f"""<p style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#4AB3F4;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.35rem;">Variables</p>
                        <table class="styled-table"><thead><tr><th>Var</th><th>Nombre</th><th>Precio</th><th>LN</th></tr></thead><tbody>
                        <tr><td>Y</td><td class="td-blue">{tkr}</td><td>{fmt_num(up,2,es_usd=True)}</td><td>{fmt_num(res["ult_precio_ln"],4)}</td></tr>
                        <tr><td>X</td><td class="td-blue">{bname}</td><td>{ub}</td><td>{ubl}</td></tr>
                        </tbody></table>""", unsafe_allow_html=True)
                        st.markdown("<br>", unsafe_allow_html=True)
                        cp_r="td-positive" if pr>0 else "td-negative"
                        cp_p="td-positive" if pp>0 else "td-negative"
                        cp_t="td-positive" if pond>0 else "td-negative"
                        st.markdown(f"""<p style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#4AB3F4;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.35rem;">Recomendación Ponderada</p>
                        <table class="styled-table"><thead><tr><th>Método</th><th>P.Obj</th><th>Potencial</th><th>Peso</th></tr></thead><tbody>
                        <tr><td>Regresión</td><td>{fmt_num(ve_,2,es_usd=True)}</td><td class="{cp_r}">{fmt_num(pr,2,es_pct=True)}</td><td class="td-blue">{peso_reg}%</td></tr>
                        <tr><td>P50</td><td>{fmt_num(p50,2,es_usd=True)}</td><td class="{cp_p}">{fmt_num(pp,2,es_pct=True)}</td><td class="td-blue">{peso_pct}%</td></tr>
                        <tr style="border-top:2px solid #4AB3F4"><td style="font-weight:600;color:#E8F4FD">⭐ Objetivo</td>
                        <td style="font-weight:600">{fmt_num(po,2,es_usd=True)}</td>
                        <td class="{cp_t}" style="font-weight:600">{fmt_num(pond,2,es_pct=True)}</td>
                        <td class="td-blue">100%</td></tr>
                        </tbody></table>""", unsafe_allow_html=True)

                    with col_r:
                        def _css(v,tp):
                            if tp=="Pot": return "td-positive" if v>0 else "td-negative"
                            if tp=="R2":  return "td-positive" if v>0.75 else "td-neutral"
                            return ""
                        cv_str=fmt_num(res["cv"],2,es_pct=True) if not np.isnan(res["cv"]) else "N/A"
                        stats=[
                            ("Factor Correlación",fmt_num(res["correl"],4),res["correl"],""),
                            ("Alfa",fmt_num(res["intercept"],4),res["intercept"],""),
                            ("Beta",fmt_num(res["slope"],4),res["slope"],""),
                            ("Coef. R²",fmt_num(r2,4),r2,"R2"),
                            ("Confiabilidad",conf,None,""),
                            ("Val. Estadística",fmt_num(ve_,2,es_usd=True),ve_,""),
                            ("Potencial Reg.",fmt_num(pr,2,es_pct=True),pr,"Pot"),
                            ("Estado",est,None,""),
                            ("Desvest 20d",fmt_num(res["dv20"],2),None,""),
                            ("Promedio 20d",fmt_num(res["pm20"],2),None,""),
                            ("Coef. Variación",cv_str,None,""),
                        ]
                        sh2='<p style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#4AB3F4;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.35rem;">Estadísticos — '+rango_ve+'</p>'
                        sh2+='<table class="styled-table"><thead><tr><th>Estadístico</th><th>Valor</th></tr></thead><tbody>'
                        for lb,vs,ns,ts in stats:
                            cs=_css(ns,ts) if ns is not None else ""
                            sh2+='<tr><td>'+lb+'</td><td class="'+cs+'">'+vs+'</td></tr>'
                        sh2+="</tbody></table>"
                        st.markdown(sh2, unsafe_allow_html=True)

                    # Gráfica regresión
                    if res["x_arr"] is not None:
                        fig_r=go.Figure()
                        fig_r.add_trace(go.Scatter(x=res["x_arr"],y=res["y_arr"],mode="markers",
                            name="Obs.",marker=dict(color="#4AB3F4",size=4,opacity=0.5)))
                        xl=np.linspace(res["x_arr"].min(),res["x_arr"].max(),100)
                        fig_r.add_trace(go.Scatter(x=xl,y=res["intercept"]+res["slope"]*xl,
                            mode="lines",name="Regresión",line=dict(color="#F39C12",width=2)))
                        lo_r=base_layout(f"Regresión — {tkr} vs {bench_ve}",350)
                        lo_r["xaxis"]["title"]=f"LN({bench_ve})"; lo_r["yaxis"]["title"]=f"LN({tkr})"
                        fig_r.update_layout(**lo_r); st.plotly_chart(fig_r,use_container_width=True)

                    # Percentiles + gráfica precio
                    cp1,cp2=st.columns([1,2])
                    with cp1:
                        ph='<p style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#4AB3F4;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.35rem;">Percentiles de Precios</p>'
                        ph+='<table class="styled-table"><thead><tr><th>Percentil</th><th>Precio</th></tr></thead><tbody>'
                        for k,v in res["p_pcts"].items():
                            ph+='<tr><td class="td-label">'+k+'</td><td>'+fmt_num(v,2,es_usd=True)+'</td></tr>'
                        cpa="td-positive" if pa<30 else ("td-negative" if pa>70 else "td-neutral")
                        cpp="td-positive" if pp>0 else "td-negative"
                        ph+=(f'<tr><td class="td-label">Percentil actual</td><td class="{cpa}">{fmt_num(pa/100,1,es_pct=True)}</td></tr>'
                             f'<tr><td class="td-label">Pot. P50</td><td class="{cpp}">{fmt_num(pp,2,es_pct=True)}</td></tr>')
                        ph+="</tbody></table>"
                        st.markdown(ph, unsafe_allow_html=True)
                    with cp2:
                        fig_e=go.Figure()
                        fig_e.add_trace(go.Scatter(x=res["p_rango"].index,y=res["p_rango"].values,
                            name=tkr,line=dict(color="#4AB3F4",width=1.8),
                            hovertemplate="%{x|%d/%m/%y}<br>$%{y:.2f}<extra></extra>"))
                        if res["val_serie"] is not None:
                            fig_e.add_trace(go.Scatter(x=res["val_serie"].index,y=res["val_serie"].values,
                                name="Val. estadística",line=dict(color="#F39C12",width=1.5,dash="dash")))
                        fig_e.add_hline(y=p50,line_color="#2ECC71",line_dash="dot",line_width=1.5,
                            annotation_text=f"P50: {fmt_num(p50,2,es_usd=True)}",annotation_font_color="#2ECC71")
                        fig_e.add_hline(y=po,line_color="#9B59B6",line_dash="dash",line_width=1.5,
                            annotation_text=f"Objetivo: {fmt_num(po,2,es_usd=True)}",annotation_font_color="#9B59B6")
                        lo_e=base_layout(f"Precio {tkr} + Valoración + Objetivo",380)
                        lo_e["xaxis"]["rangeslider"]=dict(visible=True,bgcolor="#0A1628",bordercolor="#1A3A5C")
                        fig_e.update_layout(**lo_e); st.plotly_chart(fig_e,use_container_width=True)

        else:
            st.markdown("""<div style="text-align:center;padding:4rem 2rem;color:#4A7FA5;font-family:IBM Plex Mono,monospace;">
              <div style="font-size:3rem;margin-bottom:1rem;">📐</div>
              <div style="font-size:.85rem;letter-spacing:.1em;text-transform:uppercase;">
                Selecciona el rango y presiona<br>
                <span style="color:#4AB3F4;">CALCULAR TODOS LOS ACTIVOS</span></div>
            </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 4 — VALORACIÓN TÉCNICA
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    if "precios" not in st.session_state:
        st.info("⬅️  Primero carga los datos en **01 · Análisis Básico**")
    else:
        precios_vt  = st.session_state["precios"]
        tks_vt      = st.session_state["tickers_list"]
        tks_ok_vt   = [t for t in tks_vt if t in precios_vt.columns]

        # ── CONFIGURACIÓN ─────────────────────────────────────────────────────
        with st.expander("⚙️  PARÁMETROS TÉCNICOS", expanded=True):
            st.markdown('<div class="section-header">Indicadores y Pesos</div>', unsafe_allow_html=True)
            pc1,pc2,pc3,pc4 = st.columns(4)
            with pc1:
                st.markdown('<span style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#E8F4FD;letter-spacing:.1em;text-transform:uppercase;">Medias Móviles</span>', unsafe_allow_html=True)
                mm1 = st.number_input("MM1", 1, 500, 5,   key="mm1")
                mm2 = st.number_input("MM2", 1, 500, 20,  key="mm2")
                mm3 = st.number_input("MM3", 1, 500, 100, key="mm3")
                mm4 = st.number_input("MM4", 1, 500, 200, key="mm4")
                peso_mm = st.number_input("Peso MM (%)", 0.0, 100.0, 16.67, 0.01, key="peso_mm", format="%.2f")
            with pc2:
                st.markdown('<span style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#E8F4FD;letter-spacing:.1em;text-transform:uppercase;">MACD</span>', unsafe_allow_html=True)
                macd_fast  = st.number_input("MACD Rápido", 1, 100, 12, key="macd_fast")
                macd_slow  = st.number_input("MACD Lento",  1, 200, 26, key="macd_slow")
                macd_sig   = st.number_input("Signal",      1, 100, 9,  key="macd_sig")
                peso_macd  = st.number_input("Peso MACD (%)", 0.0, 100.0, 16.67, 0.01, key="peso_macd", format="%.2f")
            with pc3:
                st.markdown('<span style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#E8F4FD;letter-spacing:.1em;text-transform:uppercase;">RSI</span>', unsafe_allow_html=True)
                rsi_per    = st.number_input("Período RSI", 1, 100, 14, key="rsi_per")
                rsi_sob    = st.number_input("Sobrecompra", 1, 100, 80, key="rsi_sob")
                rsi_sov    = st.number_input("Sobreventa",  1, 100, 20, key="rsi_sov")
                peso_rsi   = st.number_input("Peso RSI (%)", 0.0, 100.0, 50.0, 0.01, key="peso_rsi", format="%.2f")
            with pc4:
                st.markdown('<span style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#E8F4FD;letter-spacing:.1em;text-transform:uppercase;">Fibonacci & Decisión</span>', unsafe_allow_html=True)
                fibo_sens   = st.number_input("Pivote Fibo (días vecinos)", 2, 60, 10, 1, key="fibo_sens")
                peso_fibo   = st.number_input("Peso Fibo (%)", 0.0, 100.0, 16.67, 0.01, key="peso_fibo", format="%.2f")
                st.markdown("---")
                peso_dec_max = st.number_input("Umbral COMPRAR/VENDER (%)", 0.0, 100.0, 75.0, 1.0, key="peso_dec_max", format="%.1f")
                peso_dec_min = st.number_input("Umbral POSIBLE (%)",        0.0, 100.0, 50.0, 1.0, key="peso_dec_min", format="%.1f")

            st.markdown('<div class="section-header">Rango de Análisis</div>', unsafe_allow_html=True)
            vt_r1, vt_r2, vt_r3 = st.columns([2,1,1])
            with vt_r1:
                rango_vt = st.selectbox("Rango predefinido",
                    ["1 Mes","3 Meses","6 Meses","1 Año","2 Años","3 Años","5 Años","Todo"],
                    index=3, key="rango_vt")
            with vt_r2:
                usar_rango_custom = st.checkbox("Rango personalizado", value=False, key="vt_custom")
            with vt_r3:
                st.markdown("<br>", unsafe_allow_html=True)
                btn_vt = st.button("📡  CALCULAR ANÁLISIS TÉCNICO", key="btn_vt")
            if usar_rango_custom:
                vt_fc1, vt_fc2 = st.columns(2)
                with vt_fc1:
                    vt_fi_custom = st.date_input("Fecha inicio", value=datetime(2023,1,1),
                                                  min_value=datetime(1900,1,1),
                                                  max_value=datetime.today(), key="vt_fi")
                with vt_fc2:
                    vt_ff_custom = st.date_input("Fecha fin", value=datetime.today(),
                                                  min_value=datetime(1900,1,1),
                                                  max_value=datetime.today(), key="vt_ff")

        # ── FUNCIONES DE CÁLCULO ──────────────────────────────────────────────
        def calc_ema(serie, period):
            """Promedio móvil simple de `period` períodos. NaN en primeras period-1 filas."""
            return serie.rolling(window=period).mean()

        def calc_rsi(serie, period):
            """RSI usando suma rolling de period períodos."""
            delta   = serie.pct_change()
            alzas_f = delta.apply(lambda x: x if (pd.notna(x) and x>0) else 0.0)
            bajas_f = delta.apply(lambda x: x if (pd.notna(x) and x<0) else 0.0)
            sum_a   = alzas_f.rolling(window=period, min_periods=period).sum()
            sum_b   = bajas_f.rolling(window=period, min_periods=period).sum()
            rs      = sum_a / sum_b.abs().replace(0, np.nan)
            return 100 - (100 / (1 + rs))

        def calc_macd(serie, fast, slow, signal):
            ema_f  = serie.rolling(window=fast).mean()   # NaN primeras fast-1 filas
            ema_s  = serie.rolling(window=slow).mean()   # NaN primeras slow-1 filas
            macd   = ema_f - ema_s                        # NaN primeras slow-1 filas
            sig    = macd.rolling(window=signal).mean()  # NaN primeras slow+signal-2 filas
            hist   = macd - sig
            return macd, sig, hist

        def calc_fibonacci(precios_serie, n_vecinos):
            """
            Detecta pivotes locales (zigzag):
            - Pivote máximo: precio mayor que sus n_vecinos días a cada lado
            - Pivote mínimo: precio menor que sus n_vecinos días a cada lado
            Toma el último pivote máximo y el último pivote mínimo.
            Tendencia: si el último mínimo es más antiguo que el último máximo → BAJISTA
                       si el último máximo es más antiguo que el último mínimo → ALCISTA
            """
            ratios = [0.236, 0.382, 0.500, 0.618, 1.000]
            labels = ["Fibo 1 (23,6%)","Fibo 2 (38,2%)","Fibo 3 (50,0%)","Fibo 4 (61,8%)","Fibo 5 (100%)"]
            n = max(int(n_vecinos), 2)
            vals = precios_serie.values
            idx_arr = precios_serie.index

            # Detectar pivotes
            pivot_max_idx = None; pivot_min_idx = None
            for i in range(n, len(vals)-n):
                ventana = vals[i-n:i+n+1]
                if vals[i] == ventana.max():
                    pivot_max_idx = i
                if vals[i] == ventana.min():
                    pivot_min_idx = i

            # Fallback si no hay pivotes suficientes
            if pivot_max_idx is None: pivot_max_idx = int(np.argmax(vals))
            if pivot_min_idx is None: pivot_min_idx = int(np.argmin(vals))

            maximo = float(vals[pivot_max_idx])
            minimo = float(vals[pivot_min_idx])
            fecha_max = idx_arr[pivot_max_idx]
            fecha_min = idx_arr[pivot_min_idx]

            # Tendencia: el pivote más reciente define la dirección actual
            tendencia = "BAJISTA" if fecha_max > fecha_min else "ALCISTA"

            rango = maximo - minimo
            if rango == 0: rango = maximo * 0.01  # evitar div/0

            alza = {lb: maximo - r*rango for lb,r in zip(labels,ratios)}
            baja = {lb: minimo + r*rango for lb,r in zip(labels,ratios)}
            return alza, baja, maximo, minimo, tendencia, fecha_max, fecha_min

        def señal_mm_multi(precios_s, p1, p2, p3, p4):
            """
            Señal usando alineación de 4 medias móviles + Golden/Death Cross.
            p1<p2<p3<p4 (ej: 5,20,100,200)
            """
            min_len = p4 + 2
            if len(precios_s) < min_len:
                return "MANTENER", 0, 0

            mm = {p: precios_s.rolling(p).mean() for p in [p1,p2,p3,p4]}
            # Valores actuales y anteriores
            ult  = float(precios_s.iloc[-1]); ant = float(precios_s.iloc[-2])
            m1u  = mm[p1].iloc[-1]; m1a = mm[p1].iloc[-2]
            m2u  = mm[p2].iloc[-1]; m2a = mm[p2].iloc[-2]
            m3u  = mm[p3].iloc[-1] if not pd.isna(mm[p3].iloc[-1]) else m2u
            m4u  = mm[p4].iloc[-1] if not pd.isna(mm[p4].iloc[-1]) else m3u

            if any(pd.isna(v) for v in [m1u,m2u]): return "MANTENER", 0, 0

            # Golden Cross: MM corta cruza por encima de MM larga
            golden_cross = (m1u > m2u and m1a <= m2a)
            death_cross  = (m1u < m2u and m1a >= m2a)
            gc_medio     = (m2u > m3u and mm[p2].iloc[-2] <= mm[p3].iloc[-2]) if not pd.isna(m3u) else False
            dc_medio     = (m2u < m3u and mm[p2].iloc[-2] >= mm[p3].iloc[-2]) if not pd.isna(m3u) else False

            # Alineación completa alcista: MM5>MM20>MM100>MM200
            alin_alcista = m1u > m2u > m3u and ult > m2u
            alin_bajista = m1u < m2u < m3u and ult < m2u

            if alin_alcista or golden_cross or gc_medio:
                return "COMPRAR", 1, 0
            if alin_bajista or death_cross or dc_medio:
                return "VENDER", 0, 1
            if ult > m2u:
                return "ALCISTA", 0, 0
            if ult < m2u:
                return "BAJISTA", 0, 0
            return "MANTENER", 0, 0

        def señal_rsi(rsi_val, sob, sov):
            if rsi_val < sov:  return "COMPRAR", 1, 0
            if rsi_val > sob:  return "VENDER",  0, 1
            return "NEUTRAL", 0, 0

        def señal_macd_fn(macd_s, sig_s):
            m_u=macd_s.iloc[-1]; m_a=macd_s.iloc[-2]
            s_u=sig_s.iloc[-1];  s_a=sig_s.iloc[-2]
            if m_u>s_u and m_a<=s_a: return "COMPRAR", 1, 0
            if m_u<s_u and m_a>=s_a: return "VENDER",  0, 1
            if m_u>s_u:               return "ALCISTA", 0, 0
            return "BAJISTA", 0, 0

        def señal_fibo(precio_ult, alza, baja, tendencia):
            """
            Señal basada en zona de Fibonacci donde se encuentra el precio:
            ALCISTA: precio sobre Fibo1 alza → COMPRAR (retroceso leve)
                     precio entre Fibo1 y Fibo3 alza → ALCISTA (retroceso moderado)
                     precio entre Fibo3 y Fibo4 alza → NEUTRAL (zona crítica)
                     precio bajo Fibo4 alza → VENDER (retroceso profundo)
            BAJISTA: lógica inversa con niveles baja
            """
            niveles_alza = list(alza.values())  # [Fibo1..Fibo5] de mayor a menor
            niveles_baja = list(baja.values())  # [Fibo1..Fibo5] de menor a mayor

            if tendencia == "ALCISTA":
                if precio_ult >= niveles_alza[0]:                              # sobre Fibo1
                    return "COMPRAR", 1, 0
                elif precio_ult >= niveles_alza[2]:                            # entre Fibo1 y Fibo3
                    return "ALCISTA", 0, 0
                elif precio_ult >= niveles_alza[3]:                            # entre Fibo3 y Fibo4
                    return "NEUTRAL", 0, 0
                else:                                                           # bajo Fibo4
                    return "VENDER", 0, 1
            else:  # BAJISTA
                if precio_ult <= niveles_baja[0]:                              # bajo Fibo1 baja
                    return "VENDER", 0, 1
                elif precio_ult <= niveles_baja[2]:                            # entre Fibo1 y Fibo3 baja
                    return "BAJISTA", 0, 0
                elif precio_ult <= niveles_baja[3]:                            # entre Fibo3 y Fibo4 baja
                    return "NEUTRAL", 0, 0
                else:                                                           # sobre Fibo4 baja
                    return "COMPRAR", 1, 0


        def decision_final(peso_compra, peso_venta, umbral_max, umbral_min):
            if peso_compra >= umbral_max:   return "✅ COMPRAR"
            if peso_venta  >= umbral_max:   return "🔴 VENDER"
            if peso_compra >= umbral_min:   return "⚡ POSIBLE COMPRA"
            if peso_venta  >= umbral_min:   return "⚠️ POSIBLE VENTA"
            return "🟡 MANTENER"

        # ── RANGO VISUAL ──────────────────────────────────────────────────────
        dvm = {"1 Mes":30,"3 Meses":90,"6 Meses":180,"1 Año":365,"2 Años":730,
               "3 Años":1095,"5 Años":1825,"Todo":99999}
        ff_vt = pd.Timestamp(precios_vt.index[-1])
        if st.session_state.get("vt_custom", False):
            fi_vt = fecha_anterior_cercana(precios_vt.index,
                        pd.Timestamp(st.session_state.get("vt_fi", datetime(2023,1,1))))
            ff_vt = fecha_anterior_cercana(precios_vt.index,
                        pd.Timestamp(st.session_state.get("vt_ff", datetime.today())))
        else:
            dias_vt = dvm[rango_vt]
            fi_vt   = precios_vt.index[0] if dias_vt>=99999 else fecha_anterior_cercana(
                          precios_vt.index, ff_vt - pd.Timedelta(days=dias_vt))
        mask_vt = (precios_vt.index >= fi_vt) & (precios_vt.index <= ff_vt)

        def _range_buttons():
            return [
                dict(count=1,  label="1M",  step="month", stepmode="backward"),
                dict(count=3,  label="3M",  step="month", stepmode="backward"),
                dict(count=6,  label="6M",  step="month", stepmode="backward"),
                dict(count=1,  label="1A",  step="year",  stepmode="backward"),
                dict(count=3,  label="3A",  step="year",  stepmode="backward"),
                dict(step="all", label="Todo"),
            ]

        def _apply_rangeslider(lo):
            lo["xaxis"]["rangeslider"] = dict(visible=True,bgcolor="#0A1628",bordercolor="#1A3A5C",thickness=0.05)
            lo["xaxis"]["rangeselector"] = dict(
                buttons=_range_buttons(),
                bgcolor="#0A1628", activecolor="#1A5C9A",
                bordercolor="#1A3A5C", borderwidth=1,
                font=dict(color="#E8F4FD", size=10, family="IBM Plex Mono"),
                x=0, y=1.02)
            # Habilitar zoom en eje Y con doble clic o drag vertical
            lo["yaxis"]["fixedrange"] = False
            lo["dragmode"] = "zoom"
            return lo

        def _fib_chart(titulo, niveles, color_linea, tipo_linea,
                       maximo, minimo, fecha_max, fecha_min, serie_p, tkr_p, colores_p):
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=serie_p.index, y=serie_p.values,
                name=tkr_p, line=dict(color="#E8F4FD",width=1.8),
                hovertemplate="%{x|%d/%m/%y}<br>$%{y:.2f}<extra></extra>"))
            fig.add_trace(go.Scatter(x=[fecha_max], y=[maximo], mode="markers+text",
                name="Pivote Máx",
                text=[f"Máx {fmt_num(maximo,2,es_usd=True)}"],
                textposition="top center",
                textfont=dict(size=10, color="#F39C12", family="IBM Plex Mono"),
                marker=dict(color="#F39C12", size=10, symbol="triangle-down")))
            fig.add_trace(go.Scatter(x=[fecha_min], y=[minimo], mode="markers+text",
                name="Pivote Mín",
                text=[f"Mín {fmt_num(minimo,2,es_usd=True)}"],
                textposition="bottom center",
                textfont=dict(size=10, color="#4AB3F4", family="IBM Plex Mono"),
                marker=dict(color="#4AB3F4", size=10, symbol="triangle-up")))
            vals_niv = list(niveles.values())
            for i,(lb,nivel) in enumerate(niveles.items()):
                fig.add_hline(y=nivel, line_color=colores_p[i%5],
                    line_dash=tipo_linea, line_width=1.5, opacity=0.9,
                    annotation_text=f"{lb}: {fmt_num(nivel,2,es_usd=True)}",
                    annotation_font=dict(color=colores_p[i%5], size=10, family="IBM Plex Mono"),
                    annotation_position="right")
            if len(vals_niv) >= 4:
                fig.add_hrect(y0=vals_niv[2], y1=vals_niv[3],
                    fillcolor="#F39C12", opacity=0.06, line_width=0,
                    annotation_text="Zona crítica", annotation_font_color="#F39C12")
            lo = base_layout(titulo, 430)
            lo["yaxis"]["title"] = "Precio"
            lo["legend"]["font"]["color"] = "#E8F4FD"
            lo = _apply_rangeslider(lo)
            fig.update_layout(**lo)
            return fig

        if btn_vt or any(k.startswith("vt_calc_") for k in st.session_state):
            if btn_vt:
                for t in tks_ok_vt:
                    st.session_state[f"vt_calc_{t}"] = True

            # ── CALCULAR PARA TODOS LOS TICKERS ──────────────────────────────
            resultados_vt = {}
            for tkr in tks_ok_vt:
                serie = precios_vt.loc[mask_vt, tkr].dropna()
                if len(serie) < 30: continue

                # Medias móviles — alineación multimediada
                mms = {p: serie.rolling(p).mean() for p in [mm1,mm2,mm3,mm4]}
                señal_mm_str,compra_mm,venta_mm = señal_mm_multi(serie,mm1,mm2,mm3,mm4)

                # RSI
                rsi_serie = calc_rsi(serie, rsi_per)
                rsi_val   = float(rsi_serie.iloc[-1]) if not pd.isna(rsi_serie.iloc[-1]) else 50.0
                s_rsi,compra_rsi,venta_rsi = señal_rsi(rsi_val, rsi_sob, rsi_sov)

                # MACD
                macd_s,sig_s,hist_s = calc_macd(serie, macd_fast, macd_slow, macd_sig)
                s_macd,compra_macd,venta_macd = señal_macd_fn(macd_s, sig_s)

                # Fibonacci
                alza,baja,maximo,minimo,tendencia_fibo,fecha_max_fibo,fecha_min_fibo = calc_fibonacci(serie, fibo_sens)
                s_fibo,compra_fibo,venta_fibo = señal_fibo(float(serie.iloc[-1]), alza, baja, tendencia_fibo)

                # Pesos totales
                total_peso = peso_mm + peso_macd + peso_rsi + peso_fibo
                if total_peso > 0:
                    w_mm   = peso_mm   / total_peso
                    w_macd = peso_macd / total_peso
                    w_rsi  = peso_rsi  / total_peso
                    w_fibo = peso_fibo / total_peso
                else:
                    w_mm=w_macd=w_rsi=w_fibo=0.25

                peso_compra_total = (compra_mm*w_mm + compra_macd*w_macd +
                                     compra_rsi*w_rsi + compra_fibo*w_fibo)*100
                peso_venta_total  = (venta_mm*w_mm  + venta_macd*w_macd  +
                                     venta_rsi*w_rsi + venta_fibo*w_fibo)*100
                dec_final = decision_final(peso_compra_total, peso_venta_total,
                                           peso_dec_max, peso_dec_min)

                resultados_vt[tkr] = dict(
                    serie=serie, mms=mms, rsi_serie=rsi_serie, rsi_val=rsi_val,
                    macd_s=macd_s, sig_s=sig_s, hist_s=hist_s,
                    alza=alza, baja=baja, maximo=maximo, minimo=minimo,
                    señal_mm=señal_mm_str,  compra_mm=compra_mm,   venta_mm=venta_mm,
                    señal_rsi=s_rsi,        compra_rsi=compra_rsi, venta_rsi=venta_rsi,
                    señal_macd=s_macd,      compra_macd=compra_macd, venta_macd=venta_macd,
                    señal_fibo=s_fibo,      compra_fibo=compra_fibo, venta_fibo=venta_fibo,
                    tendencia_fibo=tendencia_fibo,
                    fecha_max_fibo=fecha_max_fibo, fecha_min_fibo=fecha_min_fibo,
                    w_mm=w_mm, w_macd=w_macd, w_rsi=w_rsi, w_fibo=w_fibo,
                    peso_compra=peso_compra_total, peso_venta=peso_venta_total,
                    decision=dec_final,
                )

            if not resultados_vt:
                st.warning("No hay suficientes datos. Amplía el rango."); st.stop()

            # ── RESUMEN COMPARATIVO ────────────────────────────────────────
            st.markdown('<div class="section-header">Resumen Técnico — Todos los Activos</div>',
                        unsafe_allow_html=True)
            sum_h = "<th>Activo</th><th>MM</th><th>RSI</th><th>RSI Val.</th><th>MACD</th><th>Fibonacci</th><th>% Compra</th><th>% Venta</th><th>Decisión Final</th>"
            sum_r = ""
            for t,r in resultados_vt.items():
                cc = "td-positive" if r["peso_compra"]>r["peso_venta"] else "td-negative"
                cv = "td-negative" if r["peso_venta"]>r["peso_compra"] else ""
                sum_r += (f'<tr><td class="td-blue">{t}</td>'
                          f'<td>{r["señal_mm"]}</td>'
                          f'<td>{r["señal_rsi"]}</td>'
                          f'<td class="{"td-negative" if r["rsi_val"]>rsi_sob else ("td-positive" if r["rsi_val"]<rsi_sov else "td-neutral")}">{fmt_num(r["rsi_val"]/100,1,es_pct=True)}</td>'
                          f'<td>{r["señal_macd"]}</td>'
                          f'<td>{r["señal_fibo"]}</td>'
                          f'<td class="{cc}">{fmt_num(r["peso_compra"]/100,1,es_pct=True)}</td>'
                          f'<td class="{cv}">{fmt_num(r["peso_venta"]/100,1,es_pct=True)}</td>'
                          f'<td style="font-weight:600">{r["decision"]}</td></tr>')
            st.markdown(f'<table class="styled-table"><thead><tr>{sum_h}</tr></thead><tbody>{sum_r}</tbody></table>',
                        unsafe_allow_html=True)

            # ── PESTAÑAS POR ACTIVO ────────────────────────────────────────
            st.markdown('<div class="section-header">Detalle por Activo</div>', unsafe_allow_html=True)
            tabs_vt = st.tabs([f"📡 {t}" for t in resultados_vt])


            for tab_vt_t,(tkr,res) in zip(tabs_vt,resultados_vt.items()):
                with tab_vt_t:
                    serie = res["serie"]

                    # ── TABLA DE DECISIÓN ──────────────────────────────────
                    st.markdown('<div class="section-header">Tabla de Decisión</div>',
                                unsafe_allow_html=True)
                    indicadores = [
                        ("Medias Móviles", res["señal_mm"],  res["compra_mm"],  res["venta_mm"],  res["w_mm"]),
                        ("RSI",            res["señal_rsi"], res["compra_rsi"], res["venta_rsi"], res["w_rsi"]),
                        ("MACD",           res["señal_macd"],res["compra_macd"],res["venta_macd"],res["w_macd"]),
                        ("Fibonacci",      res["señal_fibo"],res["compra_fibo"],res["venta_fibo"],res["w_fibo"]),
                    ]
                    td_h = "<th>Indicador</th><th>Señal</th><th>Compra</th><th>Venta</th><th>Peso Compra</th><th>Peso Venta</th>"
                    td_r = ""
                    for ind,señal,comp,vent,peso in indicadores:
                        pc_ind = comp*peso*100; pv_ind = vent*peso*100
                        css_s = "td-positive" if comp else ("td-negative" if vent else "td-neutral")
                        td_r += (f'<tr><td class="td-label">{ind}</td>'
                                 f'<td class="{css_s}">{señal}</td>'
                                 f'<td>{comp}</td><td>{vent}</td>'
                                 f'<td class="{"td-positive" if pc_ind>0 else ""}">{fmt_num(pc_ind/100,1,es_pct=True)}</td>'
                                 f'<td class="{"td-negative" if pv_ind>0 else ""}">{fmt_num(pv_ind/100,1,es_pct=True)}</td></tr>')
                    # Totales
                    td_r += (f'<tr style="border-top:2px solid #4AB3F4">'
                             f'<td style="font-weight:600;color:#E8F4FD">TOTAL</td><td></td><td></td><td></td>'
                             f'<td class="td-positive" style="font-weight:600">{fmt_num(res["peso_compra"]/100,1,es_pct=True)}</td>'
                             f'<td class="td-negative" style="font-weight:600">{fmt_num(res["peso_venta"]/100,1,es_pct=True)}</td></tr>')
                    st.markdown(f'<table class="styled-table"><thead><tr>{td_h}</tr></thead><tbody>{td_r}</tbody></table>',
                                unsafe_allow_html=True)

                    # Decisión final KPI
                    dec_color = ("#2ECC71" if "COMPRAR" in res["decision"] else
                                 "#E74C3C" if "VENDER"  in res["decision"] else
                                 "#F39C12" if "POSIBLE" in res["decision"] else "#4AB3F4")
                    st.markdown(f"""<div style="background:#0D1E35;border:1px solid #1A3A5C;
                        border-left:4px solid {dec_color};border-radius:6px;
                        padding:1rem 1.5rem;margin:1rem 0;font-family:IBM Plex Mono,monospace;">
                      <span style="font-size:.65rem;color:#4A7FA5;letter-spacing:.12em;
                            text-transform:uppercase;">Decisión Final — {tkr}</span><br>
                      <span style="font-size:1.4rem;font-weight:600;color:{dec_color};">
                        {res["decision"]}</span>
                    </div>""", unsafe_allow_html=True)

                    # ── TABLA FIBONACCI ────────────────────────────────────
                    precio_ult = float(serie.iloc[-1])
                    tend_color = "#2ECC71" if res["tendencia_fibo"]=="ALCISTA" else "#E74C3C"
                    tend_icon  = "📈" if res["tendencia_fibo"]=="ALCISTA" else "📉"
                    fm = fmt_fecha(res["fecha_max_fibo"]); fmi = fmt_fecha(res["fecha_min_fibo"])
                    st.markdown(f"""<div style="background:#0D1E35;border:1px solid #1A3A5C;
                        border-left:4px solid {tend_color};border-radius:6px;
                        padding:.7rem 1.2rem;margin:.5rem 0;font-family:IBM Plex Mono,monospace;">
                      <span style="font-size:.62rem;color:#4A7FA5;letter-spacing:.1em;text-transform:uppercase;">
                        Tendencia detectada — Pivote N={int(fibo_sens)} días vecinos</span><br>
                      <span style="font-size:1.1rem;font-weight:600;color:{tend_color};">
                        {tend_icon} {res["tendencia_fibo"]}</span>
                      <span style="font-size:.7rem;color:#7FA8C9;margin-left:1rem;">
                        Pivote Máx: {fmt_num(res["maximo"],2,es_usd=True)} ({fm}) &nbsp;|&nbsp;
                        Pivote Mín: {fmt_num(res["minimo"],2,es_usd=True)} ({fmi})</span>
                    </div>""", unsafe_allow_html=True)

                    st.markdown('<div class="section-header">Niveles de Fibonacci</div>',
                                unsafe_allow_html=True)
                    col_fa, col_fb = st.columns(2)
                    with col_fa:
                        fh  = '<p style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#2ECC71;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.35rem;">📈 Fibonacci Alza — Retrocesos desde Máximo</p>'
                        fh += '<table class="styled-table"><thead><tr><th>Nivel</th><th>Precio</th><th>Distancia</th></tr></thead><tbody>'
                        for lb,nivel in res["alza"].items():
                            dist = (precio_ult - nivel) / nivel * 100
                            css_d = "td-positive" if precio_ult > nivel else "td-negative"
                            fh += f'<tr><td class="td-label">{lb}</td><td>{fmt_num(nivel,2,es_usd=True)}</td><td class="{css_d}">{fmt_num(dist/100,2,es_pct=True)}</td></tr>'
                        fh += '</tbody></table>'
                        st.markdown(fh, unsafe_allow_html=True)
                    with col_fb:
                        fb  = '<p style="font-family:IBM Plex Mono,monospace;font-size:.63rem;color:#E74C3C;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.35rem;">📉 Fibonacci Baja — Recuperaciones desde Mínimo</p>'
                        fb += '<table class="styled-table"><thead><tr><th>Nivel</th><th>Precio</th><th>Distancia</th></tr></thead><tbody>'
                        for lb,nivel in res["baja"].items():
                            dist = (precio_ult - nivel) / nivel * 100
                            css_d = "td-positive" if precio_ult > nivel else "td-negative"
                            fb += f'<tr><td class="td-label">{lb}</td><td>{fmt_num(nivel,2,es_usd=True)}</td><td class="{css_d}">{fmt_num(dist/100,2,es_pct=True)}</td></tr>'
                        fb += '</tbody></table>'
                        st.markdown(fb, unsafe_allow_html=True)

                    # ── GRÁFICAS EN PESTAÑAS INDEPENDIENTES ───────────────
                    st.markdown('<div class="section-header">Gráficas Técnicas</div>',
                                unsafe_allow_html=True)
                    from plotly.subplots import make_subplots as _msp
                    gt1,gt2,gt3,gt4 = st.tabs(["📈 Precio + MM","📊 MACD","📉 RSI","🌀 Fibonacci"])

                    with gt1:
                        fig_mm = go.Figure()
                        fig_mm.add_trace(go.Scatter(x=serie.index,y=serie.values,
                            name=tkr,line=dict(color="#E8F4FD",width=1.8),
                            hovertemplate="%{x|%d/%m/%y}<br>$%{y:.2f}<extra></extra>"))
                        mm_colors = ["#4AB3F4","#F39C12","#2ECC71","#E74C3C"]
                        for i,(per,mm_s) in enumerate(res["mms"].items()):
                            fig_mm.add_trace(go.Scatter(x=mm_s.index,y=mm_s.values,
                                name=f"MM{per}",line=dict(color=mm_colors[i%4],width=1.2,dash="dot"),
                                hovertemplate=f"MM{per}: %{{y:.2f}}<extra></extra>"))
                        lo1=base_layout(f"Precio + Medias Móviles — {tkr}",420)
                        lo1["yaxis"]["title"]="Precio"
                        lo1["legend"]["font"]["color"]="#E8F4FD"
                        lo1=_apply_rangeslider(lo1)
                        fig_mm.update_layout(**lo1); st.plotly_chart(fig_mm,use_container_width=True)

                    with gt2:
                        fig_macd = _msp(rows=2,cols=1,shared_xaxes=True,
                                        row_heights=[0.4,0.6],vertical_spacing=0.05)
                        hist_vals = res["hist_s"]
                        fig_macd.add_trace(go.Bar(x=hist_vals.index,y=hist_vals.values,
                            name="Histograma",
                            marker_color=["#2ECC71" if v>=0 else "#E74C3C" for v in hist_vals.values],
                            hovertemplate="%{x|%d/%m/%y}<br>%{y:.4f}<extra></extra>"),row=1,col=1)
                        fig_macd.add_trace(go.Scatter(x=res["macd_s"].index,y=res["macd_s"].values,
                            name="MACD",line=dict(color="#4AB3F4",width=1.5),
                            hovertemplate="%{x|%d/%m/%y}<br>%{y:.4f}<extra></extra>"),row=2,col=1)
                        fig_macd.add_trace(go.Scatter(x=res["sig_s"].index,y=res["sig_s"].values,
                            name="Signal",line=dict(color="#F39C12",width=1.5),
                            hovertemplate="%{x|%d/%m/%y}<br>%{y:.4f}<extra></extra>"),row=2,col=1)
                        fig_macd.update_layout(
                            plot_bgcolor=PLOT_BG,paper_bgcolor=PAPER_BG,height=420,
                            font=dict(family="IBM Plex Sans",size=11,color=FONT_COLOR),
                            margin=dict(l=50,r=30,t=45,b=40),
                            legend=dict(bgcolor="rgba(0,0,0,0)",bordercolor=GRID_COLOR,
                                        borderwidth=1,font=dict(size=10,color="#E8F4FD")),
                            hoverlabel=dict(bgcolor="#0D1E35",bordercolor="#1A5C9A",
                                            font=dict(family="IBM Plex Mono",size=11)),
                            title=dict(text=f"MACD ({macd_fast},{macd_slow},{macd_sig}) — {tkr}",
                                       font=dict(family="IBM Plex Mono",size=11,color="#7FA8C9"),x=0.01),
                            xaxis2=dict(
                                rangeslider=dict(visible=True,bgcolor="#0A1628",bordercolor="#1A3A5C",thickness=0.05),
                                rangeselector=dict(buttons=_range_buttons(),
                                    bgcolor="#0A1628",activecolor="#1A5C9A",bordercolor="#1A3A5C",
                                    borderwidth=1,font=dict(color="#E8F4FD",size=10,family="IBM Plex Mono"),
                                    x=0,y=1.05)))
                        for row in [1,2]:
                            fig_macd.update_xaxes(gridcolor=GRID_COLOR,showgrid=True,row=row,col=1)
                            fig_macd.update_yaxes(gridcolor=GRID_COLOR,showgrid=True,fixedrange=False,row=row,col=1)
                        st.plotly_chart(fig_macd,use_container_width=True)

                    with gt3:
                        fig_rsi = go.Figure()
                        fig_rsi.add_trace(go.Scatter(x=res["rsi_serie"].index,
                            y=res["rsi_serie"].values,name="RSI",
                            line=dict(color="#9B59B6",width=1.8),
                            hovertemplate="%{x|%d/%m/%y}<br>RSI: %{y:.2f}<extra></extra>"))
                        fig_rsi.add_hline(y=rsi_sob,line_color="#E74C3C",line_dash="dash",
                            line_width=1.5,annotation_text=f"Sobrecompra ({rsi_sob})",
                            annotation_font_color="#E74C3C")
                        fig_rsi.add_hline(y=rsi_sov,line_color="#2ECC71",line_dash="dash",
                            line_width=1.5,annotation_text=f"Sobreventa ({rsi_sov})",
                            annotation_font_color="#2ECC71")
                        fig_rsi.add_hline(y=50,line_color="#4A7FA5",line_dash="dot",line_width=1,opacity=0.5)
                        fig_rsi.add_hrect(y0=rsi_sob,y1=100,fillcolor="#E74C3C",opacity=0.05,line_width=0)
                        fig_rsi.add_hrect(y0=0,y1=rsi_sov,fillcolor="#2ECC71",opacity=0.05,line_width=0)
                        lo3=base_layout(f"RSI ({rsi_per}) — {tkr}",380)
                        lo3["yaxis"].update(title="RSI",range=[0,100],fixedrange=False)
                        lo3["legend"]["font"]["color"]="#E8F4FD"
                        lo3=_apply_rangeslider(lo3)
                        fig_rsi.update_layout(**lo3); st.plotly_chart(fig_rsi,use_container_width=True)

                    with gt4:
                        colores_fib = ["#4AB3F4","#2ECC71","#F39C12","#E74C3C","#9B59B6"]
                        tend_color_g = "#2ECC71" if res["tendencia_fibo"]=="ALCISTA" else "#E74C3C"
                        tend_txt     = res["tendencia_fibo"]

                            # Gráfica Alza
                        st.markdown(f'<p style="font-family:IBM Plex Mono,monospace;font-size:.65rem;color:#2ECC71;letter-spacing:.1em;text-transform:uppercase;margin:.5rem 0 .2rem;">📈 Fibonacci Alza — {tend_txt} · Pivote N={int(fibo_sens)}</p>', unsafe_allow_html=True)
                        st.plotly_chart(_fib_chart(
                            f"Fibonacci Alza — {tkr}",
                            res["alza"], "#2ECC71", "dash",
                            res["maximo"], res["minimo"],
                            res["fecha_max_fibo"], res["fecha_min_fibo"],
                            serie, tkr, colores_fib), use_container_width=True)

                        # Gráfica Baja
                        st.markdown(f'<p style="font-family:IBM Plex Mono,monospace;font-size:.65rem;color:#E74C3C;letter-spacing:.1em;text-transform:uppercase;margin:.5rem 0 .2rem;">📉 Fibonacci Baja — {tend_txt} · Pivote N={int(fibo_sens)}</p>', unsafe_allow_html=True)
                        st.plotly_chart(_fib_chart(
                            f"Fibonacci Baja — {tkr}",
                            res["baja"], "#E74C3C", "dot",
                            res["maximo"], res["minimo"],
                            res["fecha_max_fibo"], res["fecha_min_fibo"],
                            serie, tkr, colores_fib), use_container_width=True)

        else:
            st.markdown("""<div style="text-align:center;padding:4rem 2rem;color:#4A7FA5;
                 font-family:IBM Plex Mono,monospace;">
              <div style="font-size:3rem;margin-bottom:1rem;">📡</div>
              <div style="font-size:.85rem;letter-spacing:.1em;text-transform:uppercase;">
                Configura los parámetros y presiona<br>
                <span style="color:#4AB3F4;">CALCULAR ANÁLISIS TÉCNICO</span></div>
            </div>""", unsafe_allow_html=True)




# ══════════════════════════════════════════════════════════════════════════════
# BLOQUE 5 — VALORACIÓN FUNDAMENTAL
# ══════════════════════════════════════════════════════════════════════════════
with tab5:
    if "precios" not in st.session_state:
        st.info("⬅️  Primero carga los datos en **01 · Análisis Básico**")
    else:
        precios_vf   = st.session_state["precios"]
        tks_vf       = st.session_state["tickers_list"]
        bench_vf     = st.session_state["bench"]
        rf_vf        = st.session_state.get("rf", 0.045)
        tks_ok_vf    = [t for t in tks_vf if t in precios_vf.columns]
        bench_ok_vf  = bench_vf in precios_vf.columns
        ff_vf        = pd.Timestamp(precios_vf.index[-1])
        base_year_vf = ff_vf.year

        if bench_ok_vf:
            rb_vf   = calc_rendimientos(precios_vf[[bench_vf]]).iloc[:,0]
            rm_calc = float((1 + rb_vf.mean())**365 - 1)
        else:
            rm_calc = 0.10

        # ── PASO 1: CONFIGURACIÓN GENERAL ─────────────────────────────────────
        with st.expander("⚙️  PASO 1 — Configuración General", expanded=True):
            pc1,pc2,pc3,pc4 = st.columns(4)
            with pc1:
                rm_vf_pct = st.number_input("Rm — Retorno Mercado (%)", 0.0, 50.0,
                    round(rm_calc*100,4), 0.01, format="%.4f", key="rm_vf",
                    help="Calculado del benchmark. Editable.")
                rm_vf = rm_vf_pct / 100
            with pc2:
                tasa_imp_ui = st.number_input("Tasa de Impuestos (%)", 0.0, 60.0,
                    30.0, 0.5, format="%.2f", key="tasa_imp_vf")
            with pc3:
                trm_vf = st.number_input("TRM (moneda local/USD)", 0.0, 200000.0,
                    1.0, 100.0, format="%.2f", key="trm_vf")
            with pc4:
                n_comps = st.number_input("N° comparables", 3, 20, 6, 1, key="n_comps_vf")
            comps_manual_txt = st.text_input(
                "Comparables manuales (opcional, separados por coma)",
                value="", key="comps_manual_vf")

        # ── PASO 2: TABLA DE PROYECCIÓN EDITABLE ──────────────────────────────
        st.markdown('<div class="section-header">PASO 2 — Tabla de Proyección Macroeconómica</div>',
                    unsafe_allow_html=True)

        # Precargar valores FMI por país (detectado del primer ticker)
        # Defaults: USA. Se ajustan según país.
        fmi_defaults_usa = {
            0: (1.3, 6.0),
            1: (3.6, 3.3),
            2: (3.1, 3.4),
            3: (3.2, 3.5),
            4: (3.5, 3.5),
        }
        fmi_defaults_col = {
            0: (5.1, 2.6),
            1: (5.9, 2.3),
            2: (4.5, 2.5),
            3: (3.5, 2.8),
            4: (3.2, 3.0),
        }

        anos_tabla = [base_year_vf + i for i in range(5)]

        st.markdown(f"""<div style="background:#0D1E35;border:1px solid #1A3A5C;
            border-radius:4px;padding:.6rem 1rem;margin-bottom:.8rem;
            font-family:IBM Plex Mono,monospace;font-size:.72rem;color:#7FA8C9;">
            Años de proyección: <span style="color:#4AB3F4;font-weight:600;">
            {anos_tabla[0]} → {anos_tabla[-1]}</span>
            &nbsp;·&nbsp; Último precio disponible: <span style="color:#4AB3F4;">{fmt_fecha(ff_vf)}</span>
        </div>""", unsafe_allow_html=True)

        # Tabla editable con st.data_editor
        import pandas as _pd_vf
        # Inicializar tabla solo si no existe o si cambia el año base
        needs_init = (
            "tabla_macro_vf" not in st.session_state or
            st.session_state["tabla_macro_vf"]["Año"].iloc[0] != base_year_vf
        )
        if needs_init:
            rows = []
            for i, yr in enumerate(anos_tabla):
                inf_d, pib_d = fmi_defaults_usa.get(i, (2.0, 2.0))
                rows.append({"Año": yr, "Inflación (%)": inf_d,
                             "PIB (%)": pib_d, "Prima Sector (%)": 0.0})
            st.session_state["tabla_macro_vf"] = _pd_vf.DataFrame(rows)

        # Usar clave fija para data_editor — los cambios se reflejan en session_state
        tabla_editada = st.data_editor(
            st.session_state["tabla_macro_vf"],
            key="editor_macro_vf",
            use_container_width=True,
            hide_index=True,
            on_change=None,
            column_config={
                "Año":             st.column_config.NumberColumn("Año", disabled=True, format="%d"),
                "Inflación (%)":   st.column_config.NumberColumn("Inflación (%)", min_value=-10.0, max_value=50.0, step=0.1, format="%.2f"),
                "PIB (%)":         st.column_config.NumberColumn("PIB (%)", min_value=-20.0, max_value=30.0, step=0.1, format="%.2f"),
                "Prima Sector (%)":st.column_config.NumberColumn("Prima Sector (%)", min_value=0.0, max_value=20.0, step=0.1, format="%.2f"),
            }
        )
        # Actualizar session_state solo si cambió
        if not tabla_editada.equals(st.session_state["tabla_macro_vf"]):
            st.session_state["tabla_macro_vf"] = tabla_editada

        # Calcular G por año y mostrar preview
        g_por_ano = {}
        preview_rows = ""
        for _, row in tabla_editada.iterrows():
            inf = row["Inflación (%)"] / 100
            pib = row["PIB (%)"] / 100
            prima = row["Prima Sector (%)"] / 100
            G = ((1+inf)*(1+pib)*(1+prima)) - 1
            g_por_ano[int(row["Año"])] = G
            preview_rows += (f'<tr><td class="td-label">{int(row["Año"])}</td>'
                             f'<td>{fmt_num(inf,2,es_pct=True)}</td>'
                             f'<td>{fmt_num(pib,2,es_pct=True)}</td>'
                             f'<td>{fmt_num(prima,2,es_pct=True)}</td>'
                             f'<td class="td-blue" style="font-weight:600">'
                             f'{fmt_num(G,2,es_pct=True)}</td></tr>')

        st.markdown(
            '<table class="styled-table"><thead><tr>'
            '<th>Año</th><th>Inflación</th><th>PIB</th><th>Prima</th>'
            '<th>G = ((1+Inf)×(1+PIB)×(1+Prima))-1</th>'
            '</tr></thead><tbody>' + preview_rows + '</tbody></table>',
            unsafe_allow_html=True)

        # ── PASO 3: BOTÓN CALCULAR ─────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        btn_vf = st.button("🏦  PASO 3 — CALCULAR VALORACIÓN FUNDAMENTAL", key="btn_vf",
                            type="primary")

        # ── CLASE EXTRACTOR ────────────────────────────────────────────────────
        class ExtractorYahoo:
            def __init__(self, ticker):
                self.ticker_str = ticker
                self.tk = yf.Ticker(ticker)
                try:    self.info = self.tk.info or {}
                except: self.info = {}
                self.balance_q  = self._sg("quarterly_balance_sheet")
                self.income_q   = self._sg("quarterly_income_stmt")
                self.cashflow_q = self._sg("quarterly_cashflow")
                self.balance_a  = self._sg("balance_sheet")
                self.income_a   = self._sg("income_stmt")
                self.cashflow_a = self._sg("cashflow")

            def _sg(self, attr):
                try:
                    df = getattr(self.tk, attr)
                    return df if (df is not None and not df.empty) else pd.DataFrame()
                except: return pd.DataFrame()

            def _buscar(self, df, nombres, ci=0):
                if df is None or df.empty: return None
                for n in nombres:
                    if n in df.index:
                        try:
                            v = df.loc[n].iloc[ci]
                            if pd.notna(v): return float(v)
                        except: continue
                return None

            def _ttm(self, df, nombres):
                if df is None or df.empty: return None
                for n in nombres:
                    if n in df.index:
                        try:
                            fila = df.loc[n].iloc[:4].dropna()
                            if len(fila)>=4: return float(fila.sum())
                            if len(fila)>0:  return float(fila.sum()*4/len(fila))
                        except: continue
                return None

            def get_nombre(self):    return self.info.get("longName") or self.info.get("shortName","N/A")
            def get_sector(self):    return self.info.get("sector","N/A")
            def get_industria(self): return self.info.get("industry","N/A")
            def get_moneda(self):    return self.info.get("currency","USD")
            def get_pais(self):      return self.info.get("country","US")
            def get_activo(self):    return self._buscar(self.balance_q,["Total Assets"])
            def get_pasivo(self):    return self._buscar(self.balance_q,["Total Liabilities Net Minority Interest","Total Liab","Total Liabilities"])
            def get_patrimonio(self):return self._buscar(self.balance_q,["Stockholders Equity","Total Equity","Common Stock Equity"])
            def get_num_acciones(self): return self.info.get("sharesOutstanding")
            def get_dps(self):
                # Forward Dividend: valor en $ que yfinance reporta como dividendRate
                fwd = self.info.get("dividendRate")
                return round(float(fwd), 4) if fwd else 0.0
            def get_deuda(self):
                v = self._buscar(self.balance_q,["Long Term Debt","Total Debt"])
                return v if v else (self.info.get("totalDebt") or 0)
            def get_efectivo(self):  return self._buscar(self.balance_q,["Cash And Cash Equivalents","Cash","Cash Cash Equivalents And Short Term Investments"]) or 0
            def get_inventario(self, ci=0): return self._buscar(self.balance_q,["Inventory"],ci)
            def get_cxc(self, ci=0): return self._buscar(self.balance_q,["Net Receivables","Accounts Receivable","Receivables"],ci)
            def get_cxp(self, ci=0): return self._buscar(self.balance_q,["Accounts Payable","Payables"],ci)
            def get_ppe(self, ci=0):
                v = self._buscar(self.balance_q,["Net PPE","Property Plant Equipment","Gross PPE"],ci)
                return v if v else self._buscar(self.balance_a,["Net PPE","Property Plant Equipment"],ci)
            def get_activo_ant(self):    return self._buscar(self.balance_q,["Total Assets"],1)
            def get_pasivo_ant(self):    return self._buscar(self.balance_q,["Total Liabilities Net Minority Interest","Total Liab"],1)
            def get_patrim_ant(self):    return self._buscar(self.balance_q,["Stockholders Equity","Total Equity","Common Stock Equity"],1)
            def get_ut_neta_ttm(self):   return self._ttm(self.income_q,["Net Income"])
            def get_ing_oper_ttm(self):  return self._ttm(self.income_q,["Total Revenue","Revenue"])
            def get_ebitda_ttm(self):    return self.info.get("ebitda") or self._ttm(self.income_q,["EBITDA","Normalized EBITDA"])
            def get_ut_oper_ttm(self):   return self._ttm(self.income_q,["Operating Income","EBIT"])
            def get_da_ttm(self):        return self._ttm(self.cashflow_q,["Depreciation Amortization Depletion","Depreciation And Amortization","Depreciation"])
            def get_gts_fin_ttm(self):
                v = self._ttm(self.income_q,["Interest Expense","Finance Cost"])
                if not v: v = self._ttm(self.cashflow_q,["Interest Paid","Interest Expense Paid"])
                return abs(v) if v else None
            def get_capex_ttm(self):
                v = self._ttm(self.cashflow_q,["Capital Expenditure","Purchase Of PPE","Capital Expenditures"])
                return abs(v) if v else None

        # ── FUNCIONES ──────────────────────────────────────────────────────────
        @st.cache_data(ttl=3600)
        def descargar_datos_vf(ticker, n_comp, comps_man_str):
            comps_man = [c.strip().upper() for c in comps_man_str.split(",") if c.strip()] if comps_man_str else None
            try:
                ext = ExtractorYahoo(ticker)
                def _s(v): return v if (v and not (isinstance(v,float) and np.isnan(v))) else 0
                inv0=ext.get_inventario(0); cxc0=ext.get_cxc(0); cxp0=ext.get_cxp(0)
                inv1=ext.get_inventario(1); cxc1=ext.get_cxc(1); cxp1=ext.get_cxp(1)
                datos = {
                    "ticker": ticker,
                    "nombre": ext.get_nombre(), "sector": ext.get_sector(),
                    "industria": ext.get_industria(), "moneda": ext.get_moneda(),
                    "pais": ext.get_pais(),
                    "activo": ext.get_activo(),    "pasivo": ext.get_pasivo(),
                    "patrim": ext.get_patrimonio(),
                    "activo_ant": ext.get_activo_ant(), "pasivo_ant": ext.get_pasivo_ant(),
                    "patrim_ant": ext.get_patrim_ant(),
                    "num_acc": ext.get_num_acciones(),
                    "dps": ext.get_dps(),
                    "deuda": ext.get_deuda(),   "efectivo": ext.get_efectivo(),
                    "ut_neta": ext.get_ut_neta_ttm(),
                    "ing_op":  ext.get_ing_oper_ttm(),
                    "ebitda":  ext.get_ebitda_ttm(),
                    "ut_op":   ext.get_ut_oper_ttm(),
                    "dep_am":  ext.get_da_ttm(),
                    "gts_fin": ext.get_gts_fin_ttm(),
                    "capex":   ext.get_capex_ttm(),
                    "inv0": inv0, "cxc0": cxc0, "cxp0": cxp0,
                    "inv1": inv1, "cxc1": cxc1, "cxp1": cxp1,
                    "ppe0": ext.get_ppe(0), "ppe1": ext.get_ppe(1),
                    "ktno0": _s(inv0)+_s(cxc0)-_s(cxp0),
                    "ktno1": _s(inv1)+_s(cxc1)-_s(cxp1),
                }
                # Comparables — usar campos directos de yfinance info
                comps = []
                try:
                    src = comps_man if comps_man else list(yf.Industry(datos["industria"]).top_companies.index[:n_comp+3])
                    for tkc in src:
                        if tkc == ticker: continue
                        try:
                            ec   = ExtractorYahoo(tkc)
                            ic   = ec.info
                            mc   = ic.get("marketCap",0) or 0
                            if mc<=0: continue
                            # Activos del balance (más confiable que info)
                            ta   = ec.get_activo() or ic.get("totalAssets",0) or 0
                            deu  = ec.get_deuda() or ic.get("totalDebt",0) or 0
                            ebd  = ec.get_ebitda_ttm() or ic.get("ebitda",0) or 0
                            rev  = ec.get_ing_oper_ttm() or ic.get("totalRevenue",1) or 1
                            pr   = ic.get("currentPrice",np.nan)
                            pat  = ec.get_patrimonio() or 0
                            ut_n = ec.get_ut_neta_ttm() or 0
                            n_ac = ec.get_num_acciones() or 1
                            ef_c = ec.get_efectivo() or 0
                            ev_c = mc + deu - ef_c
                            # Q Tobin = (MarketCap + Deuda) / Total Activos
                            qt_c  = (mc+deu)/ta if (ta and ta>0) else np.nan
                            # P/E
                            upa_c = ut_n/n_ac if (ut_n and n_ac) else None
                            rpg_c = ic.get("trailingPE") or (pr/upa_c if (pr and upa_c and upa_c>0) else np.nan)
                            # Yield = dividendo / precio
                            dps_c = ic.get("dividendRate",0) or 0
                            div_c = dps_c/pr if (dps_c and pr) else (ic.get("dividendYield",0) or 0)
                            if div_c and div_c>1: div_c = div_c/100
                            # ROE
                            roe_c = ic.get("returnOnEquity") or (ut_n/pat if (ut_n and pat and pat!=0) else np.nan)
                            # Margen EBITDA
                            mg_c  = ic.get("ebitdaMargins") or (ebd/rev if (ebd and rev) else np.nan)
                            # EV/EBITDA
                            evebd = ic.get("enterpriseToEbitda") or (ev_c/ebd if (ev_c and ebd) else np.nan)
                            comps.append({
                                "Ticker":     tkc, "Nombre": ic.get("longName",""),
                                "_qt":        qt_c,   "_rpg":       rpg_c,
                                "_yield":     div_c,  "_roe":       roe_c,
                                "_mg_ebitda": mg_c,   "_ev_ebitda": evebd,
                                "_mkt_cap":   mc,     "_num_acc":   n_ac,
                                "_precio":    pr,     "_dps":       dps_c,
                                "_ut_neta":   ut_n,   "_ing_op":    rev,
                                "_ebitda":    ebd,    "_deuda":     deu,
                                "_efectivo":  ef_c,
                            })
                            if len(comps)>=n_comp: break
                        except: continue
                except: pass
                datos["comps"] = comps
                return datos
            except Exception as e:
                return {"error": str(e)}

        def calc_beta_vf(tkr):
            try:
                ra = calc_rendimientos(precios_vf[[tkr]]).iloc[:,0]
                rb = calc_rendimientos(precios_vf[[bench_vf]]).iloc[:,0]
                df = pd.concat([ra,rb],axis=1).dropna()
                if len(df)<30: return 1.0
                from scipy import stats as sp
                sl,*_ = sp.linregress(df.iloc[:,1].values, df.iloc[:,0].values)
                return round(sl,4)
            except: return 1.0

        def calcular_fcf_proyeccion(datos, precio_mdo, ke, kd, tasa_imp,
                                        g_por_ano_d, trm, base_year, prima_sector=0.0):
            """
            base_year  = último año del rango (ff_vf.year)
            g_por_ano_d= {año: G} — el primer año es base_year (sin VP),
                         los siguientes se proyectan y descuentan.
            Fix 3: año base muestra FCF₀ sin VP; proyección arranca el año siguiente.
            Fix 4: VT = FCF_último / (WACC - prima_sector)
            Fix 5: VP_VT = VT / (1+WACC)^(n_años_proyectados)
            """
            num   = datos.get("num_acc") or 0
            deuda = datos.get("deuda")   or 0
            ef    = datos.get("efectivo") or 0
            ut_op = datos.get("ut_op")   or 0
            dep   = abs(datos.get("dep_am") or 0)

            anos_sorted = sorted(g_por_ano_d.keys())
            # Años de proyección = todos excepto el año base
            anos_proy = [yr for yr in anos_sorted if yr > base_year]
            n = len(anos_proy)   # número de períodos descontados

            # prima_sector viene como parámetro explícito (columna "Prima Sector (%)")

            # FCF base (año base, t=0)
            imp_rta = ut_op * tasa_imp
            uodi    = ut_op - imp_rta
            fcb     = uodi - dep
            vktno   = datos["ktno0"] - datos["ktno1"]
            ppe0    = datos.get("ppe0") or 0
            ppe1    = datos.get("ppe1") or 0
            vcap    = ppe0 - ppe1
            fcf0    = fcb - vktno - vcap

            # Iteración WACC ↔ Precio
            precio_iter = precio_mdo; wacc_f = ke
            for _ in range(100):
                E = num*precio_iter; D = deuda; tot = E+D
                if tot<=0: break
                we = E/tot; wd = D/tot
                wacc_new = ke*we + kd*(1-tasa_imp)*wd
                if wacc_new <= prima_sector: break
                # Proyección solo años > base_year
                fcf_prev = fcf0; vp_sum_iter = 0
                for t, yr in enumerate(anos_proy, 1):
                    G_t   = g_por_ano_d[yr]
                    fcf_t = fcf_prev*(1+G_t)
                    vp_sum_iter += fcf_t/(1+wacc_new)**t
                    fcf_prev = fcf_t
                # VT = FCF_último / (WACC - prima_sector)  [Fix 4]
                vt_iter = fcf_prev / (wacc_new - prima_sector) if wacc_new > prima_sector else 0
                # VP_VT descontado n períodos  [Fix 5]
                vp_vt_iter = vt_iter / (1+wacc_new)**n if n>0 else 0
                # Para iteración usamos misma estructura EV
                activo_i  = datos.get("activo")  or 0
                inv0_i    = datos.get("inv0")    or 0
                cxc0_i    = datos.get("cxc0")    or 0
                ppe0_i    = datos.get("ppe0")    or 0
                pasivo_i  = datos.get("pasivo")  or 0
                cxp0_i    = datos.get("cxp0")    or 0
                oa_i = activo_i - inv0_i - cxc0_i - ppe0_i
                op_i = pasivo_i - cxp0_i
                eq_iter = vp_sum_iter + vp_vt_iter
                ev_iter = eq_iter + oa_i - op_i
                p_new   = ev_iter/num if num>0 else 0
                if abs(p_new-precio_iter)<0.001:
                    precio_iter=p_new; wacc_f=wacc_new; break
                precio_iter=p_new; wacc_f=wacc_new

            # Proyección final con wacc_f
            E_f=num*precio_iter; D_f=deuda; tot_f=E_f+D_f
            we_f=E_f/tot_f if tot_f>0 else 0.5
            wd_f=D_f/tot_f if tot_f>0 else 0.5

            # Año base sin VP
            proyt=[{"year":base_year,"G":None,"fcf":fcf0,"vp":None}]
            fcf_prev=fcf0
            for t, yr in enumerate(anos_proy, 1):
                G_t   = g_por_ano_d[yr]
                fcf_t = fcf_prev*(1+G_t)
                vp_t  = fcf_t/(1+wacc_f)**t
                proyt.append({"year":yr,"G":G_t,"fcf":fcf_t,"vp":vp_t})
                fcf_prev = fcf_t

            vp_sum  = sum(x["vp"] for x in proyt if x["vp"] is not None)
            fcf_ult = proyt[-1]["fcf"]
            # Fix 4: VT = FCF_último / (WACC - prima_sector)
            vt_val  = fcf_ult/(wacc_f-prima_sector) if wacc_f>prima_sector else 0
            # Fix 5: VP_VT = VT / (1+WACC)^n
            vp_vt   = vt_val/(1+wacc_f)**n if n>0 else 0
            # Equity Value = Σ VP proyectados + VP Valor Terminal (sin año base)
            equity_value = vp_sum + vp_vt

            # Otros Activos = Activo Total - Inventario - CxC - PP&E  (año actual)
            activo  = datos.get("activo")  or 0
            inv0    = datos.get("inv0")    or 0
            cxc0    = datos.get("cxc0")    or 0
            ppe0    = datos.get("ppe0")    or 0
            pasivo  = datos.get("pasivo")  or 0
            cxp0    = datos.get("cxp0")    or 0
            otros_activos = activo - inv0 - cxc0 - ppe0
            otros_pasivos = pasivo - cxp0

            # Enterprise Value = Equity Value + Otros Activos - Otros Pasivos
            enterprise_value = equity_value + otros_activos - otros_pasivos
            po_usd = enterprise_value/num if num>0 else np.nan
            po_cop = po_usd*trm if po_usd else np.nan

            return dict(
                tasa_imp=tasa_imp, imp_rta=imp_rta, uodi=uodi,
                dep_am=dep, fcb=fcb, vktno=vktno, vcap=vcap, fcf=fcf0,
                wacc=wacc_f, we=we_f, wd=wd_f, ke=ke, kd=kd,
                E_val=E_f, D_val=D_f, prima_sector=prima_sector,
                proyt=proyt, vp_sum=vp_sum, n_proy=n,
                vt_val=vt_val, vp_vt=vp_vt,
                equity_value=equity_value,
                otros_activos=otros_activos, otros_pasivos=otros_pasivos,
                enterprise_value=enterprise_value,
                po_usd=po_usd, po_cop=po_cop,
            )

        # ── CÁLCULO ────────────────────────────────────────────────────────────
        if btn_vf or any(k.startswith("vf_calc_") for k in st.session_state):
            if btn_vf:
                for t in tks_ok_vf: st.session_state[f"vf_calc_{t}"] = True

            tasa_imp = tasa_imp_ui/100
            trm      = trm_vf

            resultados_vf = {}
            pb = st.progress(0,"Descargando estados financieros...")
            for i_t,tkr in enumerate(tks_ok_vf):
                pb.progress((i_t+1)/len(tks_ok_vf),f"Procesando {tkr}...")
                datos = descargar_datos_vf(tkr, int(n_comps), comps_manual_txt)
                if "error" in datos:
                    resultados_vf[tkr]={"error":datos["error"]}; continue

                precio_mdo = float(precios_vf[tkr].iloc[-1])
                datos["precio"]           = precio_mdo
                datos["fecha_extraccion"] = fmt_fecha(ff_vf)

                beta = calc_beta_vf(tkr) if bench_ok_vf else 1.0
                ke   = rf_vf + (rm_vf - rf_vf)*beta
                kd   = abs(datos["gts_fin"]/datos["deuda"]) if (
                    datos.get("gts_fin") and datos.get("deuda") and datos["deuda"]!=0) else 0.05

                # Prima sector = primer valor "Prima Sector (%)" de la tabla
                prima_sec_val = float(tabla_editada["Prima Sector (%)"].iloc[0])/100 if len(tabla_editada)>0 else 0.0
                res  = calcular_fcf_proyeccion(datos, precio_mdo, ke, kd,
                                               tasa_imp, g_por_ano, trm,
                                               base_year_vf, prima_sec_val)
                comps = datos.get("comps",[])

                # Múltiplos comparables (pesos del doc del profe)
                pesos_mult = {"RPG (P/E)":0.10,"Yield":0.05,"ROE":0.15,
                              "Mg EBITDA":0.30,"EV/EBITDA":0.40}
                num_acc = datos.get("num_acc") or 1
                pat_obj = datos.get("patrim") or 0
                ut_n    = datos.get("ut_neta") or 0
                ing_op  = datos.get("ing_op")  or 1
                ebd_obj = datos.get("ebitda")  or 0
                ev_obj  = (precio_mdo*num_acc + datos.get("deuda",0) - datos.get("efectivo",0))
                upa_obj = ut_n/num_acc if num_acc else np.nan
                dps_obj = datos.get("dps",0)
                roe_obj = ut_n/pat_obj if pat_obj!=0 else np.nan
                mg_obj  = ebd_obj/ing_op if ing_op else np.nan
                ev_ebd_obj = ev_obj/ebd_obj if ebd_obj else np.nan
                qt_obj  = (precio_mdo*num_acc)/(datos.get("activo") or 1)
                rpg_obj = precio_mdo/upa_obj if (upa_obj and upa_obj>0) else np.nan
                yld_obj = dps_obj/precio_mdo if precio_mdo else 0

                obj_vals = {"Q Tobin":qt_obj,"RPG (P/E)":rpg_obj,"Yield":yld_obj,
                            "ROE":roe_obj,"Mg EBITDA":mg_obj,"EV/EBITDA":ev_ebd_obj}

                eficiencia_pond = 0.0
                ef_details = {}
                for campo, peso in pesos_mult.items():
                    vals_c = [c[campo] for c in comps
                              if c.get(campo) and not np.isnan(c[campo])]
                    if vals_c and not np.isnan(obj_vals.get(campo,np.nan)):
                        prom_c = float(np.mean(vals_c))
                        v_obj  = obj_vals[campo]
                        ef_i   = (v_obj - prom_c)/prom_c if prom_c!=0 else 0
                        ef_details[campo] = {"obj":v_obj,"prom":prom_c,
                                             "ef":ef_i,"peso":peso}
                        eficiencia_pond += ef_i*peso

                precio_sug = precio_mdo*(1+eficiencia_pond)

                def _pot(po): return (po/precio_mdo-1) if (po and precio_mdo) else np.nan
                pots=[p for p in [_pot(res.get("po_usd")),_pot(precio_sug)] if isinstance(p,float) and not np.isnan(p)]
                pot_p=float(np.mean(pots)) if pots else np.nan
                señal=("🟢 COMPRAR" if pot_p>0.15 else "🟡 POSIBLE COMPRA" if pot_p>0.05
                       else "🔴 VENDER" if pot_p<-0.15 else "⚠️ POSIBLE VENTA" if pot_p<-0.05
                       else "⚪ MANTENER") if not np.isnan(pot_p) else "N/A"

                resultados_vf[tkr]={**datos,**res,
                    "beta":beta,"g_por_ano":g_por_ano,
                    "comps":comps,
                    "pot_fcf":_pot(res.get("po_usd")),
                    "pot_pond":pot_p,"señal":señal}

            pb.empty()
            st.session_state["resultados_vf"]=resultados_vf

        if "resultados_vf" in st.session_state:
            resultados_vf = st.session_state["resultados_vf"]

            # ── RESUMEN ────────────────────────────────────────────────────────
            st.markdown('<div class="section-header">Resumen Valoración Fundamental</div>',
                        unsafe_allow_html=True)
            rsh = ("<th>Ticker</th><th>Precio</th><th>PO FCF USD</th>"
                   "<th>PO FCF Local</th><th>Pot.FCF</th><th>Señal FCF</th>"
                   "<th>Precio Múltiplos</th><th>Ratio vs Precio</th>"
                   "<th>Señal Múltiplos</th><th>WACC</th><th>G Perp</th>")
            rsr = ""
            for t,r in resultados_vf.items():
                if "error" in r:
                    rsr += ('<tr><td class="td-blue">'+str(t)+'</td>'
                            '<td colspan="10" class="td-negative">Error: '+r["error"][:50]+'</td></tr>')
                    continue
                def _fs(v,f="p"):
                    if not isinstance(v,(int,float)) or np.isnan(v): return "N/A"
                    return fmt_num(v,2,es_usd=True) if f=="p" else fmt_num(v,2,es_pct=True)
                cf="td-positive" if (r.get("pot_fcf",0) or 0)>0 else "td-negative"
                ratio_v = r.get("ratio_mult", np.nan)
                css_ratio = ("td-positive" if (isinstance(ratio_v,float) and not np.isnan(ratio_v) and ratio_v>1.5)
                             else "td-negative" if (isinstance(ratio_v,float) and not np.isnan(ratio_v) and ratio_v<0.9)
                             else "td-neutral")
                precio_cv_r = st.session_state.get(f"precio_conv_{t}", r.get("precio",np.nan))
                rsr += ('<tr><td class="td-blue">'+str(t)+'</td>'
                        +'<td>'+_fs(r.get("precio"))+'</td>'
                        +'<td>'+_fs(r.get("po_usd"))+'</td>'
                        +'<td>'+_fs(r.get("po_cop"))+'</td>'
                        +'<td class="'+cf+'">'+_fs(r.get("pot_fcf"),"pct")+'</td>'
                        +'<td style="font-weight:600">'+r.get("señal_fcf","N/A")+'</td>'
                        +'<td>'+_fs(precio_cv_r)+'</td>'
                        +'<td class="'+css_ratio+'">'+(fmt_num(ratio_v*100,2,es_pct=True) if isinstance(ratio_v,float) and not np.isnan(ratio_v) else "N/A")+'</td>'
                        +'<td style="font-weight:600">'+r.get("señal_mult","N/A")+'</td>'
                        +'<td>'+_fs(r.get("wacc"),"pct")+'</td>'
                        +'<td>'+_fs(r.get("G_perp"),"pct")+'</td>'
                        +'</tr>')
            st.markdown('<table class="styled-table"><thead><tr>'+rsh+'</tr></thead><tbody>'+rsr+'</tbody></table>',
                        unsafe_allow_html=True)

            # ── PESTAÑAS POR ACTIVO ────────────────────────────────────────────
            st.markdown('<div class="section-header">Detalle por Activo</div>',
                        unsafe_allow_html=True)
            tabs_vf = st.tabs([f"🏦 {t}" for t in resultados_vf])

            for tab_vf_t,(tkr,r) in zip(tabs_vf, resultados_vf.items()):
                with tab_vf_t:
                    if "error" in r:
                        st.error(f"Error {tkr}: {r['error']}"); continue

                    def _M(v,f="usd"):
                        if not isinstance(v,(int,float)) or np.isnan(v): return "N/A"
                        if f=="pct":  return fmt_num(v,2,es_pct=True)
                        if f=="dec":  return fmt_num(v,4)
                        if f=="dec2": return fmt_num(v,2)
                        if f=="price":return fmt_num(v,2,es_usd=True)
                        av=abs(v); s="-" if v<0 else ""
                        if av>=1e9: return s+fmt_num(av/1e9,3)+"B"
                        if av>=1e6: return s+fmt_num(av/1e6,3)+"M"
                        return fmt_num(v,2,es_usd=True)

                    poc=r.get("pot_fcf",0) or 0; pp=r.get("pot_pond",0) or 0
                    tend_c=("#2ECC71" if "COMPRAR" in r.get("señal","") else
                            "#E74C3C" if "VENDER"  in r.get("señal","") else
                            "#F39C12" if "POSIBLE" in r.get("señal","") else "#4AB3F4")
                    mkt_cap=(r.get("num_acc") or 0)*(r.get("precio") or 0)

                    st.markdown(f"""<div class="kpi-grid">
                      <div class="kpi-card"><div class="kpi-label">Precio ({fmt_fecha(ff_vf)})</div>
                        <div class="kpi-value">{_M(r.get("precio",np.nan),"price")}</div></div>
                      <div class="kpi-card"><div class="kpi-label">PO FCF (USD)</div>
                        <div class="kpi-value {'positive' if poc>0 else 'negative'}">{_M(r.get("po_usd",np.nan),"price")}</div>
                        <div class="kpi-delta">Pot: {_M(poc,"pct")}</div></div>
                      <div class="kpi-card"><div class="kpi-label">PO Local (×{fmt_num(trm_vf,0)})</div>
                        <div class="kpi-value" style="font-size:1rem">{_M(r.get("po_cop",np.nan),"price")}</div></div>
                      <div class="kpi-card"><div class="kpi-label">WACC</div>
                        <div class="kpi-value">{_M(r.get("wacc",np.nan),"pct")}</div></div>
                      <div class="kpi-card"><div class="kpi-label">G Perpetuidad</div>
                        <div class="kpi-value">{_M(r.get("G_perp",np.nan),"pct")}</div></div>
                      <div class="kpi-card"><div class="kpi-label">Señal</div>
                        <div class="kpi-value" style="font-size:.8rem;color:{tend_c}">{r.get("señal","N/A")}</div></div>
                    </div>""",unsafe_allow_html=True)

                    col_a,col_b = st.columns(2)

                    with col_a:
                        # Identificación
                        st.markdown('<div class="section-header">Datos de la Empresa</div>',
                                    unsafe_allow_html=True)
                        id_h='<table class="styled-table"><thead><tr><th>Campo</th><th>Año Actual</th><th>Año Anterior</th></tr></thead><tbody>'
                        id_h+=('<tr><td class="td-label">Ticker</td><td colspan="2">'+tkr+'</td></tr>'
                               +'<tr><td class="td-label">Nombre</td><td colspan="2">'+r.get("nombre","N/A")+'</td></tr>'
                               +'<tr><td class="td-label">Sector</td><td colspan="2">'+r.get("sector","N/A")+'</td></tr>'
                               +'<tr><td class="td-label">Industria</td><td colspan="2">'+r.get("industria","N/A")+'</td></tr>'
                               +'<tr><td class="td-label">Moneda</td><td colspan="2">'+r.get("moneda","USD")+'</td></tr>'
                               +'<tr><td class="td-label">Fecha extracción</td><td colspan="2">'+r.get("fecha_extraccion","N/A")+'</td></tr>')
                        def _row2(lbl,v0,v1=None,f="usd"):
                            ex="<td>"+_M(v1,f)+"</td>" if v1 is not None else "<td></td>"
                            return '<tr><td class="td-label">'+lbl+"</td><td>"+_M(v0,f)+"</td>"+ex+"</tr>"
                        id_h+=_row2("Precio Mercado",r.get("precio"),None,"price")
                        id_h+=_row2("Mkt Cap",mkt_cap)
                        id_h+=_row2("# Acciones",r.get("num_acc"),None,"dec")
                        id_h+=_row2("Dividendo Yield (%)",r.get("dps",0),None,"dec2")
                        id_h+=_row2("Activo Total",r.get("activo"),r.get("activo_ant"))
                        id_h+=_row2("Pasivo Total",r.get("pasivo"),r.get("pasivo_ant"))
                        id_h+=_row2("Patrimonio",r.get("patrim"),r.get("patrim_ant"))
                        id_h+=_row2("Deuda Financiera",r.get("deuda"))
                        id_h+=_row2("Efectivo",r.get("efectivo"))
                        id_h+='<tr><td colspan="3" style="background:#0A1628;color:#4AB3F4;font-size:.62rem;letter-spacing:.1em;padding:.4rem;text-transform:uppercase">TTM — Últimos 12 Meses</td></tr>'
                        id_h+=_row2("UO 12M",r.get("ut_op"))
                        id_h+=_row2("GASTOS DEP, AMOR Y PROV 12M",r.get("dep_am"))
                        id_h+=_row2("EBITDA TTM",r.get("ebitda"))
                        id_h+=_row2("Ut. Neta TTM",r.get("ut_neta"))
                        id_h+=_row2("Ingresos Operacionales",r.get("ing_op"))
                        id_h+=_row2("Gastos Financieros TTM",r.get("gts_fin"))
                        id_h+=_row2("CAPEX TTM",r.get("capex"))
                        id_h+='<tr><td colspan="3" style="background:#0A1628;color:#4AB3F4;font-size:.62rem;letter-spacing:.1em;padding:.4rem;text-transform:uppercase">Capital de Trabajo</td></tr>'
                        id_h+=_row2("Inventario",r.get("inv0"),r.get("inv1"))
                        id_h+=_row2("CxC",r.get("cxc0"),r.get("cxc1"))
                        id_h+=_row2("CxP",r.get("cxp0"),r.get("cxp1"))
                        id_h+=_row2("PP&E",r.get("ppe0"),r.get("ppe1"))
                        id_h+=('<tr><td class="td-label" style="font-weight:600">KTNO</td>'
                               +'<td class="td-blue">'+_M(r.get("ktno0"))+'</td>'
                               +'<td class="td-blue">'+_M(r.get("ktno1"))+'</td></tr>')
                        id_h+="</tbody></table>"
                        st.markdown(id_h,unsafe_allow_html=True)

                    with col_b:
                        # FCF
                        st.markdown('<div class="section-header">1. FCF</div>',
                                    unsafe_allow_html=True)
                        fh='<table class="styled-table"><thead><tr><th>Concepto</th><th>Valor</th></tr></thead><tbody>'
                        fh+='<tr><td colspan="2" style="background:#0A1628;color:#4AB3F4;font-size:.62rem;letter-spacing:.1em;padding:.4rem;text-transform:uppercase">Estructura de Capital</td></tr>'
                        fh+=('<tr><td class="td-label">UO 12M</td><td>'+_M(r.get("ut_op"))+'</td></tr>'
                             +'<tr><td class="td-label">GASTOS DEP, AMOR Y PROV 12M</td><td>'+_M(r.get("dep_am"))+'</td></tr>'
                             +'<tr><td class="td-label">Tasa Impuestos</td><td>'+_M(r.get("tasa_imp",np.nan),"pct")+'</td></tr>'
                             +'<tr><td class="td-label">Risk Free (RF)</td><td>'+_M(rf_vf,"pct")+'</td></tr>'
                             +'<tr><td class="td-label">R Mercado</td><td>'+_M(rm_vf,"pct")+'</td></tr>'
                             +'<tr><td class="td-label">Beta</td><td>'+_M(r.get("beta",np.nan),"dec")+'</td></tr>'
                             +'<tr><td class="td-label">Ke = RF+(Rm-RF)×Beta</td><td class="td-blue">'+_M(r.get("ke",np.nan),"pct")+'</td></tr>'
                             +'<tr><td class="td-label">Kd = Gts.Fin/Deuda</td><td>'+_M(r.get("kd",np.nan),"pct")+'</td></tr>'
                             +'<tr><td class="td-label">Mkt Cap (E)</td><td>'+_M(mkt_cap)+'</td></tr>'
                             +'<tr><td class="td-label">W% Equity</td><td>'+_M(r.get("we",np.nan),"pct")+'</td></tr>'
                             +'<tr><td class="td-label">W% Deuda</td><td>'+_M(r.get("wd",np.nan),"pct")+'</td></tr>'
                             +'<tr><td class="td-label" style="font-weight:600">WACC (iterado)</td><td class="td-blue" style="font-weight:600">'+_M(r.get("wacc",np.nan),"pct")+'</td></tr>')
                        fh+='<tr><td colspan="2" style="background:#1A5C9A;padding:.5rem;color:#E8F4FD;font-size:.75rem;font-weight:600;letter-spacing:.1em;text-transform:uppercase">FREE CASH FLOW</td></tr>'
                        fh+=('<tr><td class="td-label">Utilidad Operacional 12M</td><td>'+_M(r.get("ut_op"))+'</td></tr>'
                             +'<tr><td class="td-label">(−) Impuesto de Renta</td><td>'+_M(r.get("imp_rta"))+'</td></tr>'
                             +'<tr><td class="td-label" style="font-weight:600">= UODI</td><td class="td-blue">'+_M(r.get("uodi"))+'</td></tr>'
                             +'<tr><td class="td-label">(−) Gas, Dep y Amor 12M</td><td>'+_M(r.get("dep_am"))+'</td></tr>'
                             +'<tr><td class="td-label" style="font-weight:600">= Flujo Caja Bruto</td><td class="td-blue">'+_M(r.get("fcb"))+'</td></tr>'
                             +'<tr><td class="td-label">(−) Var $ KTNO</td><td>'+_M(r.get("vktno"))+'</td></tr>'
                             +'<tr><td class="td-label">(−) Var $ Capex (PP&E)</td><td>'+_M(r.get("vcap"))+'</td></tr>')
                        fcf_v=r.get("fcf",np.nan)
                        css_f="td-positive" if isinstance(fcf_v,float) and not np.isnan(fcf_v) and fcf_v>0 else "td-negative"
                        fh+=('<tr style="border-top:2px solid #4AB3F4"><td style="font-weight:600;color:#E8F4FD">= FREE CASH FLOW</td>'
                             +'<td class="'+css_f+'" style="font-weight:600">'+_M(fcf_v)+'</td></tr>')
                        fh+="</tbody></table>"
                        st.markdown(fh,unsafe_allow_html=True)

                    # Proyección
                    st.markdown('<div class="section-header">Proyección FCF</div>',
                                unsafe_allow_html=True)
                    proyt=r.get("proyt",[])
                    if proyt:
                        ph="<th>Concepto</th>"+"".join(f"<th>{x['year']}</th>" for x in proyt)+"<th>Vlr Terminal</th>"
                        # G row: base year has no G
                        g_cells="".join(
                            "<td>—</td>" if x["G"] is None
                            else f"<td>{_M(x['G'],'pct')}</td>"
                            for x in proyt)
                        g_row="<tr><td class='td-label'>G año</td>"+g_cells+"<td></td></tr>"
                        # FCF row
                        fcf_row=("<tr><td class='td-label'>FCF</td>"
                                 +"".join(f"<td>{_M(x['fcf'])}</td>" for x in proyt)
                                 +f"<td>{_M(r.get('vt_val'))}</td></tr>")
                        # VP row: base year has no VP
                        vp_cells="".join(
                            "<td>—</td>" if x["vp"] is None
                            else f"<td class='td-blue'>{_M(x['vp'])}</td>"
                            for x in proyt)
                        vp_row=("<tr><td class='td-label'>VP</td>"+vp_cells
                                +f"<td class='td-blue'>{_M(r.get('vp_vt'))}</td></tr>")
                        st.markdown('<table class="styled-table"><thead><tr>'+ph+'</tr></thead><tbody>'+g_row+fcf_row+vp_row+'</tbody></table>',
                                    unsafe_allow_html=True)
                        # Info note
                        n_p = r.get("n_proy",0)
                        st.markdown(f'<div style="font-family:IBM Plex Mono,monospace;font-size:.62rem;color:#4A7FA5;margin-top:.3rem;">'
                                    f'VT = FCF_{proyt[-1]["year"]} / (WACC − Prima) &nbsp;·&nbsp; '
                                    f'VP_VT = VT / (1+WACC)^{n_p} &nbsp;·&nbsp; '
                                    f'G perpetuidad (Prima sector) = {_M(r.get("prima_sector",0),"pct")}'
                                    f'</div>', unsafe_allow_html=True)

                    # Equity Value y PO
                    col_ev1,col_ev2=st.columns(2)
                    with col_ev1:
                        # Tabla con el orden exacto de la imagen del profe
                        ev_v  = r.get("equity_value", np.nan)
                        oa_v  = r.get("otros_activos", np.nan)
                        op_v  = r.get("otros_pasivos", np.nan)
                        entv  = r.get("enterprise_value", np.nan)
                        po_u  = r.get("po_usd", np.nan)
                        po_c  = r.get("po_cop", np.nan)
                        n_acc = r.get("num_acc", np.nan)

                        ev_h = '<table class="styled-table"><thead><tr><th>Concepto</th><th>Valor</th></tr></thead><tbody>'
                        # Equity Value
                        ev_h += ('<tr style="border-bottom:1px solid #1A3A5C">'
                                 +'<td style="font-weight:600;color:#E8F4FD">Equity Value</td>'
                                 +'<td style="font-weight:600">'+_M(ev_v)+'</td></tr>')
                        # + Otros activos
                        ev_h += '<tr><td class="td-label">(+) Otros Activos</td><td>'+_M(oa_v)+'</td></tr>'
                        ev_h += '<td class="td-neutral" style="font-size:.6rem;padding:.2rem .5rem" colspan="2">Activo Total − Inventario − CxC − PP&E</td>'
                        # - Otros pasivos
                        ev_h += '<tr><td class="td-label">(−) Otros Pasivos</td><td>'+_M(op_v)+'</td></tr>'
                        ev_h += '<td class="td-neutral" style="font-size:.6rem;padding:.2rem .5rem" colspan="2">Pasivo Total − CxP</td>'
                        # Enterprise Value
                        ev_h += ('<tr style="border-top:2px solid #4AB3F4">'
                                 +'<td style="font-weight:600;color:#E8F4FD">Enterprise Value</td>'
                                 +'<td class="td-blue" style="font-weight:600">'+_M(entv)+'</td></tr>')
                        # Acciones
                        ev_h += '<tr><td class="td-label">Acción Circul.</td><td>'+_M(n_acc,"dec")+'</td></tr>'
                        # PO
                        ev_h += ('<tr><td class="td-label" style="font-weight:600">PO (USD)</td>'
                                 +'<td class="td-blue" style="font-weight:600">'+_M(po_u,"price")+'</td></tr>')
                        if trm_vf > 1:
                            ev_h += '<tr><td class="td-label">TRM</td><td>'+fmt_num(trm_vf,0)+'</td></tr>'
                            ev_h += ('<tr><td class="td-label" style="font-weight:600">PO COP</td>'
                                     +'<td class="td-blue" style="font-weight:600">'+_M(po_c,"price")+'</td></tr>')
                        ev_h += "</tbody></table>"
                        st.markdown(ev_h, unsafe_allow_html=True)





                    # ══════════════════════════════════════════════════════════
                    # SECCIÓN 2 — VALORACIÓN POR MÚLTIPLOS
                    # ══════════════════════════════════════════════════════════
                    st.markdown("""<div style="background:#0B2E1A;border:1px solid #2ECC71;
                        border-radius:6px;padding:.7rem 1.2rem;margin:1.5rem 0 .8rem;
                        font-family:IBM Plex Mono,monospace;">
                      <span style="font-size:.65rem;color:#2ECC71;letter-spacing:.12em;
                            text-transform:uppercase;font-weight:600;">
                        ▌ Sección 2 — Valoración por Múltiplos Comparables</span>
                    </div>""", unsafe_allow_html=True)

                    comps          = r.get("comps", [])
                    precio_inicial = r.get("precio", np.nan)
                    num_acc_obj    = r.get("num_acc") or 1
                    ut_neta_obj    = r.get("ut_neta") or 0
                    ing_op_obj     = r.get("ing_op") or 1
                    ebd_obj        = r.get("ebitda") or 0
                    deuda_obj      = r.get("deuda") or 0
                    ef_obj         = r.get("efectivo") or 0
                    div_obj        = r.get("dps", 0) or 0

                    # Patrimonio = MktCap inicial (FIJO — base para Vlr Libros y ROE)
                    patrim_display = precio_inicial * num_acc_obj
                    # Vlr Libros = precio_inicial (FIJO — no cambia con iteraciones)
                    vlr_libros     = precio_inicial
                    upa_obj        = ut_neta_obj / num_acc_obj if num_acc_obj else np.nan
                    # ROE usa MktCap inicial (fijo)
                    roe_fijo       = (ut_neta_obj / patrim_display
                                      if patrim_display else np.nan)
                    mg_fijo        = ebd_obj / ing_op_obj if ing_op_obj else np.nan

                    # Recuperar precio convergido
                    precio_key     = f"precio_conv_{tkr}"
                    # Recuperar precio convergido — resetear si es irrazonable (>100x precio inicial)
                    _precio_stored = st.session_state.get(precio_key, precio_inicial)
                    if (isinstance(_precio_stored, float) and precio_inicial and
                        (_precio_stored > precio_inicial * 100 or _precio_stored < precio_inicial * 0.01)):
                        _precio_stored = precio_inicial
                        if precio_key in st.session_state: del st.session_state[precio_key]
                    precio_conv = _precio_stored

                    # MKT CAP dinámico (cambia con iteración)
                    mkt_cap_iter   = precio_conv * num_acc_obj
                    ev_iter        = mkt_cap_iter + deuda_obj - ef_obj

                    def _fmt_m(v):
                        if not isinstance(v,(int,float)) or np.isnan(v): return "N/A"
                        return fmt_num(v, 2)

                    def _calc_mult_iter(precio_p):
                        """Calcula múltiplos de MI acción con precio dado."""
                        mc_p   = precio_p * num_acc_obj          # MKT CAP dinámico
                        ev_p   = mc_p + deuda_obj - ef_obj
                        qt_p   = precio_p / vlr_libros if vlr_libros else np.nan
                        rpg_p  = precio_p / upa_obj if (upa_obj and upa_obj>0) else np.nan
                        yld_p  = (div_obj/precio_p if (div_obj>0 and precio_p)
                                  else (upa_obj/precio_p if (upa_obj and upa_obj>0 and precio_p) else np.nan))
                        roe_p  = roe_fijo          # ROE usa Patrimonio fijo
                        mg_p   = mg_fijo           # Margen EBITDA no depende del precio
                        eve_p  = ev_p/ebd_obj if ebd_obj else np.nan
                        return qt_p, rpg_p, yld_p, roe_p, mg_p, eve_p, mc_p, ev_p

                    qt_conv,rpg_conv,yld_conv,roe_conv,mg_conv,eve_conv,mc_conv,ev_conv = _calc_mult_iter(precio_conv)

                    # ── Tabla 1: Datos empresa ─────────────────────────────────
                    st.markdown('<div class="section-header">Datos de la Empresa (Múltiplos)</div>',
                                unsafe_allow_html=True)
                    t1_cols = ["Empresa","Patrimonio (fijo)","MKT CAP","#Acciones",
                               "Vlr Libros","Precio Mdo","Dividendo","UPA 12",
                               "Utl Neta 12","Ing Ope 12","EBITDA 12",
                               "Deuda Fin","Efectivo","EV"]
                    t1h = "".join(f"<th>{c}</th>" for c in t1_cols)
                    t1r = (f'<tr>'
                           +f'<td class="td-blue" style="font-weight:600">{tkr}</td>'
                           +f'<td>{_M(patrim_display)}</td>'
                           +f'<td>{_M(mc_conv)}</td>'
                           +f'<td>{_M(num_acc_obj,"dec")}</td>'
                           +f'<td>{_M(vlr_libros,"price")}</td>'
                           +f'<td>{_M(precio_conv,"price")}</td>'
                           +f'<td>{_M(div_obj,"price")}</td>'
                           +f'<td>{_M(upa_obj,"price")}</td>'
                           +f'<td>{_M(ut_neta_obj)}</td>'
                           +f'<td>{_M(ing_op_obj)}</td>'
                           +f'<td>{_M(ebd_obj)}</td>'
                           +f'<td>{_M(deuda_obj)}</td>'
                           +f'<td>{_M(ef_obj)}</td>'
                           +f'<td>{_M(ev_conv)}</td>'
                           +'</tr>')
                    st.markdown('<table class="styled-table"><thead><tr>'+t1h
                                +'</tr></thead><tbody>'+t1r+'</tbody></table>',
                                unsafe_allow_html=True)

                    # ── Promedios sector desde yfinance directamente ───────────
                    def _prom_pos(lst):
                        vals=[v for v in lst if isinstance(v,(int,float))
                              and not np.isnan(v) and v>0]
                        return float(np.mean(vals)) if vals else np.nan

                    qt_list=[]; rpg_list=[]; yld_list=[]; roe_list=[]; mg_list=[]; eve_list=[]
                    comp_rows = ""
                    def _fmtc(v, pct=False):
                        if not isinstance(v,(int,float)) or np.isnan(v): return "N/A"
                        return fmt_num(v,2,es_pct=True) if pct else fmt_num(v,2)

                    for comp in comps:
                        tkc    = comp.get("Ticker","")
                        # Usar directamente los campos de yfinance — sin recalcular
                        qt_c   = comp.get("_qt", np.nan)
                        rpg_c  = comp.get("_rpg", np.nan)
                        yld_c  = comp.get("_yield", np.nan)
                        roe_c  = comp.get("_roe", np.nan)
                        mg_c   = comp.get("_mg_ebitda", np.nan)
                        eve_c  = comp.get("_ev_ebitda", np.nan)
                        for lst,v in zip([qt_list,rpg_list,yld_list,roe_list,mg_list,eve_list],
                                         [qt_c,rpg_c,yld_c,roe_c,mg_c,eve_c]):
                            if isinstance(v,(int,float)) and not np.isnan(v) and v>0:
                                lst.append(v)
                        comp_rows += (f'<tr><td class="td-label">{tkc}</td>'
                                      +f'<td>{_fmtc(qt_c)}</td>'
                                      +f'<td>{_fmtc(rpg_c)}</td>'
                                      +f'<td>{_fmtc(yld_c,True)}</td>'
                                      +f'<td>{_fmtc(roe_c,True)}</td>'
                                      +f'<td>{_fmtc(mg_c,True)}</td>'
                                      +f'<td>{_fmtc(eve_c)}</td></tr>')

                    prom_qt  = _prom_pos(qt_list)
                    prom_rpg = _prom_pos(rpg_list)
                    prom_yld = _prom_pos(yld_list)
                    prom_roe = _prom_pos(roe_list)
                    prom_mg  = _prom_pos(mg_list)
                    prom_eve = _prom_pos(eve_list)

                    # ── Tabla 2: Múltiplos ─────────────────────────────────────
                    st.markdown('<div class="section-header">Múltiplos</div>',
                                unsafe_allow_html=True)
                    t2h = ("<th>Empresa</th><th>Q Tobin</th><th>RPG (P/E)</th>"
                           "<th>YIELD</th><th>ROE</th><th>Margen EBITDA</th><th>EV/EBITDA</th>")
                    def _fmtm(v, pct=False):
                        if not isinstance(v,(int,float)) or np.isnan(v): return "N/A"
                        return fmt_num(v,2,es_pct=True) if pct else fmt_num(v,2)
                    # Mi acción (precio convergido)
                    t2r  = (f'<tr style="border-bottom:2px solid #2ECC71">'
                            +f'<td class="td-blue" style="font-weight:600">{tkr}</td>'
                            +f'<td>{_fmtm(qt_conv)}</td>'
                            +f'<td>{_fmtm(rpg_conv)}</td>'
                            +f'<td>{_fmtm(yld_conv,True)}</td>'
                            +f'<td>{_fmtm(roe_conv,True)}</td>'
                            +f'<td>{_fmtm(mg_conv,True)}</td>'
                            +f'<td>{_fmtm(eve_conv)}</td></tr>')
                    t2r += comp_rows
                    t2r += ('<tr style="border-top:2px solid #1A3A5C">'
                            +'<td class="td-neutral" style="font-weight:600">Promedio Sector</td>'
                            +f'<td class="td-neutral">{_fmtm(prom_qt)}</td>'
                            +f'<td class="td-neutral">{_fmtm(prom_rpg)}</td>'
                            +f'<td class="td-neutral">{_fmtm(prom_yld,True)}</td>'
                            +f'<td class="td-neutral">{_fmtm(prom_roe,True)}</td>'
                            +f'<td class="td-neutral">{_fmtm(prom_mg,True)}</td>'
                            +f'<td class="td-neutral">{_fmtm(prom_eve)}</td></tr>')
                    st.markdown('<table class="styled-table"><thead><tr>'+t2h
                                +'</tr></thead><tbody>'+t2r+'</tbody></table>',
                                unsafe_allow_html=True)

                    # ── Botón interpolación ────────────────────────────────────
                    st.markdown("<br>", unsafe_allow_html=True)
                    btn_interp = st.button("▶  Ejecutar Interpolación (100 iter.)",
                                           key=f"btn_interp_{tkr}", type="primary")
                    if btn_interp:
                        try:
                            precio_mdo_iter = precio_inicial  # F80 inicial
                            for _ in range(100):
                                # Calcular múltiplos con precio_mdo_iter actual
                                qt_i,rpg_i,yld_i,roe_i,mg_i,eve_i,_mc,_ev = _calc_mult_iter(precio_mdo_iter)
                                pesos_i = {"RPG":0.10,"YIELD":0.05,"ROE":0.15,
                                           "Mg EBITDA":0.30,"EV/EBITDA":0.40}
                                proms_i = {"RPG":prom_rpg,"YIELD":prom_yld,"ROE":prom_roe,
                                           "Mg EBITDA":prom_mg,"EV/EBITDA":prom_eve}
                                objs_i  = {"RPG":rpg_i,"YIELD":yld_i,"ROE":roe_i,
                                           "Mg EBITDA":mg_i,"EV/EBITDA":eve_i}
                                # Eficiencia ponderada
                                # RPG, EV/EBITDA: Promedio/Mi_acción - 1
                                # YIELD, ROE, Mg EBITDA: Mi_acción/Promedio - 1
                                _campos_inv = {"YIELD", "ROE", "Mg EBITDA"}
                                ef_p = 0.0
                                for campo,peso in pesos_i.items():
                                    ov=objs_i.get(campo,np.nan); pv=proms_i.get(campo,np.nan)
                                    if (isinstance(ov,float) and not np.isnan(ov) and ov!=0 and
                                        isinstance(pv,float) and not np.isnan(pv) and pv!=0):
                                        if campo in _campos_inv:
                                            ef_p += (ov/pv - 1)*peso
                                        else:
                                            ef_p += (pv/ov - 1)*peso
                                # C109: Precio sugerido = Q Tobin Sugerido × Vlr Libros
                                qt_sug_i = prom_qt*(1+ef_p) if not np.isnan(prom_qt) else np.nan
                                pr_sug_i = qt_sug_i*vlr_libros if not np.isnan(qt_sug_i) else np.nan
                                if pr_sug_i is None or np.isnan(pr_sug_i): break
                                # D109: ¿round(C109,2) == round(F80,2)?
                                if round(pr_sug_i,2) == round(precio_mdo_iter,2):
                                    break  # Convergió — igual que cuando D109=True en Excel
                                # Guard: si diverge más de 50x del precio inicial, detener
                                if pr_sug_i > precio_inicial * 50 or pr_sug_i < precio_inicial * 0.02:
                                    pr_sug_i = precio_mdo_iter  # mantener último valor razonable
                                    break
                                # Si no convergió: F80 = C109
                                precio_mdo_iter = pr_sug_i
                            st.session_state[precio_key] = precio_mdo_iter
                        except Exception as e_iter:
                            st.error(f"Error en iteración: {e_iter}")
                            precio_mdo_iter = precio_inicial
                        st.rerun()

                    # ── Tabla eficiencia (con precio convergido) ───────────────
                    st.markdown('<div class="section-header">Tabla de Eficiencia</div>',
                                unsafe_allow_html=True)
                    pesos_m = {"RPG":0.10,"YIELD":0.05,"ROE":0.15,
                               "Mg EBITDA":0.30,"EV/EBITDA":0.40}
                    proms_m = {"RPG":prom_rpg,"YIELD":prom_yld,"ROE":prom_roe,
                               "Mg EBITDA":prom_mg,"EV/EBITDA":prom_eve}
                    objs_m  = {"RPG":rpg_conv,"YIELD":yld_conv,"ROE":roe_conv,
                               "Mg EBITDA":mg_conv,"EV/EBITDA":eve_conv}
                    # RPG, EV/EBITDA: Promedio/Mi_acción - 1
                    # YIELD, ROE, Mg EBITDA: Mi_acción/Promedio - 1
                    campos_inversos = {"YIELD", "ROE", "Mg EBITDA"}
                    ef_vals_m={}; ef_pond_m={}
                    for campo,peso in pesos_m.items():
                        ov=objs_m.get(campo,np.nan); pv=proms_m.get(campo,np.nan)
                        if (isinstance(ov,float) and not np.isnan(ov) and ov!=0 and
                            isinstance(pv,float) and not np.isnan(pv) and pv!=0):
                            if campo in campos_inversos:
                                ef_vals_m[campo] = ov/pv - 1  # Mi_acción/Promedio - 1
                            else:
                                ef_vals_m[campo] = pv/ov - 1  # Promedio/Mi_acción - 1
                            ef_pond_m[campo] = ef_vals_m[campo]*peso
                        else:
                            ef_vals_m[campo]=np.nan; ef_pond_m[campo]=np.nan
                    ef_tot   = sum(v for v in ef_vals_m.values() if not np.isnan(v))
                    ef_p_tot = sum(v for v in ef_pond_m.values() if not np.isnan(v))
                    def _css_e(v):
                        if not isinstance(v,float) or np.isnan(v): return ""
                        return "td-positive" if v>0 else "td-negative"
                    t3h = '<th>Concepto</th>'+''.join(f'<th>{c}</th>' for c in pesos_m)+'<th>Total</th>'
                    p_row  = '<tr><td class="td-label">Pesos</td>'+''.join(f'<td>{fmt_num(p,2,es_pct=True)}</td>' for p in pesos_m.values())+'<td></td></tr>'
                    ef_row = ('<tr><td class="td-label">Eficiencia</td>'
                              +''.join(f'<td class="{_css_e(ef_vals_m[c])}">'
                                       +(_fmtm(ef_vals_m[c],True) if not np.isnan(ef_vals_m[c]) else "N/A")
                                       +'</td>' for c in pesos_m)
                              +f'<td class="{_css_e(ef_tot)}">{_fmtm(ef_tot,True)}</td></tr>')
                    ep_row = ('<tr><td class="td-label" style="font-weight:600">Ef. Ponderada</td>'
                              +''.join(f'<td class="{_css_e(ef_pond_m[c])}">'
                                       +(_fmtm(ef_pond_m[c],True) if not np.isnan(ef_pond_m[c]) else "N/A")
                                       +'</td>' for c in pesos_m)
                              +f'<td class="{_css_e(ef_p_tot)}" style="font-weight:600">'
                              +_fmtm(ef_p_tot,True)+'</td></tr>')
                    st.markdown('<table class="styled-table"><thead><tr>'+t3h
                                +'</tr></thead><tbody>'+p_row+ef_row+ep_row+'</tbody></table>',
                                unsafe_allow_html=True)

                    # ── Precio sugerido ────────────────────────────────────────
                    st.markdown('<div class="section-header">Precio Sugerido por Múltiplos</div>',
                                unsafe_allow_html=True)
                    qt_sug_show = prom_qt*(1+ef_p_tot) if not np.isnan(prom_qt) else np.nan
                    pr_sug_show = qt_sug_show*vlr_libros if not np.isnan(qt_sug_show) else np.nan
                    pot_mult_v  = (pr_sug_show/precio_inicial-1) if (pr_sug_show and precio_inicial) else np.nan
                    css_ps      = "td-positive" if (isinstance(pot_mult_v,float) and not np.isnan(pot_mult_v) and pot_mult_v>0) else "td-negative"
                    convergido  = precio_key in st.session_state
                    conv_ok     = (convergido and pr_sug_show and
                                   abs(round(pr_sug_show,2)-round(precio_conv,2))<0.01)
                    t5h = '<th>Concepto</th><th>Valor</th>'
                    t5r = (f'<tr><td class="td-label">Vlr Libros (fijo)</td><td>{_M(vlr_libros,"price")}</td></tr>'
                           +f'<tr><td class="td-label">Promedio Q Tobin Sector</td><td>{_fmtm(prom_qt)}</td></tr>'
                           +f'<tr><td class="td-label">Q Tobin Sugerido</td><td>{_fmtm(qt_sug_show)}</td></tr>'
                           +f'<tr style="border-top:2px solid #4AB3F4"><td style="font-weight:600;color:#E8F4FD">Precio Sugerido</td>'
                           +f'<td class="td-blue" style="font-weight:600">{_M(pr_sug_show,"price")}</td></tr>'
                           +f'<tr><td class="td-label">Precio Mercado (última iter.)</td>'
                           +f'<td>{_M(precio_conv,"price")}</td></tr>'
                           +f'<tr><td class="td-label">Potencial Múltiplos</td>'
                           +f'<td class="{css_ps}">{_M(pot_mult_v,"pct")}</td></tr>'
                           +(f'<tr><td class="td-label">Estado</td><td class="{"td-positive" if conv_ok else "td-negative"}" style="font-weight:600">{"✅ CONVERGIÓ" if conv_ok else "⚠️ No convergió"}</td></tr>'
                             if convergido else ''))
                    st.markdown('<table class="styled-table"><thead><tr>'+t5h
                                +'</tr></thead><tbody>'+t5r+'</tbody></table>',
                                unsafe_allow_html=True)
                    if not convergido:
                        st.info("Presiona **▶ Ejecutar Interpolación** para encontrar el precio convergido.")

                        # Señal final

                        # Señal final
                        st.markdown(f"""<div style="background:#0D1E35;border:1px solid #1A3A5C;
                            border-left:4px solid {tend_c};border-radius:6px;
                            padding:1rem 1.5rem;margin:1rem 0;font-family:IBM Plex Mono,monospace;">
                          <span style="font-size:.62rem;color:#4A7FA5;letter-spacing:.1em;text-transform:uppercase;">Señal Final — {tkr}</span><br>
                          <span style="font-size:1.3rem;font-weight:600;color:{tend_c};">{r.get("señal","N/A")}</span>
                        </div>""",unsafe_allow_html=True)
        else:
            st.markdown("""<div style="text-align:center;padding:4rem 2rem;color:#4A7FA5;
                 font-family:IBM Plex Mono,monospace;">
              <div style="font-size:3rem;margin-bottom:1rem;">🏦</div>
              <div style="font-size:.85rem;letter-spacing:.1em;text-transform:uppercase;">
                Configura los parámetros, edita la tabla y presiona<br>
                <span style="color:#4AB3F4;">PASO 3 — CALCULAR VALORACIÓN FUNDAMENTAL</span></div>
            </div>""",unsafe_allow_html=True)

# ── FUNDAMENTAL EXPORT FUNCTIONS ──────────────────────────────────────────────

def _construir_fundamental_xl():
    """Build fundamental valuation data for Excel export."""
    if "resultados_vf" not in st.session_state:
        return None
    return st.session_state["resultados_vf"]

def _exportar_fundamental_excel(resultados_vf):
    """Generate Excel workbook with fundamental valuation sheets."""
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    import io, numpy as _np

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Color palette
    HDR_FILL  = PatternFill("solid", fgColor="0B2545")
    HDR2_FILL = PatternFill("solid", fgColor="1A3A5C")
    BLUE_FILL = PatternFill("solid", fgColor="1B4F7E")
    GRN_FILL  = PatternFill("solid", fgColor="0B2E1A")
    ALT_FILL  = PatternFill("solid", fgColor="0D1E35")
    WHT_FONT  = Font(name="Arial", color="E8F4FD", bold=False, size=10)
    HDR_FONT  = Font(name="Arial", color="4AB3F4", bold=True, size=10)
    TTL_FONT  = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    GRN_FONT  = Font(name="Arial", color="2ECC71", bold=True, size=10)
    RED_FONT  = Font(name="Arial", color="E74C3C", bold=True, size=10)
    thin = Side(style="thin", color="1A3A5C")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _v(v):
        if v is None: return "N/A"
        if isinstance(v, float) and _np.isnan(v): return "N/A"
        return v

    def _fmt_n(v, decimals=2):
        if v is None or (isinstance(v, float) and _np.isnan(v)): return "N/A"
        return round(float(v), decimals)

    def hdr_row(ws, row, cols, title=None, fill=None):
        if title:
            ws.merge_cells(start_row=row, start_column=1,
                          end_row=row, end_column=len(cols))
            cell = ws.cell(row=row, column=1, value=title)
            cell.font = TTL_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
            row += 1
        for c, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=col)
            cell.font = HDR_FONT
            cell.fill = fill or HDR2_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = brd
        return row + 1

    def data_row(ws, row, vals, alt=False):
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=_fmt_n(v) if isinstance(v, float) else _v(v))
            cell.font = WHT_FONT
            cell.fill = ALT_FILL if alt else PatternFill("solid", fgColor="0A1628")
            cell.alignment = Alignment(horizontal="right" if c>1 else "left")
            cell.border = brd

    for tkr, r in resultados_vf.items():
        if "error" in r: continue

        # ── Hoja 1: FCF ────────────────────────────────────────────────────────
        ws1 = wb.create_sheet(f"{tkr} - FCF")
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions["A"].width = 36
        for col in ["B","C","D"]: ws1.column_dimensions[col].width = 22

        # Título
        ws1.merge_cells("A1:D1")
        t = ws1["A1"]; t.value = f"VALORACIÓN FUNDAMENTAL — FCF — {tkr}"
        t.font = Font(name="Arial", color="4AB3F4", bold=True, size=13)
        t.fill = HDR_FILL; t.alignment = Alignment(horizontal="center")
        ws1.row_dimensions[1].height = 28

        # Datos empresa
        row = hdr_row(ws1, 3, ["Concepto","Año Actual","Año Anterior","Notas"],
                     "DATOS DE LA EMPRESA", HDR_FILL)
        empresa_data = [
            ("Ticker",           tkr,                    "",         ""),
            ("Sector",           _v(r.get("sector")),    "",         ""),
            ("Industria",        _v(r.get("industria")), "",         ""),
            ("Moneda",           _v(r.get("moneda")),    "",         ""),
            ("Fecha Extracción", _v(r.get("fecha_extraccion")), "",  ""),
            ("Precio Mercado",   _fmt_n(r.get("precio"),2), "",      "USD"),
            ("# Acciones",       _fmt_n(r.get("num_acc"),0), "",     ""),
            ("Activo Total",     _fmt_n(r.get("activo")),_fmt_n(r.get("activo_ant")), ""),
            ("Pasivo Total",     _fmt_n(r.get("pasivo")),_fmt_n(r.get("pasivo_ant")), ""),
            ("Patrimonio",       _fmt_n(r.get("patrim")),_fmt_n(r.get("patrim_ant")), ""),
            ("Deuda Financiera", _fmt_n(r.get("deuda")), "",         ""),
            ("Efectivo",         _fmt_n(r.get("efectivo")), "",      ""),
            ("UO 12M (TTM)",     _fmt_n(r.get("ut_op")), "",         ""),
            ("DAP 12M (TTM)",    _fmt_n(r.get("dep_am")), "",        ""),
            ("EBITDA 12M (TTM)", _fmt_n(r.get("ebitda")), "",        ""),
            ("Ut. Neta 12M",     _fmt_n(r.get("ut_neta")), "",       ""),
            ("Ingresos Oper.",   _fmt_n(r.get("ing_op")), "",        ""),
            ("Gts. Financieros", _fmt_n(r.get("gts_fin")), "",       ""),
            ("CAPEX TTM",        _fmt_n(r.get("capex")), "",         ""),
            ("Inventario",       _fmt_n(r.get("inv0")), _fmt_n(r.get("inv1")), ""),
            ("CxC",              _fmt_n(r.get("cxc0")), _fmt_n(r.get("cxc1")), ""),
            ("CxP",              _fmt_n(r.get("cxp0")), _fmt_n(r.get("cxp1")), ""),
            ("PP&E",             _fmt_n(r.get("ppe0")), _fmt_n(r.get("ppe1")), ""),
            ("KTNO",             _fmt_n(r.get("ktno0")), _fmt_n(r.get("ktno1")), "Inv+CxC-CxP"),
        ]
        for i, vals in enumerate(empresa_data):
            data_row(ws1, row, vals, alt=(i%2==0)); row += 1

        # FCF
        row += 1
        row = hdr_row(ws1, row, ["Concepto","Valor","",""], "FREE CASH FLOW", BLUE_FILL)
        fcf_data = [
            ("Utilidad Operacional 12M",  _fmt_n(r.get("ut_op"))),
            ("(-) Impuesto de Renta",      _fmt_n(r.get("imp_rta"))),
            ("= UODI",                     _fmt_n(r.get("uodi"))),
            ("(-) Gas, Dep y Amor 12M",    _fmt_n(r.get("dep_am"))),
            ("= Flujo de Caja Bruto",      _fmt_n(r.get("fcb"))),
            ("(-) Variación KTNO",         _fmt_n(r.get("vktno"))),
            ("(-) Variación CAPEX (PP&E)", _fmt_n(r.get("vcap"))),
            ("= FREE CASH FLOW",           _fmt_n(r.get("fcf"))),
        ]
        for i, (label, val) in enumerate(fcf_data):
            cell_a = ws1.cell(row=row, column=1, value=label)
            cell_b = ws1.cell(row=row, column=2, value=val)
            cell_a.font = WHT_FONT; cell_b.font = WHT_FONT
            fill = BLUE_FILL if label.startswith("=") else (ALT_FILL if i%2==0 else PatternFill("solid", fgColor="0A1628"))
            cell_a.fill = fill; cell_b.fill = fill
            cell_a.border = brd; cell_b.border = brd
            if label == "= FREE CASH FLOW":
                cell_a.font = Font(name="Arial", color="2ECC71", bold=True, size=10)
                cell_b.font = Font(name="Arial", color="2ECC71", bold=True, size=10)
            row += 1

        # WACC
        row += 1
        row = hdr_row(ws1, row, ["Variable","Valor","",""], "ESTRUCTURA DE CAPITAL — WACC", HDR2_FILL)
        wacc_data = [
            ("Beta",              _fmt_n(r.get("beta"),4)),
            ("RF",                _fmt_n(r.get("ke"), 4)),
            ("Ke",                _fmt_n(r.get("ke"), 4)),
            ("Kd",                _fmt_n(r.get("kd"), 4)),
            ("W% Equity",         _fmt_n(r.get("we"), 4)),
            ("W% Deuda",          _fmt_n(r.get("wd"), 4)),
            ("WACC (iterado)",    _fmt_n(r.get("wacc"), 4)),
            ("G perpetuidad",     _fmt_n(r.get("G_perp"), 4)),
        ]
        for i, (label, val) in enumerate(wacc_data):
            data_row(ws1, row, [label, val, "", ""], alt=(i%2==0)); row += 1

        # ── Hoja 2: Proyección ──────────────────────────────────────────────────
        ws2 = wb.create_sheet(f"{tkr} - Proyección")
        ws2.sheet_view.showGridLines = False
        ws2.column_dimensions["A"].width = 30

        ws2.merge_cells(f"A1:H1")
        t2 = ws2["A1"]; t2.value = f"PROYECCIÓN FCF Y VALORACIÓN — {tkr}"
        t2.font = Font(name="Arial", color="4AB3F4", bold=True, size=13)
        t2.fill = HDR_FILL; t2.alignment = Alignment(horizontal="center")
        ws2.row_dimensions[1].height = 28

        proyt = r.get("proyt", [])
        n_cols = len(proyt) + 2  # concepto + años + VT
        for i in range(n_cols):
            col_l = get_column_letter(i+1)
            ws2.column_dimensions[col_l].width = 18

        # Header años
        row2 = 3
        headers = ["Concepto"] + [str(x["year"]) for x in proyt] + ["Vlr Terminal"]
        row2 = hdr_row(ws2, row2, headers, "PROYECCIÓN FCF", HDR_FILL)

        # G row
        g_vals = ["G año"] + [("—" if x["G"] is None else _fmt_n(x["G"],4)) for x in proyt] + [""]
        data_row(ws2, row2, g_vals); row2 += 1
        # FCF row
        fcf_vals = ["FCF"] + [_fmt_n(x["fcf"]) for x in proyt] + [_fmt_n(r.get("vt_val"))]
        data_row(ws2, row2, fcf_vals, alt=True); row2 += 1
        # VP row
        vp_vals = ["VP"] + [("—" if x["vp"] is None else _fmt_n(x["vp"])) for x in proyt] + [_fmt_n(r.get("vp_vt"))]
        data_row(ws2, row2, vp_vals); row2 += 1

        # Equity Value section
        row2 += 1
        row2 = hdr_row(ws2, row2, ["Concepto","Valor","",""],
                      "EQUITY VALUE — ENTERPRISE VALUE — PO", BLUE_FILL)
        ev_data = [
            ("VP Período Relevante",  _fmt_n(r.get("vp_sum"))),
            ("VP Valor Terminal",     _fmt_n(r.get("vp_vt"))),
            ("Equity Value",          _fmt_n(r.get("equity_value"))),
            ("(+) Otros Activos",     _fmt_n(r.get("otros_activos"))),
            ("(-) Otros Pasivos",     _fmt_n(r.get("otros_pasivos"))),
            ("Enterprise Value",      _fmt_n(r.get("enterprise_value"))),
            ("Acciones Circulación",  _fmt_n(r.get("num_acc"),0)),
            ("PO (USD)",              _fmt_n(r.get("po_usd"),2)),
            ("PO (Local)",            _fmt_n(r.get("po_cop"),2)),
            ("Precio Actual",         _fmt_n(r.get("precio"),2)),
            ("Potencial FCF",         _fmt_n(r.get("pot_fcf"),4)),
        ]
        for i, (label, val) in enumerate(ev_data):
            data_row(ws2, row2, [label, val, "", ""], alt=(i%2==0)); row2 += 1

        # ── Hoja 3: Múltiplos ───────────────────────────────────────────────────
        ws3 = wb.create_sheet(f"{tkr} - Múltiplos")
        ws3.sheet_view.showGridLines = False
        ws3.column_dimensions["A"].width = 30

        ws3.merge_cells("A1:H1")
        t3 = ws3["A1"]; t3.value = f"VALORACIÓN POR MÚLTIPLOS — {tkr}"
        t3.font = Font(name="Arial", color="2ECC71", bold=True, size=13)
        t3.fill = GRN_FILL; t3.alignment = Alignment(horizontal="center")
        ws3.row_dimensions[1].height = 28

        for col in ["B","C","D","E","F","G","H"]: ws3.column_dimensions[col].width = 18

        # Tabla datos empresa
        row3 = 3
        t1_cols = ["Empresa","Patrimonio","MKT CAP","#Acciones","Vlr Libros",
                  "Precio Mdo","Dividendo","UPA 12","Utl Neta 12",
                  "Ing Ope 12","EBITDA 12","Deuda Fin","Efectivo","EV"]
        row3 = hdr_row(ws3, row3, t1_cols, "DATOS DE LA EMPRESA", GRN_FILL)
        precio_c = r.get("precio",0) or 0
        n_acc    = r.get("num_acc",1) or 1
        pat_disp = precio_c * n_acc
        mc_conv  = st.session_state.get(f"precio_conv_{tkr}", precio_c) * n_acc
        vlr_lib  = precio_c
        upa_v    = (r.get("ut_neta",0) or 0) / n_acc
        ev_v     = mc_conv + (r.get("deuda",0) or 0) - (r.get("efectivo",0) or 0)
        empresa_row = [tkr, _fmt_n(pat_disp), _fmt_n(mc_conv), _fmt_n(n_acc,0),
                      _fmt_n(vlr_lib,2), _fmt_n(precio_c,2),
                      _fmt_n(r.get("dps",0),2), _fmt_n(upa_v,2),
                      _fmt_n(r.get("ut_neta")), _fmt_n(r.get("ing_op")),
                      _fmt_n(r.get("ebitda")), _fmt_n(r.get("deuda")),
                      _fmt_n(r.get("efectivo")), _fmt_n(ev_v)]
        data_row(ws3, row3, empresa_row); row3 += 2

        # Tabla múltiplos
        t2_cols = ["Empresa","Q Tobin","RPG (P/E)","YIELD","ROE","Margen EBITDA","EV/EBITDA"]
        row3 = hdr_row(ws3, row3, t2_cols, "MÚLTIPLOS", HDR2_FILL)

        # Mi acción
        precio_cv = st.session_state.get(f"precio_conv_{tkr}", precio_c)
        qt_o = precio_cv / vlr_lib if vlr_lib else None
        rpg_o = precio_cv / upa_v if (upa_v and upa_v>0) else None
        dps_v = r.get("dps",0) or 0
        yld_o = (dps_v/precio_cv if dps_v>0 else upa_v/precio_cv) if precio_cv else None
        roe_o = (r.get("ut_neta",0) or 0)/pat_disp if pat_disp else None
        mg_o  = (r.get("ebitda",0) or 0)/(r.get("ing_op",1) or 1)
        eve_o = ev_v / (r.get("ebitda",1) or 1)
        obj_row = [tkr, _fmt_n(qt_o,4), _fmt_n(rpg_o,4), _fmt_n(yld_o,4),
                  _fmt_n(roe_o,4), _fmt_n(mg_o,4), _fmt_n(eve_o,4)]
        for c, v in enumerate(obj_row, 1):
            cell = ws3.cell(row=row3, column=c, value=v)
            cell.font = GRN_FONT; cell.fill = GRN_FILL; cell.border = brd
        row3 += 1

        # Comparables
        comps = r.get("comps", [])
        qt_l=[]; rpg_l=[]; yld_l=[]; roe_l=[]; mg_l=[]; eve_l=[]
        for i, comp in enumerate(comps):
            qt_c=comp.get("_qt"); rpg_c=comp.get("_rpg")
            yld_c=comp.get("_yield"); roe_c=comp.get("_roe")
            mg_c=comp.get("_mg_ebitda"); eve_c=comp.get("_ev_ebitda")
            for lst,v in zip([qt_l,rpg_l,yld_l,roe_l,mg_l,eve_l],
                             [qt_c,rpg_c,yld_c,roe_c,mg_c,eve_c]):
                if v and not (isinstance(v,float) and _np.isnan(v)) and v>0: lst.append(v)
            comp_row = [comp.get("Ticker",""), _fmt_n(qt_c,4), _fmt_n(rpg_c,4),
                       _fmt_n(yld_c,4), _fmt_n(roe_c,4), _fmt_n(mg_c,4), _fmt_n(eve_c,4)]
            data_row(ws3, row3, comp_row, alt=(i%2==0)); row3 += 1

        # Promedio
        def _pm(lst): return round(sum(lst)/len(lst),4) if lst else "N/A"
        prom_row = ["Promedio Sector",_pm(qt_l),_pm(rpg_l),_pm(yld_l),
                   _pm(roe_l),_pm(mg_l),_pm(eve_l)]
        for c, v in enumerate(prom_row, 1):
            cell = ws3.cell(row=row3, column=c, value=v)
            cell.font = Font(name="Arial", color="F39C12", bold=True, size=10)
            cell.fill = HDR2_FILL; cell.border = brd
        row3 += 2

        # Eficiencia
        pesos = {"RPG":0.10,"YIELD":0.05,"ROE":0.15,"Mg EBITDA":0.30,"EV/EBITDA":0.40}
        proms = {"RPG":_pm(rpg_l),"YIELD":_pm(yld_l),"ROE":_pm(roe_l),
                "Mg EBITDA":_pm(mg_l),"EV/EBITDA":_pm(eve_l)}
        objs  = {"RPG":rpg_o,"YIELD":yld_o,"ROE":roe_o,"Mg EBITDA":mg_o,"EV/EBITDA":eve_o}
        _inv  = {"YIELD","ROE","Mg EBITDA"}

        ef_cols = ["Concepto"] + list(pesos.keys()) + ["Total"]
        row3 = hdr_row(ws3, row3, ef_cols, "TABLA DE EFICIENCIA", HDR_FILL)
        peso_row_v = ["Pesos"] + [p for p in pesos.values()] + [""]
        data_row(ws3, row3, peso_row_v); row3 += 1

        ef_v = {}; ef_p_v = {}
        for campo, peso in pesos.items():
            ov=objs.get(campo); pv=proms.get(campo)
            if (ov and pv and isinstance(ov,float) and isinstance(pv,float)
                and not _np.isnan(ov) and not _np.isnan(pv) and ov!=0 and pv!=0):
                ef_v[campo] = ov/pv-1 if campo in _inv else pv/ov-1
                ef_p_v[campo] = ef_v[campo]*peso
            else:
                ef_v[campo]=None; ef_p_v[campo]=None

        ef_tot = sum(v for v in ef_v.values() if v is not None)
        ef_p_tot = sum(v for v in ef_p_v.values() if v is not None)

        ef_row_v  = ["Eficiencia"] + [_fmt_n(ef_v.get(c),4) for c in pesos] + [_fmt_n(ef_tot,4)]
        efp_row_v = ["Ef. Ponderada"] + [_fmt_n(ef_p_v.get(c),4) for c in pesos] + [_fmt_n(ef_p_tot,4)]
        data_row(ws3, row3, ef_row_v, alt=True); row3 += 1
        data_row(ws3, row3, efp_row_v); row3 += 2

        # Q Tobin sugerido y precio
        qt_sug = _pm(qt_l)*(1+ef_p_tot) if qt_l and ef_p_tot else None
        pr_sug = qt_sug * vlr_lib if qt_sug else None
        row3 = hdr_row(ws3, row3, ["Concepto","Valor","",""],
                      "PRECIO SUGERIDO POR MÚLTIPLOS", GRN_FILL)
        sug_data = [
            ("Vlr Libros (fijo)",        _fmt_n(vlr_lib,2)),
            ("Promedio Q Tobin Sector",  _fmt_n(_pm(qt_l),4)),
            ("Q Tobin Sugerido",         _fmt_n(qt_sug,4)),
            ("Precio Sugerido",          _fmt_n(pr_sug,2)),
            ("Precio Convergido",        _fmt_n(st.session_state.get(f"precio_conv_{tkr}", precio_c),2)),
            ("Precio Inicial",           _fmt_n(precio_c,2)),
        ]
        for i, (label, val) in enumerate(sug_data):
            data_row(ws3, row3, [label, val, "", ""], alt=(i%2==0)); row3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generar_pdf_fundamental(resultados_vf):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, PageBreak, HRFlowable,
                                     KeepTogether)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing, Rect, String, Line
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    import io, numpy as _np, datetime as _dt

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=0.55*inch, rightMargin=0.55*inch,
                            topMargin=0.55*inch, bottomMargin=0.55*inch)

    # ── Palette ────────────────────────────────────────────────────────────────
    C_BG    = colors.HexColor("#0B2545")
    C_MID   = colors.HexColor("#1A3A5C")
    C_BLUE  = colors.HexColor("#4AB3F4")
    C_GREEN = colors.HexColor("#2ECC71")
    C_RED   = colors.HexColor("#E74C3C")
    C_AMBER = colors.HexColor("#F39C12")
    C_WHITE = colors.white
    C_LGRAY = colors.HexColor("#E8F4FD")
    C_MGRAY = colors.HexColor("#7FA8C9")
    C_ALT   = colors.HexColor("#0D1E35")

    def señal_color(s):
        if "COMPRAR" in str(s): return C_GREEN
        if "VENDER"  in str(s): return C_RED
        if "POSIBLE" in str(s) or "MANTENER" in str(s): return C_AMBER
        return C_BLUE

    def _fv(v, fmt="f"):
        if v is None or (isinstance(v,float) and _np.isnan(v)): return "N/A"
        try:
            f = float(v)
            if fmt=="pct": return f"{f*100:.2f}%"
            if fmt=="usd": return f"${f:,.2f}"
            if fmt=="big":
                av=abs(f); s="-" if f<0 else ""
                if av>=1e9: return f"{s}${av/1e9:.2f}B"
                if av>=1e6: return f"{s}${av/1e6:.2f}M"
                return f"${f:,.2f}"
            if fmt=="x": return f"{f:.2f}x"
            if fmt=="ratio": return f"{f:.1%}"
            return f"{f:.4f}"
        except: return str(v)

    # ── Styles ─────────────────────────────────────────────────────────────────
    def sty(name, **kw):
        base = getSampleStyleSheet()["Normal"]
        return ParagraphStyle(name, parent=base, **kw)

    S_TITLE   = sty("title",  fontSize=24, textColor=C_BLUE,   fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4)
    S_TICKER  = sty("ticker", fontSize=36, textColor=C_WHITE,  fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=2)
    S_SUB     = sty("sub",    fontSize=10, textColor=C_MGRAY,  fontName="Helvetica",       alignment=TA_CENTER, spaceAfter=12)
    S_H1      = sty("h1",     fontSize=12, textColor=C_WHITE,  fontName="Helvetica-Bold",  alignment=TA_CENTER, spaceAfter=4, backColor=C_MID, borderPadding=6)
    S_H2      = sty("h2",     fontSize=10, textColor=C_BLUE,   fontName="Helvetica-Bold",  spaceAfter=3)
    S_BODY    = sty("body",   fontSize=9,  textColor=C_LGRAY,  fontName="Helvetica",       leading=13, spaceAfter=6)
    S_FOOTER  = sty("footer", fontSize=7,  textColor=C_MGRAY,  fontName="Helvetica",       alignment=TA_CENTER)
    S_CAPTION = sty("cap",    fontSize=7,  textColor=C_MGRAY,  fontName="Helvetica",       alignment=TA_CENTER, spaceAfter=4)

    thin = 0.3
    GRID = colors.HexColor("#1A3A5C")

    def make_kpi_bar(labels, values, colors_list, width=7*inch, height=0.9*inch):
        """Horizontal KPI bar with colored boxes."""
        d = Drawing(width, height)
        n = len(labels)
        w = width / n
        for i,(lbl,val,col) in enumerate(zip(labels,values,colors_list)):
            x = i*w
            d.add(Rect(x+2, 2, w-4, height-4, fillColor=C_MID, strokeColor=GRID, strokeWidth=0.5))
            d.add(String(x+w/2, height-16, str(lbl), fontSize=7, textAnchor="middle",
                        fillColor=C_MGRAY, fontName="Helvetica"))
            d.add(String(x+w/2, height/2-6, str(val), fontSize=11, textAnchor="middle",
                        fillColor=col, fontName="Helvetica-Bold"))
        return d

    def make_bar_chart(categories, series_data, series_labels, series_colors,
                      width=6*inch, height=2.2*inch):
        """Grouped bar chart comparing values."""
        d = Drawing(width, height)
        bc = VerticalBarChart()
        bc.x = 50; bc.y = 30
        bc.width  = width - 65
        bc.height = height - 45
        bc.data   = series_data
        bc.categoryAxis.categoryNames = categories
        bc.categoryAxis.labels.fontSize = 8
        bc.categoryAxis.labels.fillColor = C_LGRAY
        bc.categoryAxis.labels.fontName = "Helvetica"
        bc.valueAxis.labels.fontSize    = 8
        bc.valueAxis.labels.fillColor   = C_LGRAY
        bc.valueAxis.labels.fontName    = "Helvetica"
        bc.groupSpacing = 5
        bc.barSpacing   = 2
        for i, col in enumerate(series_colors):
            bc.bars[i].fillColor   = col
            bc.bars[i].strokeColor = C_BG
            bc.bars[i].strokeWidth = 0.3
        d.add(bc)
        return d

    def tbl(data, col_widths, header_rows=1, highlight_last=False):
        """Styled dark table."""
        t = Table(data, colWidths=col_widths)
        style = [
            ("BACKGROUND", (0,0), (-1, header_rows-1), C_MID),
            ("TEXTCOLOR",  (0,0), (-1, header_rows-1), C_BLUE),
            ("FONTNAME",   (0,0), (-1, header_rows-1), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("FONTNAME",   (0,header_rows), (-1,-1), "Helvetica"),
            ("TEXTCOLOR",  (0,header_rows), (-1,-1), C_LGRAY),
            ("ROWBACKGROUNDS",(0,header_rows),(-1,-1),[C_BG, C_ALT]),
            ("GRID",       (0,0), (-1,-1), thin, GRID),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN",      (1,0), (-1,-1), "RIGHT"),
            ("ALIGN",      (0,0), (0,-1), "LEFT"),
            ("TOPPADDING",    (0,0),(-1,-1), 3),
            ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ]
        if highlight_last and len(data) > header_rows:
            style += [
                ("BACKGROUND", (0,-1),(-1,-1), C_MID),
                ("TEXTCOLOR",  (0,-1),(-1,-1), C_AMBER),
                ("FONTNAME",   (0,-1),(-1,-1), "Helvetica-Bold"),
            ]
        t.setStyle(TableStyle(style))
        return t

    # ── BUILD STORY ────────────────────────────────────────────────────────────
    story = []
    today = _dt.datetime.today().strftime("%d/%m/%Y")

    for tkr, r in resultados_vf.items():
        if "error" in r: continue

        señal_m  = r.get("señal_mult","N/A")
        señal_f  = r.get("señal_fcf","N/A")
        sc_m     = señal_color(señal_m)
        precio_a = r.get("precio",0) or 0
        po_usd   = r.get("po_usd")
        wacc_v   = r.get("wacc")
        g_perp   = r.get("G_perp")
        fcf_v    = r.get("fcf")
        precio_cv= st.session_state.get(f"precio_conv_{tkr}", precio_a)
        ratio_m  = r.get("ratio_mult", np.nan)
        pot_fcf  = r.get("pot_fcf")

        # ══════════════════════════════════════════════════════════════════════
        # PORTADA
        # ══════════════════════════════════════════════════════════════════════
        story.append(Spacer(1, 0.4*inch))

        # Header band
        header_data = [[Paragraph(f"REPORTE DE VALORACIÓN FUNDAMENTAL", S_TITLE)]]
        ht = Table(header_data, colWidths=[7*inch])
        ht.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), C_BG),
            ("TOPPADDING",(0,0),(-1,-1), 10),
            ("BOTTOMPADDING",(0,0),(-1,-1), 10),
        ]))
        story.append(ht)
        story.append(Spacer(1, 0.1*inch))

        story.append(Paragraph(tkr, S_TICKER))
        story.append(Paragraph(
            f"{r.get('nombre','')} &nbsp;·&nbsp; {r.get('sector','')} &nbsp;·&nbsp; {r.get('moneda','USD')} &nbsp;·&nbsp; {today}",
            S_SUB))

        story.append(HRFlowable(width="100%", thickness=2, color=sc_m, spaceAfter=12, spaceBefore=4))

        # Señal grande
        señal_box = Table([[Paragraph(f"Señal: {señal_m}", sty("sb",
            fontSize=18, textColor=sc_m, fontName="Helvetica-Bold", alignment=TA_CENTER))]],
            colWidths=[7*inch])
        señal_box.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), C_BG),
            ("TOPPADDING",(0,0),(-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ("BOX",(0,0),(-1,-1), 1.5, sc_m),
        ]))
        story.append(señal_box)
        story.append(Spacer(1, 0.15*inch))

        # KPI bar
        kpi_labels = ["Precio Actual","PO (FCF)","Precio Múltiplos","WACC","FCF"]
        kpi_values = [_fv(precio_a,"usd"), _fv(po_usd,"usd"), _fv(precio_cv,"usd"),
                     _fv(wacc_v,"pct"), _fv(fcf_v,"big")]
        kpi_colors = [C_BLUE, C_GREEN if (po_usd and po_usd>precio_a) else C_RED,
                     sc_m, C_BLUE, C_GREEN if (fcf_v and fcf_v>0) else C_RED]
        story.append(make_kpi_bar(kpi_labels, kpi_values, kpi_colors))
        story.append(Spacer(1, 0.15*inch))

        # Datos empresa mini-tabla
        emp_data = [
            ["Empresa", r.get("nombre","N/A"), "Industria", r.get("industria","N/A")],
            ["Sector",  r.get("sector","N/A"),  "Moneda",    r.get("moneda","USD")],
            ["Fecha",   r.get("fecha_extraccion","N/A"), "País", r.get("pais","N/A")],
        ]
        et = Table(emp_data, colWidths=[1.2*inch,2.3*inch,1.2*inch,2.3*inch])
        et.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(0,-1), C_MID),("BACKGROUND",(2,0),(2,-1), C_MID),
            ("TEXTCOLOR",(0,0),(0,-1), C_BLUE),("TEXTCOLOR",(2,0),(2,-1), C_BLUE),
            ("TEXTCOLOR",(1,0),(1,-1), C_LGRAY),("TEXTCOLOR",(3,0),(3,-1), C_LGRAY),
            ("FONTNAME",(0,0),(-1,-1),"Helvetica"),("FONTSIZE",(0,0),(-1,-1),8),
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
            ("BACKGROUND",(1,0),(1,-1),C_BG),("BACKGROUND",(3,0),(3,-1),C_BG),
            ("GRID",(0,0),(-1,-1),thin,GRID),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ]))
        story.append(et)

        # ══════════════════════════════════════════════════════════════════════
        # SECCIÓN 1 — PORTAFOLIO (si disponible)
        # ══════════════════════════════════════════════════════════════════════
        if "port_calc" in st.session_state:
            story.append(PageBreak())
            story.append(Paragraph("SECCIÓN 1 — PORTAFOLIO ÓPTIMO", S_H1))
            story.append(Spacer(1,0.1*inch))

            pc = st.session_state["port_calc"]
            tks_port = [k for k in pc.get("rets_e",{}).keys()]
            pesos_p  = pc.get("pesos_norm",[])
            rets_e   = pc.get("rets_e",{})
            ret_p    = pc.get("ret_p",np.nan)
            vol_p    = pc.get("vol_p",np.nan)
            sh_p     = pc.get("sh_p",np.nan)

            story.append(Paragraph("Composición Óptima del Portafolio", S_H2))
            port_rows = [["Activo","% Asignación","Retorno Esp."]]
            for i,t in enumerate(tks_port):
                w = pesos_p[i]*100 if i<len(pesos_p) else 0
                re= rets_e.get(t,0)*100
                port_rows.append([t, f"{w:.1f}%", f"{re:.2f}%"])
            port_rows.append(["PORTAFOLIO", "100.0%", f"{(ret_p or 0)*100:.2f}%"])
            story.append(tbl(port_rows, [1.5*inch,2*inch,2*inch], highlight_last=True))
            story.append(Spacer(1,0.1*inch))

            # Métricas portafolio
            port_kpi = [["Retorno Esp.","Volatilidad","Sharpe"],
                       [f"{(ret_p or 0)*100:.2f}%", f"{(vol_p or 0)*100:.2f}%", f"{(sh_p or 0):.4f}"]]
            story.append(tbl(port_kpi, [2.33*inch,2.33*inch,2.34*inch]))

            # Bar chart pesos
            if tks_port and pesos_p:
                story.append(Spacer(1,0.1*inch))
                story.append(Paragraph("Distribución del Portafolio", S_H2))
                try:
                    chart_data = [[pesos_p[i]*100 if i<len(pesos_p) else 0 for i in range(len(tks_port))]]
                    chart = make_bar_chart(tks_port, chart_data, ["Peso (%)"],
                                         [C_BLUE], width=6.5*inch, height=2*inch)
                    story.append(chart)
                    story.append(Paragraph("Porcentaje óptimo por activo (Markowitz)", S_CAPTION))
                except: pass

        # ══════════════════════════════════════════════════════════════════════
        # SECCIÓN 2 — FCF
        # ══════════════════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph("SECCIÓN 2 — VALORACIÓN POR FLUJO DE CAJA LIBRE", S_H1))
        story.append(Spacer(1,0.1*inch))

        col_left, col_right = [], []

        # FCF cálculo
        story.append(Paragraph("Cálculo del FCF", S_H2))
        fcf_rows = [["Concepto","Valor"],
            ["UO 12M",               _fv(r.get("ut_op"),"big")],
            ["(-) Impuesto Renta",   _fv(r.get("imp_rta"),"big")],
            ["= UODI",               _fv(r.get("uodi"),"big")],
            ["(-) Dep, Amor y Prov", _fv(r.get("dep_am"),"big")],
            ["= Flujo Caja Bruto",   _fv(r.get("fcb"),"big")],
            ["(-) Var KTNO",         _fv(r.get("vktno"),"big")],
            ["(-) Var CAPEX",        _fv(r.get("vcap"),"big")],
            ["= FREE CASH FLOW",     _fv(r.get("fcf"),"big")],
        ]
        ft = tbl(fcf_rows,[3.2*inch,2.5*inch])
        ft.setStyle(TableStyle([
            ("BACKGROUND",(0,8),(-1,8), colors.HexColor("#0B2E1A")),
            ("TEXTCOLOR",(0,8),(-1,8), C_GREEN),
            ("FONTNAME",(0,8),(-1,8),"Helvetica-Bold"),
        ]))
        story.append(ft)
        story.append(Spacer(1,0.1*inch))

        # WACC
        story.append(Paragraph("Estructura de Capital — WACC", S_H2))
        wacc_rows = [["Variable","Valor","Variable","Valor"],
            ["Beta",  _fv(r.get("beta")), "Ke", _fv(r.get("ke"),"pct")],
            ["Kd",    _fv(r.get("kd"),"pct"), "W% Eq.", _fv(r.get("we"),"pct")],
            ["WACC",  _fv(r.get("wacc"),"pct"), "G Perp.", _fv(r.get("G_perp"),"pct")],
        ]
        story.append(tbl(wacc_rows,[1.5*inch,1.5*inch,1.5*inch,1.5*inch]))
        story.append(Spacer(1,0.1*inch))

        # Proyección
        story.append(Paragraph("Proyección FCF", S_H2))
        proyt = r.get("proyt",[])
        if proyt:
            ph  = [""] + [str(x["year"]) for x in proyt] + ["VT"]
            gr  = ["G"]   + [("—" if x["G"] is None else f"{x['G']*100:.1f}%") for x in proyt] + [""]
            fcr = ["FCF"] + [_fv(x["fcf"],"big") for x in proyt] + [_fv(r.get("vt_val"),"big")]
            vpr = ["VP"]  + [("—" if x["vp"] is None else _fv(x["vp"],"big")) for x in proyt] + [_fv(r.get("vp_vt"),"big")]
            n_p = len(proyt)+2
            cw_p= [0.9*inch] + [0.85*inch]*(len(proyt)) + [1.1*inch]
            story.append(tbl([ph,gr,fcr,vpr], cw_p))
        story.append(Spacer(1,0.1*inch))

        # EV/PO
        story.append(Paragraph("Enterprise Value y Precio Objetivo", S_H2))
        ev_rows = [["Concepto","Valor"],
            ["Equity Value",      _fv(r.get("equity_value"),"big")],
            ["(+) Otros Activos", _fv(r.get("otros_activos"),"big")],
            ["(-) Otros Pasivos", _fv(r.get("otros_pasivos"),"big")],
            ["Enterprise Value",  _fv(r.get("enterprise_value"),"big")],
            ["Acciones Circ.",    f"{(r.get('num_acc') or 0):,.0f}"],
            ["PO (USD)",          _fv(r.get("po_usd"),"usd")],
            ["Potencial FCF",     _fv(r.get("pot_fcf"),"pct")],
        ]
        et2 = tbl(ev_rows,[3.2*inch,2.5*inch])
        story.append(et2)

        # Bar chart: precio actual vs PO
        story.append(Spacer(1,0.1*inch))
        story.append(Paragraph("Comparativo de Precios", S_H2))
        labels_p = ["Precio Actual","PO (FCF)","Precio Múltiplos"]
        vals_p   = [v for v in [precio_a, po_usd, precio_cv] if v and not (isinstance(v,float) and _np.isnan(v))]
        lbls_p   = labels_p[:len(vals_p)]
        if len(vals_p) >= 2:
            try:
                chart2 = make_bar_chart(lbls_p, [vals_p], ["Precio USD"],
                                       [C_BLUE], width=6.5*inch, height=2*inch)
                story.append(chart2)
                story.append(Paragraph("Precio actual vs precios objetivo (USD)", S_CAPTION))
            except: pass

        # Señal FCF
        s_fcf_col = señal_color(señal_f)
        sfb = Table([[Paragraph(f"Señal FCF: {señal_f}", sty("sf",
            fontSize=12, textColor=s_fcf_col, fontName="Helvetica-Bold", alignment=TA_CENTER))]],
            colWidths=[7*inch])
        sfb.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),C_BG),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("BOX",(0,0),(-1,-1),1,s_fcf_col)]))
        story.append(Spacer(1,0.1*inch))
        story.append(sfb)

        # ══════════════════════════════════════════════════════════════════════
        # SECCIÓN 3 — MÚLTIPLOS
        # ══════════════════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph("SECCIÓN 3 — VALORACIÓN POR MÚLTIPLOS COMPARABLES", S_H1))
        story.append(Spacer(1,0.1*inch))

        # Tabla múltiplos
        story.append(Paragraph("Múltiplos del Sector", S_H2))
        comps   = r.get("comps",[])
        n_acc   = r.get("num_acc",1) or 1
        pat_d   = precio_a * n_acc
        vlr_l   = precio_a
        upa_v   = (r.get("ut_neta",0) or 0)/n_acc
        dps_v   = r.get("dps",0) or 0
        ev_ov   = precio_cv*n_acc + (r.get("deuda",0) or 0) - (r.get("efectivo",0) or 0)
        qt_ov   = precio_cv/vlr_l if vlr_l else None
        rpg_ov  = precio_cv/upa_v if (upa_v and upa_v>0) else None
        yld_ov  = (dps_v/precio_cv if dps_v>0 else upa_v/precio_cv) if precio_cv else None
        roe_ov  = (r.get("ut_neta",0) or 0)/pat_d if pat_d else None
        mg_ov   = (r.get("ebitda",0) or 0)/(r.get("ing_op",1) or 1)
        eve_ov  = ev_ov/(r.get("ebitda",1) or 1)

        mult_hdr = ["Empresa","Q Tobin","RPG","YIELD","ROE","Mg EBITDA","EV/EBITDA"]
        mult_rows2= [mult_hdr,
                    [tkr, _fv(qt_ov), _fv(rpg_ov), _fv(yld_ov,"pct") if yld_ov else "N/A",
                     _fv(roe_ov,"pct") if roe_ov else "N/A", _fv(mg_ov,"pct"), _fv(eve_ov)]]
        qt_l2=[]; rpg_l2=[]; yld_l2=[]; roe_l2=[]; mg_l2=[]; eve_l2=[]
        for comp in comps:
            qt_c=comp.get("_qt"); rpg_c=comp.get("_rpg"); yld_c=comp.get("_yield")
            roe_c=comp.get("_roe"); mg_c=comp.get("_mg_ebitda"); eve_c=comp.get("_ev_ebitda")
            for lst,v in zip([qt_l2,rpg_l2,yld_l2,roe_l2,mg_l2,eve_l2],[qt_c,rpg_c,yld_c,roe_c,mg_c,eve_c]):
                if v and not(isinstance(v,float) and _np.isnan(v)) and v>0: lst.append(v)
            mult_rows2.append([comp.get("Ticker",""),_fv(qt_c),_fv(rpg_c),
                              _fv(yld_c,"pct") if yld_c else "N/A",
                              _fv(roe_c,"pct") if roe_c else "N/A",
                              _fv(mg_c,"pct") if mg_c else "N/A",_fv(eve_c)])

        def _pm2(lst): return sum(lst)/len(lst) if lst else None
        mult_rows2.append(["Promedio",_fv(_pm2(qt_l2)),_fv(_pm2(rpg_l2)),
                          _fv(_pm2(yld_l2),"pct") if _pm2(yld_l2) else "N/A",
                          _fv(_pm2(roe_l2),"pct") if _pm2(roe_l2) else "N/A",
                          _fv(_pm2(mg_l2),"pct") if _pm2(mg_l2) else "N/A",_fv(_pm2(eve_l2))])
        cw_m = [1.4*inch]+[0.87*inch]*6
        mt2  = tbl(mult_rows2, cw_m, highlight_last=True)
        story.append(mt2)
        story.append(Spacer(1,0.1*inch))

        # Eficiencia
        story.append(Paragraph("Tabla de Eficiencia Ponderada", S_H2))
        pesos2 = {"RPG":0.10,"YIELD":0.05,"ROE":0.15,"Mg EBITDA":0.30,"EV/EBITDA":0.40}
        proms2 = {"RPG":_pm2(rpg_l2),"YIELD":_pm2(yld_l2),"ROE":_pm2(roe_l2),
                 "Mg EBITDA":_pm2(mg_l2),"EV/EBITDA":_pm2(eve_l2)}
        objs2  = {"RPG":rpg_ov,"YIELD":yld_ov,"ROE":roe_ov,"Mg EBITDA":mg_ov,"EV/EBITDA":eve_ov}
        _inv2  = {"YIELD","ROE","Mg EBITDA"}
        ef2={}; efp2={}
        for campo,peso in pesos2.items():
            ov=objs2.get(campo); pv=proms2.get(campo)
            if ov and pv and not(isinstance(ov,float) and _np.isnan(ov)) and not(isinstance(pv,float) and _np.isnan(pv)) and ov!=0 and pv!=0:
                ef2[campo]=ov/pv-1 if campo in _inv2 else pv/ov-1
                efp2[campo]=ef2[campo]*peso
            else: ef2[campo]=None; efp2[campo]=None
        ef_tot2   = sum(v for v in ef2.values() if v is not None)
        ef_p_tot2 = sum(v for v in efp2.values() if v is not None)

        ef_tbl2 = [
            [""]+list(pesos2.keys())+["Total"],
            ["Pesos"]+[f"{p*100:.0f}%" for p in pesos2.values()]+["100%"],
            ["Eficiencia"]+[_fv(ef2.get(c),"pct") if ef2.get(c) is not None else "N/A" for c in pesos2]+[_fv(ef_tot2,"pct")],
            ["Ef.Pond."]+[_fv(efp2.get(c),"pct") if efp2.get(c) is not None else "N/A" for c in pesos2]+[_fv(ef_p_tot2,"pct")],
        ]
        story.append(tbl(ef_tbl2,[1.2*inch]+[0.87*inch]*5+[0.83*inch]))
        story.append(Spacer(1,0.1*inch))

        # Precio sugerido
        story.append(Paragraph("Precio Sugerido por Múltiplos (Convergido)", S_H2))
        qt_sug2 = _pm2(qt_l2)*(1+ef_p_tot2) if qt_l2 else None
        pr_sug2 = qt_sug2*vlr_l if qt_sug2 else None
        pot_m   = (pr_sug2/precio_a-1) if (pr_sug2 and precio_a) else None
        sug_rows= [["Concepto","Valor"],
            ["Vlr Libros (fijo)",   _fv(vlr_l,"usd")],
            ["Q Tobin Sector",      _fv(_pm2(qt_l2))],
            ["Q Tobin Sugerido",    _fv(qt_sug2)],
            ["Precio Sugerido",     _fv(pr_sug2,"usd")],
            ["Precio Convergido",   _fv(precio_cv,"usd")],
            ["Ratio vs Precio Act.",_fv(ratio_m,"ratio") if ratio_m and not _np.isnan(ratio_m) else "N/A"],
            ["Potencial Múltiplos", _fv(pot_m,"pct")],
        ]
        story.append(tbl(sug_rows,[3.2*inch,2.5*inch]))

        # Señal múltiplos
        smb = Table([[Paragraph(f"Señal Múltiplos: {señal_m}", sty("sm",
            fontSize=12, textColor=sc_m, fontName="Helvetica-Bold", alignment=TA_CENTER))]],
            colWidths=[7*inch])
        smb.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),C_BG),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("BOX",(0,0),(-1,-1),1.5,sc_m)]))
        story.append(Spacer(1,0.1*inch))
        story.append(smb)

        # ══════════════════════════════════════════════════════════════════════
        # CONCLUSIÓN
        # ══════════════════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph("CONCLUSIÓN Y RECOMENDACIÓN", S_H1))
        story.append(Spacer(1,0.15*inch))

        # Tabla resumen comparativa
        conc_rows = [
            ["Método","Precio Objetivo","Potencial","Señal"],
            ["FCF (WACC iterado)",     _fv(po_usd,"usd"),   _fv(pot_fcf,"pct"), señal_f],
            ["Múltiplos (convergido)", _fv(precio_cv,"usd"), _fv(pot_m,"pct"),  señal_m],
        ]
        ct = Table(conc_rows, colWidths=[2.3*inch,1.8*inch,1.4*inch,1.5*inch])
        ct.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), C_MID),
            ("TEXTCOLOR",(0,0),(-1,0), C_BLUE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("BACKGROUND",(0,1),(-1,-1), C_BG),
            ("TEXTCOLOR",(0,1),(-1,-1), C_LGRAY),
            ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("GRID",(0,0),(-1,-1),thin,GRID),
            ("ALIGN",(1,0),(-1,-1),"CENTER"),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        story.append(ct)
        story.append(Spacer(1,0.15*inch))

        # Narrativa conclusión
        pot_fcf_s = f"{(pot_fcf or 0)*100:.1f}%" if pot_fcf else "N/A"
        pot_m_s   = f"{(pot_m or 0)*100:.1f}%" if pot_m else "N/A"
        ratio_s   = f"{(ratio_m or 0)*100:.1f}%" if (ratio_m and not _np.isnan(ratio_m)) else "N/A"
        conclusion = (
            f"El análisis fundamental de <b>{tkr}</b> combina dos metodologías complementarias. "
            f"El modelo DCF (Flujo de Caja Descontado) arroja un precio objetivo de <b>{_fv(po_usd,'usd')}</b> "
            f"con un potencial de <b>{pot_fcf_s}</b>, utilizando un WACC de <b>{_fv(wacc_v,'pct')}</b>. "
            f"La valoración por múltiplos comparables del sector converge en un precio de <b>{_fv(precio_cv,'usd')}</b>, "
            f"representando el <b>{ratio_s}</b> del precio actual. "
            f"Señal consolidada: <b>{señal_m}</b>."
        )
        story.append(Paragraph(conclusion, S_BODY))
        story.append(Spacer(1,0.15*inch))

        # Señal final grande
        sfinal = Table([[Paragraph(f"SEÑAL FINAL: {señal_m}", sty("sfinal",
            fontSize=16, textColor=sc_m, fontName="Helvetica-Bold", alignment=TA_CENTER))]],
            colWidths=[7*inch])
        sfinal.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), C_BG),
            ("TOPPADDING",(0,0),(-1,-1), 12),("BOTTOMPADDING",(0,0),(-1,-1), 12),
            ("BOX",(0,0),(-1,-1), 2, sc_m),
        ]))
        story.append(sfinal)
        story.append(Spacer(1, 0.3*inch))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C_MGRAY))
        story.append(Spacer(1,0.05*inch))
        story.append(Paragraph(
            f"Generado el {today} · Datos: yfinance · Solo para fines académicos e informativos · No constituye asesoramiento de inversión.",
            S_FOOTER))

        if tkr != list(resultados_vf.keys())[-1]:
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    return buf.read()



with tab6:
    tiene_datos = "precios"   in st.session_state
    tiene_port  = "port_calc" in st.session_state
    tiene_ve    = any(k.startswith("ve_calc_") for k in st.session_state)

    st.markdown('''<style>
    .cb-label span { color: #E8F4FD !important; font-family: IBM Plex Mono, monospace !important;
                     font-size: .72rem !important; letter-spacing: .08em !important; }
    </style>''', unsafe_allow_html=True)

    st.markdown('<div class="section-header">📊 Exportar Excel</div>', unsafe_allow_html=True)

    if not tiene_datos:
        st.info("Primero carga los datos en **01 · Análisis Básico** para poder exportar.")
    else:
        st.markdown('<p style="font-family:IBM Plex Mono,monospace;font-size:.72rem;color:#E8F4FD;letter-spacing:.08em;">Selecciona las secciones a incluir en el reporte:</p>', unsafe_allow_html=True)

        # Inyectar CSS para hacer blanco el texto de los checkboxes
        st.markdown('''<style>
        [data-testid="stCheckbox"] label p,
        [data-testid="stCheckbox"] label span { color: #E8F4FD !important; }
        </style>''', unsafe_allow_html=True)

        rc1,rc2,rc3,rc4,rc5 = st.columns(5)
        with rc1:
            inc_basico = st.checkbox("📊 Análisis Básico", value=True,
                                      disabled=False, key="inc_basico")
            st.markdown('<span style="font-size:.65rem;color:#E8F4FD;">Siempre disponible</span>', unsafe_allow_html=True)
        with rc2:
            inc_port = st.checkbox("📐 Portafolio", value=tiene_port,
                                    disabled=not tiene_port, key="inc_port")
            st.markdown(f'<span style="font-size:.65rem;color:{"#2ECC71" if tiene_port else "#E74C3C"};">{"✓ Calculado" if tiene_port else "Calcular en 02"}</span>', unsafe_allow_html=True)
        with rc3:
            inc_ve = st.checkbox("📈 Val. Estadística", value=tiene_ve,
                                  disabled=not tiene_ve, key="inc_ve")
            st.markdown(f'<span style="font-size:.65rem;color:{"#2ECC71" if tiene_ve else "#E74C3C"};">{"✓ Calculado" if tiene_ve else "Calcular en 03"}</span>', unsafe_allow_html=True)
        with rc4:
            tiene_vt = any(k.startswith("vt_calc_") for k in st.session_state)
            inc_vt   = st.checkbox("📡 Val. Técnica", value=tiene_vt,
                                    disabled=not tiene_vt, key="inc_vt")
            st.markdown(f'<span style="font-size:.65rem;color:{"#2ECC71" if tiene_vt else "#E74C3C"};">{"✓ Calculado" if tiene_vt else "Calcular en 04"}</span>', unsafe_allow_html=True)
        with rc5:
            tiene_vf = "resultados_vf" in st.session_state
            inc_vf   = st.checkbox("🏦 Val. Fundamental", value=tiene_vf,
                                    disabled=not tiene_vf, key="inc_vf")
            st.markdown(f'<span style="font-size:.65rem;color:{"#2ECC71" if tiene_vf else "#E74C3C"};">{"✓ Calculado" if tiene_vf else "Calcular en 05"}</span>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("📊 Generar Excel", key="btn_excel_tab6"):
                with st.spinner("Generando Excel..."):
                    _port_data = _construir_port_data_xl() if (inc_port and tiene_port) else None
                    _tiene_ve = any(k.startswith("ve_calc_") for k in st.session_state)
                    _tiene_vt = any(k.startswith("vt_calc_") for k in st.session_state)
                    _ve_data  = _construir_ve_data_xl() if _tiene_ve else None
                    _vt_data  = _construir_vt_data_xl() if _tiene_vt else None
                    _excel    = _exportar_excel(
                        inc_basico=inc_basico, port_data=_port_data,
                        ve_data=_ve_data, vt_data=_vt_data)
                    if _excel:
                        st.session_state["excel_bytes_tab6"] = _excel
                        partes = (["Análisis Básico"] if inc_basico else [])
                        if _port_data: partes.append("Portafolio")
                        if _ve_data:   partes.append("Val. Estadística")
                        if _vt_data:   partes.append("Val. Técnica")
                        st.success(f"✅ Excel listo: {' + '.join(partes)}")

            if "excel_bytes_tab6" in st.session_state:
                st.download_button("⬇️  Descargar Excel",
                    data=st.session_state["excel_bytes_tab6"],
                    file_name=f"dashboard_financiero_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel_tab6")

        with col_btn2:
            if st.button("📊 Generar Excel — Val. Fundamental", key="btn_excel_vf",
                        disabled=not tiene_vf):
                with st.spinner("Generando Excel Fundamental..."):
                    _vf_data = _construir_fundamental_xl()
                    if _vf_data:
                        _excel_vf = _exportar_fundamental_excel(_vf_data)
                        st.session_state["excel_vf_bytes"] = _excel_vf
                        st.success("✅ Excel Fundamental listo")

            if "excel_vf_bytes" in st.session_state:
                st.download_button("⬇️  Descargar Excel Fundamental",
                    data=st.session_state["excel_vf_bytes"],
                    file_name=f"valoracion_fundamental_{datetime.today().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel_vf")

    st.markdown('<div class="section-header" style="margin-top:2rem;">📄 Exportar PDF — Valoración Fundamental</div>', unsafe_allow_html=True)
    tiene_vf_pdf = "resultados_vf" in st.session_state
    if not tiene_vf_pdf:
        st.info("Primero calcula la Valoración Fundamental en la pestaña 05.")
    else:
        if st.button("📄 Generar PDF — Reporte Fundamental", key="btn_pdf_vf"):
            with st.spinner("Generando PDF..."):
                try:
                    import importlib, subprocess, sys
                    if importlib.util.find_spec("reportlab") is None:
                        subprocess.check_call([sys.executable, "-m", "pip", "install",
                                               "reportlab", "--quiet"])
                    _pdf_bytes = _generar_pdf_fundamental(st.session_state["resultados_vf"])
                    st.session_state["pdf_vf_bytes"] = _pdf_bytes
                    st.success("✅ PDF listo")
                except Exception as e_pdf:
                    st.error(f"Error generando PDF: {e_pdf}")
        if "pdf_vf_bytes" in st.session_state:
            st.download_button("⬇️  Descargar PDF Fundamental",
                data=st.session_state["pdf_vf_bytes"],
                file_name=f"reporte_fundamental_{datetime.today().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                key="dl_pdf_vf")
